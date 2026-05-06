import unittest
from pathlib import Path

from openpyxl import Workbook

from lingxing_excel_processor import (
    WorkbookSelection,
    build_mul_box_groups,
)


class UpsCartonNumberTests(unittest.TestCase):
    def test_mul_sku_uses_carton_number_not_carton_name(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        headers = ["序号", "MSKU", "FNSKU", "品名", "SKU", "发货数量", "已装箱数", "第1箱"]
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=3, column=col_idx).value = header

        worksheet.cell(row=4, column=1).value = "1"
        worksheet.cell(row=4, column=2).value = "WT-9002BGZZ"
        worksheet.cell(row=4, column=3).value = "X004DX3PZN"
        worksheet.cell(row=4, column=5).value = "211F9002BG3"
        worksheet.cell(row=4, column=6).value = 2
        worksheet.cell(row=4, column=7).value = 2
        worksheet.cell(row=4, column=8).value = 2

        worksheet.cell(row=5, column=1).value = "2"
        worksheet.cell(row=5, column=2).value = "WT-9002CP"
        worksheet.cell(row=5, column=3).value = "X004AGMB1R"
        worksheet.cell(row=5, column=5).value = "211F9002CP1"
        worksheet.cell(row=5, column=6).value = 1
        worksheet.cell(row=5, column=7).value = 1
        worksheet.cell(row=5, column=8).value = 1

        worksheet.cell(row=11, column=5).value = "箱号"
        worksheet.cell(row=11, column=8).value = "FBA19BBYFD4FU000001"
        worksheet.cell(row=12, column=5).value = "箱子名称"
        worksheet.cell(row=12, column=8).value = "P1 - B1"

        selection = WorkbookSelection(
            path=Path("synthetic.xlsx"),
            sheet_name=worksheet.title,
            header_row=3,
            headers={header: index for index, header in enumerate(headers, start=1)},
        )
        source_rows = [
            {
                "MSKU": worksheet.cell(row=row_idx, column=2).value,
                "FNSKU": worksheet.cell(row=row_idx, column=3).value,
                "SKU": worksheet.cell(row=row_idx, column=5).value,
                "_worksheet_row": row_idx,
            }
            for row_idx in (4, 5)
        ]

        groups = build_mul_box_groups(worksheet, selection, source_rows, [(1, 8)])

        self.assertEqual(groups[0]["carton_numbers"], ["FBA19BBYFD4FU000001"])


if __name__ == "__main__":
    unittest.main()
