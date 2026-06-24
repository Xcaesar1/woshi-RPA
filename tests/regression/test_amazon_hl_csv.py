import io
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app.core.config import AMAZON_AGL_WORKFLOW_NAME
from app.core.config import EXAMPLE_MANIFESTS
from lingxing_excel_processor import (
    classify_source_workbook,
    extract_detail_rows,
    locate_msku_mapping_file,
    process_workbooks,
    read_source_metadata,
)

from app.services.amazon_hl_csv_service import (
    convert_amazon_hl_csv_to_source_workbook,
    parse_amazon_hl_csv,
    parse_amazon_hl_csv_shipments,
)
from app.services.workflow_service import create_task_submission, get_workflow_options, process_amazon_hl_csv_task, validate_hl_upload_filename


AMAZON_HL_SAMPLE = '''"工作流程名称","wffec879c4-0f2d-42f5-adba-17db47eeaa3e"
"货件编号","FBA19GPGFQ5Q"
"货件名称","FBA STA (06/23/2026 08:10)-POC1"
"配送地址","FONTANA, CA"
"箱子数量","30"
"SKU 数量","1"
"商品数量","150"

"原厂包装发货（30 箱）"
"SKU","商品名称","ASIN","FNSKU","原厂包装模板名称","状况","预处理类型","包装箱重量（磅）","箱子长度（英寸）","箱子宽度（英寸）","箱子高度（英寸）","每箱件数","箱子总数","商品总数","箱号"
"B2207-BN","Roman Bathtub Faucet Deck Mount Tub Faucet for Bathroom Widespread Elegant Classic Spout with 3 Holes 2 Handle Valve Cartridge Included, Brush Nickel","B09VGDD6DF","X0036T55NF","B2207","新品","无需进行预处理","44.22","19.69","17.91","13.58","5","30","150","FBA19GPGFQ5QU000030,FBA19GPGFQ5QU000011,FBA19GPGFQ5QU000021,FBA19GPGFQ5QU000010,FBA19GPGFQ5QU000014,FBA19GPGFQ5QU000029,FBA19GPGFQ5QU000007,FBA19GPGFQ5QU000026,FBA19GPGFQ5QU000020,FBA19GPGFQ5QU000027,FBA19GPGFQ5QU000013,FBA19GPGFQ5QU000024,FBA19GPGFQ5QU000017,FBA19GPGFQ5QU000012,FBA19GPGFQ5QU000005,FBA19GPGFQ5QU000002,FBA19GPGFQ5QU000001,FBA19GPGFQ5QU000003,FBA19GPGFQ5QU000009,FBA19GPGFQ5QU000015,FBA19GPGFQ5QU000004,FBA19GPGFQ5QU000018,FBA19GPGFQ5QU000006,FBA19GPGFQ5QU000022,FBA19GPGFQ5QU000008,FBA19GPGFQ5QU000019,FBA19GPGFQ5QU000016,FBA19GPGFQ5QU000025,FBA19GPGFQ5QU000028,FBA19GPGFQ5QU000023"
'''


AMAZON_HL_SAME_TABLE_MULTI_FBA = '''"工作流程名称","wffec879c4-0f2d-42f5-adba-17db47eeaa3e"
"货件名称","AGL batch"
"配送地址","FONTANA, CA"

"原厂包装发货（3 箱）"
"货件编号","SKU","商品名称","ASIN","FNSKU","原厂包装模板名称","状况","预处理类型","包装箱重量（磅）","箱子长度（英寸）","箱子宽度（英寸）","箱子高度（英寸）","每箱件数","箱子总数","商品总数","箱号"
"FBA19GPGFQ5Q","B2207-BN","Roman Bathtub Faucet Deck Mount Tub Faucet","B09VGDD6DF","X0036T55NF","B2207","新品","无需进行预处理","44.22","19.69","17.91","13.58","5","2","10","FBA19GPGFQ5QU000001,FBA19GPGFQ5QU000002"
"FBA19GPGFQ6R","B2207-BN","Roman Bathtub Faucet Deck Mount Tub Faucet","B09VGDD6DF","X0036T55NF","B2207","新品","无需进行预处理","44.22","19.69","17.91","13.58","5","1","5","FBA19GPGFQ6RU000001"
'''


class DummyUpload:
    def __init__(self, filename: str, content: str):
        self.filename = filename
        self.file = io.BytesIO(content.encode("utf-8-sig"))


class AmazonHlCsvTests(unittest.TestCase):
    def write_sample(self, directory: Path, name: str = "FBA19GPGFQ5Q.csv") -> Path:
        path = directory / name
        path.write_text(AMAZON_HL_SAMPLE, encoding="utf-8-sig")
        return path

    def test_agl_user_facing_names_do_not_say_hl(self) -> None:
        workflow = next(item for item in get_workflow_options() if item["name"] == AMAZON_AGL_WORKFLOW_NAME)

        self.assertEqual(workflow["label"], "AGL 发货 Amazon CSV 整理")
        self.assertIn("amazon_agl_shipment.csv", EXAMPLE_MANIFESTS)
        self.assertNotIn("amazon_hl_shipment.csv", EXAMPLE_MANIFESTS)

    def sample_for_fba(self, fba_code: str, cargo_name: str | None = None) -> str:
        return AMAZON_HL_SAMPLE.replace("FBA19GPGFQ5Q", fba_code).replace(
            "FBA STA (06/23/2026 08:10)-POC1",
            cargo_name or f"FBA STA ({fba_code})",
        )

    def test_parse_amazon_hl_csv_extracts_metadata_and_detail_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = self.write_sample(Path(tmp))

            shipment = parse_amazon_hl_csv(csv_path)

            self.assertEqual(shipment.fba_code, "FBA19GPGFQ5Q")
            self.assertEqual(shipment.cargo_name, "FBA STA (06/23/2026 08:10)-POC1")
            self.assertEqual(shipment.destination, "FONTANA, CA")
            self.assertEqual(shipment.carton_count, 30)
            self.assertEqual(shipment.sku_count, 1)
            self.assertEqual(shipment.total_quantity, 150)
            self.assertEqual(len(shipment.items), 1)
            item = shipment.items[0]
            self.assertEqual(item.msku, "B2207-BN")
            self.assertEqual(item.factory_sku, "B2207")
            self.assertEqual(item.fnsku, "X0036T55NF")
            self.assertEqual(item.quantity_per_box, 5)
            self.assertEqual(item.carton_count, 30)
            self.assertEqual(item.total_quantity, 150)
            self.assertEqual(item.box_range, "1-30")

    def test_parse_amazon_hl_csv_shipments_keeps_single_csv_compatible(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = self.write_sample(Path(tmp))

            shipments = parse_amazon_hl_csv_shipments(csv_path)

            self.assertEqual([shipment.fba_code for shipment in shipments], ["FBA19GPGFQ5Q"])
            self.assertEqual(len(shipments[0].items), 1)

    def test_parse_amazon_hl_csv_shipments_splits_repeated_metadata_blocks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_path = root / "multi_block.csv"
            csv_path.write_text(
                self.sample_for_fba("FBA19GPGFQ5Q", "FBA STA (20260623)-A")
                + "\n"
                + self.sample_for_fba("FBA19GPGFQ6R", "FBA STA (20260623)-B"),
                encoding="utf-8-sig",
            )

            shipments = parse_amazon_hl_csv_shipments(csv_path)

            self.assertEqual([shipment.fba_code for shipment in shipments], ["FBA19GPGFQ5Q", "FBA19GPGFQ6R"])
            self.assertEqual([shipment.items[0].box_range for shipment in shipments], ["1-30", "1-30"])

    def test_parse_amazon_hl_csv_shipments_groups_same_table_fba_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = Path(tmp) / "same_table.csv"
            csv_path.write_text(AMAZON_HL_SAME_TABLE_MULTI_FBA, encoding="utf-8-sig")

            shipments = parse_amazon_hl_csv_shipments(csv_path)

            self.assertEqual([shipment.fba_code for shipment in shipments], ["FBA19GPGFQ5Q", "FBA19GPGFQ6R"])
            self.assertEqual([shipment.carton_count for shipment in shipments], [2, 1])
            self.assertEqual([shipment.sku_count for shipment in shipments], [1, 1])
            self.assertEqual([shipment.total_quantity for shipment in shipments], [10, 5])
            self.assertEqual([shipment.items[0].box_range for shipment in shipments], ["1-2", "1-1"])

    def test_convert_amazon_hl_csv_creates_existing_one_sku_source_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_path = self.write_sample(root)

            workbook_path, shipment = convert_amazon_hl_csv_to_source_workbook(
                csv_path,
                root / "source",
                resource_dir=Path.cwd(),
            )

            self.assertEqual(shipment.fba_code, "FBA19GPGFQ5Q")
            self.assertTrue(workbook_path.name.startswith("FBA19GPGFQ5Q"))
            self.assertTrue(workbook_path.name.endswith("_AMAZON_AGL_NO_PIC.xlsx"))
            source_info = classify_source_workbook(workbook_path)
            self.assertIsNotNone(source_info)
            self.assertEqual(source_info.format_type, "ONE_SKU")

            metadata = read_source_metadata(source_info)
            self.assertEqual(metadata["货件单号"], "FBA19GPGFQ5Q")
            self.assertEqual(metadata["货件名称"], "FBA STA (06/23/2026 08:10)-POC1")
            self.assertEqual(metadata["配送地址"], "FONTANA, CA")

            workbook = load_workbook(workbook_path, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames[0], "Amazon AGL")
                rows = extract_detail_rows(workbook[source_info.selection.sheet_name], source_info.selection)
                self.assertEqual(len(rows), 1)
                self.assertEqual(rows[0]["序号"], 1)
                self.assertEqual(rows[0]["MSKU"], "B2207-BN")
                self.assertEqual(rows[0]["SKU"], "B2207")
                self.assertEqual(rows[0]["FNSKU"], "X0036T55NF")
                self.assertEqual(rows[0]["品名"], "22系列三件套缸边浴缸龙头拉丝封油")
                self.assertEqual(rows[0]["品线"], "缸边浴缸")
                self.assertEqual(rows[0]["发货量"], 150)
                self.assertEqual(rows[0]["单箱数量"], 5)
                self.assertEqual(rows[0]["箱数"], 30)
                self.assertEqual(rows[0]["箱号"], "1-30")
            finally:
                workbook.close()

    def test_process_amazon_hl_csv_uses_mapping_table_product_name_and_line(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_path = self.write_sample(root)
            source_dir = root / "source"
            output_dir = root / "output"
            convert_amazon_hl_csv_to_source_workbook(csv_path, source_dir, resource_dir=Path.cwd())

            report = process_workbooks(Path.cwd(), source_dir, output_dir)

            output_path = output_dir / report["output_workbook"]
            workbook = load_workbook(output_path, data_only=True)
            try:
                worksheet = workbook[workbook.sheetnames[0]]
                header_row = None
                headers = {}
                for row_idx in range(1, worksheet.max_row + 1):
                    row_headers = {
                        str(worksheet.cell(row=row_idx, column=col_idx).value).strip(): col_idx
                        for col_idx in range(1, worksheet.max_column + 1)
                        if worksheet.cell(row=row_idx, column=col_idx).value is not None
                    }
                    if "品名" in row_headers and "备注/品线" in row_headers:
                        header_row = row_idx
                        headers = row_headers
                        break

                self.assertIsNotNone(header_row)
                data_row = header_row + 1
                self.assertEqual(worksheet.cell(row=data_row, column=headers["品名"]).value, "22系列三件套缸边浴缸龙头拉丝封油")
                self.assertEqual(worksheet.cell(row=data_row, column=headers["备注/品线"]).value, "缸边浴缸")
            finally:
                workbook.close()

    def test_create_agl_submission_accepts_multiple_csv_files_in_one_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            job_paths = {
                "job_dir": root / "job",
                "input": root / "job" / "input",
                "downloads": root / "job" / "downloads",
                "output": root / "job" / "output",
                "logs": root / "job" / "logs",
                "screenshots": root / "job" / "screenshots",
                "reports": root / "job" / "reports",
            }
            for path in job_paths.values():
                path.mkdir(parents=True, exist_ok=True)
            captured: dict = {}

            def fake_create_task(**kwargs):
                captured.update(kwargs)
                return {
                    "id": kwargs["task_id"],
                    "workflow_name": kwargs["workflow_name"],
                    "original_filename": kwargs["original_filename"],
                    "submitter": kwargs["submitter"],
                    "remark": kwargs["remark"],
                    "status": "QUEUED",
                    "created_at": "2026-06-23 18:00:00",
                    "queued_at": "2026-06-23 18:00:00",
                    "started_at": None,
                    "heartbeat_at": None,
                    "finished_at": None,
                    "upload_path": kwargs["upload_path"],
                    "job_dir": kwargs["job_dir"],
                    "result_zip_path": None,
                    "result_primary_file": None,
                    "log_path": kwargs["log_path"],
                    "error_message": None,
                    "total_fba_count": kwargs["total_fba_count"],
                    "success_fba_count": 0,
                    "failed_fba_count": 0,
                }

            uploads = [
                DummyUpload("agl.csv", self.sample_for_fba("FBA19GPGFQ5Q")),
                DummyUpload("agl.csv", self.sample_for_fba("FBA19GPGFQ6R")),
            ]
            with (
                patch("app.services.workflow_service.build_job_directories", return_value=job_paths),
                patch("app.services.workflow_service.generate_task_id", return_value="task-multi-hl"),
                patch("app.services.workflow_service.create_task", side_effect=fake_create_task),
                patch("app.services.workflow_service.enqueue_task", return_value=True),
            ):
                task = create_task_submission(
                    manifest_uploads=uploads,
                    workflow_name=AMAZON_AGL_WORKFLOW_NAME,
                    submitter="测试",
                )

            self.assertEqual(captured["total_fba_count"], 2)
            self.assertEqual(captured["original_filename"], "agl.csv 等 2 个文件")
            self.assertEqual(task["task_display_id"], "FBA19GPGFQ5Q 等 2 个")
            self.assertEqual(sorted(path.name for path in job_paths["input"].glob("*.csv")), ["agl.csv", "agl_2.csv"])

    def test_create_agl_submission_rejects_duplicate_fba_codes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            job_paths = {
                "job_dir": root / "job",
                "input": root / "job" / "input",
                "downloads": root / "job" / "downloads",
                "output": root / "job" / "output",
                "logs": root / "job" / "logs",
                "screenshots": root / "job" / "screenshots",
                "reports": root / "job" / "reports",
            }
            for path in job_paths.values():
                path.mkdir(parents=True, exist_ok=True)

            uploads = [
                DummyUpload("first.csv", self.sample_for_fba("FBA19GPGFQ5Q")),
                DummyUpload("second.csv", self.sample_for_fba("FBA19GPGFQ5Q")),
            ]
            with (
                patch("app.services.workflow_service.build_job_directories", return_value=job_paths),
                patch("app.services.workflow_service.generate_task_id", return_value="task-duplicate-hl"),
                patch("app.services.workflow_service.create_task") as create_task_mock,
                patch("app.services.workflow_service.enqueue_task", return_value=True),
            ):
                with self.assertRaisesRegex(ValueError, "重复"):
                    create_task_submission(
                        manifest_uploads=uploads,
                        workflow_name=AMAZON_AGL_WORKFLOW_NAME,
                        submitter="测试",
                    )

            create_task_mock.assert_not_called()

    def test_process_amazon_hl_csv_task_processes_multiple_saved_csv_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            job_dir = root / "job"
            input_dir = job_dir / "input"
            input_dir.mkdir(parents=True, exist_ok=True)
            (input_dir / "first.csv").write_text(self.sample_for_fba("FBA19GPGFQ5Q"), encoding="utf-8-sig")
            (input_dir / "second.csv").write_text(self.sample_for_fba("FBA19GPGFQ6R"), encoding="utf-8-sig")
            (job_dir / "logs").mkdir(parents=True, exist_ok=True)

            report = process_amazon_hl_csv_task(
                {
                    "id": "task-process-multi-hl",
                    "job_dir": str(job_dir),
                    "log_path": str(job_dir / "logs" / "task.log"),
                },
                lambda _message: None,
            )

            self.assertEqual(report["status"], "success")
            self.assertEqual(report["success_count"], 2)
            self.assertEqual(report["failed_count"], 0)
            self.assertEqual(report["fba_codes"], ["FBA19GPGFQ5Q", "FBA19GPGFQ6R"])
            self.assertEqual(report["source"], "amazon_agl_csv")
            self.assertEqual(len(report["results"]), 2)
            for result in report["results"]:
                self.assertIn("amazon_agl", result["downloads_dir"])
                self.assertIn("amazon_agl", result["output_dir"])
                self.assertIn("amazon_agl_summary", result)
                self.assertNotIn("amazon_hl_summary", result)
                self.assertEqual(result["downloaded_files"][0]["warehouse_code"], "AMAZON-AGL")
                self.assertTrue(result["downloaded_files"][0]["filename"].endswith("_AMAZON_AGL_NO_PIC.xlsx"))

    def test_agl_upload_filename_accepts_only_csv(self) -> None:
        validate_hl_upload_filename("FBA19GPGFQ5Q.csv")

        for name in ["fba_manifest.txt", "fba_manifest.xlsx", "FBA19GPGFQ5Q"]:
            with self.subTest(name=name):
                with self.assertRaises(ValueError):
                    validate_hl_upload_filename(name)

    def test_msku_mapping_file_prefers_updated_product_code_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_mapping = root / "MSKU对应品线表.xlsx"
            updated_mapping = root / "MSKU对应品线表_已更新产品编码_20260508_193206.xlsx"
            old_mapping.touch()
            updated_mapping.touch()

            self.assertEqual(locate_msku_mapping_file(root), updated_mapping)


if __name__ == "__main__":
    unittest.main()
