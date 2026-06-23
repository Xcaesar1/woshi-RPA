from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from lingxing_excel_processor import build_lookup_index, find_matching_sheet, locate_msku_mapping_file, normalize_lookup_key
from lingxing_rpa_runner import sanitize_filename_part


AMAZON_HL_DETAIL_HEADERS = [
    "SKU",
    "商品名称",
    "FNSKU",
    "原厂包装模板名称",
    "每箱件数",
    "箱子总数",
    "商品总数",
    "箱号",
]

AMAZON_HL_FBA_HEADERS = ["货件编号", "FBA号", "FBA 号"]

SOURCE_HEADERS = ["序号", "MSKU", "FNSKU", "品名", "SKU", "发货量", "单箱数量", "箱数", "箱号", "品线"]

MSKU_MAPPING_HEADERS = ["MSKU", "产品名称", "品线"]


@dataclass(frozen=True)
class AmazonHlItem:
    sequence: int
    msku: str
    product_name: str
    fnsku: str
    factory_sku: str
    quantity_per_box: int
    carton_count: int
    total_quantity: int
    box_range: str
    raw_box_numbers: list[str]


@dataclass(frozen=True)
class AmazonHlShipment:
    fba_code: str
    cargo_name: str
    destination: str
    carton_count: int
    sku_count: int
    total_quantity: int
    items: list[AmazonHlItem]


@dataclass(frozen=True)
class AmazonHlProductMapping:
    product_name: str | None
    product_line: str | None


def _clean_cell(value: Any) -> str:
    return str(value or "").strip().lstrip("\ufeff")


def _read_csv_rows(path: Path) -> list[list[str]]:
    if not path.exists():
        raise FileNotFoundError(f"Amazon CSV 文件不存在：{path}")

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [[_clean_cell(cell) for cell in row] for row in csv.reader(handle)]


def _parse_int(value: str, field_name: str) -> int:
    text = _clean_cell(value).replace(",", "")
    if not text:
        raise ValueError(f"Amazon CSV 缺少 {field_name}")
    try:
        number = float(text)
    except ValueError as exc:
        raise ValueError(f"Amazon CSV 的 {field_name} 不是有效数字：{value}") from exc
    if number < 0 or number != int(number):
        raise ValueError(f"Amazon CSV 的 {field_name} 必须是非负整数：{value}")
    return int(number)


def _required_metadata(metadata: dict[str, str], key: str) -> str:
    value = _clean_cell(metadata.get(key))
    if not value:
        raise ValueError(f"Amazon CSV 缺少 {key}")
    return value


def _find_detail_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows):
        header_map = {cell: index for index, cell in enumerate(row) if cell}
        if all(header in header_map for header in AMAZON_HL_DETAIL_HEADERS):
            return row_index, header_map
    raise ValueError("Amazon CSV 未找到原厂包装发货明细表头")


def _parse_box_numbers(value: str) -> list[str]:
    return [item.strip() for item in _clean_cell(value).split(",") if item.strip()]


def _nonblank_cells(row: list[str]) -> list[str]:
    return [_clean_cell(cell) for cell in row if _clean_cell(cell)]


def _is_fba_metadata_row(row: list[str]) -> bool:
    cells = _nonblank_cells(row)
    return len(cells) == 2 and cells[0] == "货件编号" and cells[1].upper().startswith("FBA")


def _split_repeated_metadata_blocks(rows: list[list[str]]) -> list[list[list[str]]]:
    starts = [index for index, row in enumerate(rows) if _is_fba_metadata_row(row)]
    if len(starts) <= 1:
        return []

    blocks: list[list[list[str]]] = []
    for index, start in enumerate(starts):
        end = starts[index + 1] if index + 1 < len(starts) else len(rows)
        block = rows[start:end]
        if any(_nonblank_cells(row) for row in block):
            blocks.append(block)
    return blocks


def _extract_metadata(rows: list[list[str]]) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row in rows:
        if len(row) >= 2 and row[0] and row[1]:
            metadata[row[0]] = row[1]
    return metadata


def _find_existing_header(header_map: dict[str, int], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in header_map:
            return candidate
    return None


def _row_value(row: list[str], header_map: dict[str, int], header: str) -> str:
    index = header_map[header]
    return _clean_cell(row[index] if index < len(row) else "")


def _has_any_detail_value(row: list[str], header_map: dict[str, int]) -> bool:
    return any(_row_value(row, header_map, header) for header in AMAZON_HL_DETAIL_HEADERS)


def _is_non_detail_marker_row(row: list[str]) -> bool:
    first_cell = _clean_cell(row[0] if row else "")
    return first_cell in {
        "工作流程名称",
        "货件编号",
        "货件名称",
        "配送地址",
        "箱子数量",
        "SKU 数量",
        "商品数量",
    } or first_cell.startswith("原厂包装发货")


def _build_item_from_row(
    row: list[str],
    detail_headers: dict[str, int],
    *,
    sequence: int,
    box_start: int,
) -> tuple[AmazonHlItem, int]:
    row_carton_count = _parse_int(_row_value(row, detail_headers, "箱子总数"), "箱子总数")
    if row_carton_count <= 0:
        raise ValueError("Amazon CSV 的 箱子总数 必须大于 0")

    box_end = box_start + row_carton_count - 1
    item = AmazonHlItem(
        sequence=sequence,
        msku=_row_value(row, detail_headers, "SKU"),
        product_name=_row_value(row, detail_headers, "商品名称"),
        fnsku=_row_value(row, detail_headers, "FNSKU"),
        factory_sku=_row_value(row, detail_headers, "原厂包装模板名称"),
        quantity_per_box=_parse_int(_row_value(row, detail_headers, "每箱件数"), "每箱件数"),
        carton_count=row_carton_count,
        total_quantity=_parse_int(_row_value(row, detail_headers, "商品总数"), "商品总数"),
        box_range=f"{box_start}-{box_end}",
        raw_box_numbers=_parse_box_numbers(_row_value(row, detail_headers, "箱号")),
    )
    return item, box_end + 1


def _validate_items(items: list[AmazonHlItem]) -> None:
    if not items:
        raise ValueError("Amazon CSV 未找到可处理的 SKU 明细")

    for item in items:
        if not item.msku:
            raise ValueError(f"Amazon CSV 第 {item.sequence} 行缺少 SKU")
        if not item.product_name:
            raise ValueError(f"Amazon CSV 第 {item.sequence} 行缺少 商品名称")
        if not item.fnsku:
            raise ValueError(f"Amazon CSV 第 {item.sequence} 行缺少 FNSKU")
        if not item.factory_sku:
            raise ValueError(f"Amazon CSV 第 {item.sequence} 行缺少 原厂包装模板名称")


def _unique_nonblank(values: list[Any]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _clean_cell(value)
        if not text or text in seen:
            continue
        seen.add(text)
        output.append(text)
    return output


def load_msku_product_mapping(
    resource_dir: Path | None,
    msku_values: list[str] | None = None,
) -> dict[str, AmazonHlProductMapping]:
    if resource_dir is None:
        return {}

    wanted_keys = {
        normalize_lookup_key(msku)
        for msku in (msku_values or [])
        if normalize_lookup_key(msku)
    }
    mapping_path = locate_msku_mapping_file(resource_dir)
    selection = find_matching_sheet(mapping_path, MSKU_MAPPING_HEADERS)
    rows_by_msku = build_lookup_index(mapping_path, selection, "MSKU")

    mapping: dict[str, AmazonHlProductMapping] = {}
    for msku_key, rows in rows_by_msku.items():
        if wanted_keys and msku_key not in wanted_keys:
            continue
        product_names = _unique_nonblank([row.get("产品名称") for row in rows])
        product_lines = _unique_nonblank([row.get("品线") for row in rows])
        if len(product_names) > 1 or len(product_lines) > 1:
            raise ValueError(f"MSKU对应品线表.xlsx 中 MSKU 匹配到多条且产品名称或品线不一致：{rows[0].get('MSKU')}")
        mapping[msku_key] = AmazonHlProductMapping(
            product_name=product_names[0] if product_names else None,
            product_line=product_lines[0] if product_lines else None,
        )
    return mapping


def _parse_single_shipment_rows(rows: list[list[str]]) -> AmazonHlShipment:
    detail_header_index, detail_headers = _find_detail_header(rows)
    metadata = _extract_metadata(rows[:detail_header_index])

    fba_code = _required_metadata(metadata, "货件编号").upper()
    cargo_name = _required_metadata(metadata, "货件名称")
    destination = _required_metadata(metadata, "配送地址")
    carton_count = _parse_int(_required_metadata(metadata, "箱子数量"), "箱子数量")
    sku_count = _parse_int(_required_metadata(metadata, "SKU 数量"), "SKU 数量")
    total_quantity = _parse_int(_required_metadata(metadata, "商品数量"), "商品数量")

    items: list[AmazonHlItem] = []
    next_box_start = 1
    for row in rows[detail_header_index + 1 :]:
        if not any(_clean_cell(cell) for cell in row):
            continue
        if _is_non_detail_marker_row(row):
            continue
        if not _has_any_detail_value(row, detail_headers):
            continue

        item, next_box_start = _build_item_from_row(
            row,
            detail_headers,
            sequence=len(items) + 1,
            box_start=next_box_start,
        )
        items.append(item)
    _validate_items(items)

    return AmazonHlShipment(
        fba_code=fba_code,
        cargo_name=cargo_name,
        destination=destination,
        carton_count=carton_count,
        sku_count=sku_count,
        total_quantity=total_quantity,
        items=items,
    )


def _parse_same_table_shipments(rows: list[list[str]]) -> list[AmazonHlShipment]:
    detail_header_index, detail_headers = _find_detail_header(rows)
    fba_header = _find_existing_header(detail_headers, AMAZON_HL_FBA_HEADERS)
    if fba_header is None:
        raise ValueError("Amazon CSV 明细表未找到货件编号列")

    metadata = _extract_metadata(rows[:detail_header_index])
    groups: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row_index, row in enumerate(rows[detail_header_index + 1 :], start=detail_header_index + 2):
        if not any(_clean_cell(cell) for cell in row):
            continue
        if _is_non_detail_marker_row(row):
            continue
        if not _has_any_detail_value(row, detail_headers):
            continue

        fba_code = _row_value(row, detail_headers, fba_header).upper()
        if not fba_code:
            raise ValueError(f"Amazon CSV 第 {row_index} 行缺少货件编号")

        cargo_name = _row_value(row, detail_headers, "货件名称") if "货件名称" in detail_headers else _clean_cell(metadata.get("货件名称"))
        destination = _row_value(row, detail_headers, "配送地址") if "配送地址" in detail_headers else _clean_cell(metadata.get("配送地址"))
        if fba_code not in groups:
            groups[fba_code] = {
                "cargo_name": cargo_name,
                "destination": destination,
                "items": [],
                "next_box_start": 1,
            }
            order.append(fba_code)
        group = groups[fba_code]
        if cargo_name and group["cargo_name"] and cargo_name != group["cargo_name"]:
            raise ValueError(f"Amazon CSV 中 {fba_code} 出现多个货件名称")
        if destination and group["destination"] and destination != group["destination"]:
            raise ValueError(f"Amazon CSV 中 {fba_code} 出现多个配送地址")
        if cargo_name and not group["cargo_name"]:
            group["cargo_name"] = cargo_name
        if destination and not group["destination"]:
            group["destination"] = destination

        items = group["items"]
        item, next_box_start = _build_item_from_row(
            row,
            detail_headers,
            sequence=len(items) + 1,
            box_start=group["next_box_start"],
        )
        items.append(item)
        group["next_box_start"] = next_box_start

    shipments: list[AmazonHlShipment] = []
    for fba_code in order:
        group = groups[fba_code]
        cargo_name = _clean_cell(group.get("cargo_name"))
        destination = _clean_cell(group.get("destination"))
        if not cargo_name:
            raise ValueError(f"Amazon CSV 的 {fba_code} 缺少货件名称")
        if not destination:
            raise ValueError(f"Amazon CSV 的 {fba_code} 缺少配送地址")
        items = group["items"]
        _validate_items(items)
        shipments.append(
            AmazonHlShipment(
                fba_code=fba_code,
                cargo_name=cargo_name,
                destination=destination,
                carton_count=sum(item.carton_count for item in items),
                sku_count=len({item.msku for item in items if item.msku}),
                total_quantity=sum(item.total_quantity for item in items),
                items=items,
            )
        )
    if not shipments:
        raise ValueError("Amazon CSV 未找到可处理的 FBA 明细")
    return shipments


def parse_amazon_hl_csv_shipments(path: Path) -> list[AmazonHlShipment]:
    rows = _read_csv_rows(path)
    if not rows:
        raise ValueError("Amazon CSV 是空文件")

    blocks = _split_repeated_metadata_blocks(rows)
    if blocks:
        return [_parse_single_shipment_rows(block) for block in blocks]

    _, detail_headers = _find_detail_header(rows)
    if _find_existing_header(detail_headers, AMAZON_HL_FBA_HEADERS):
        return _parse_same_table_shipments(rows)
    return [_parse_single_shipment_rows(rows)]


def parse_amazon_hl_csv(path: Path) -> AmazonHlShipment:
    return parse_amazon_hl_csv_shipments(path)[0]


def convert_amazon_hl_shipment_to_source_workbook(
    shipment: AmazonHlShipment,
    output_dir: Path,
    *,
    resource_dir: Path | None = None,
) -> tuple[Path, AmazonHlShipment]:
    product_mapping = load_msku_product_mapping(resource_dir, [item.msku for item in shipment.items])
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / f"{sanitize_filename_part(shipment.fba_code)}_AMAZON_HL_NO_PIC.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Amazon HL"

    metadata_rows = [
        ("货件单号", shipment.fba_code),
        ("货件名称", shipment.cargo_name),
        ("配送地址", shipment.destination),
        ("箱子数量", shipment.carton_count),
        ("SKU数量", shipment.sku_count),
        ("商品数量", shipment.total_quantity),
    ]
    for row_index, (label, value) in enumerate(metadata_rows, start=1):
        worksheet.cell(row=row_index, column=1).value = label
        worksheet.cell(row=row_index, column=2).value = value

    header_row = len(metadata_rows) + 2
    for col_index, header in enumerate(SOURCE_HEADERS, start=1):
        worksheet.cell(row=header_row, column=col_index).value = header

    for item_index, item in enumerate(shipment.items, start=1):
        row_index = header_row + item_index
        mapped_product = product_mapping.get(normalize_lookup_key(item.msku))
        values = {
            "序号": item.sequence,
            "MSKU": item.msku,
            "FNSKU": item.fnsku,
            "品名": mapped_product.product_name if mapped_product and mapped_product.product_name else item.product_name,
            "SKU": item.factory_sku,
            "发货量": item.total_quantity,
            "单箱数量": item.quantity_per_box,
            "箱数": item.carton_count,
            "箱号": item.box_range,
            "品线": mapped_product.product_line if mapped_product else None,
        }
        for col_index, header in enumerate(SOURCE_HEADERS, start=1):
            worksheet.cell(row=row_index, column=col_index).value = values[header]

    workbook.save(workbook_path)
    workbook.close()
    return workbook_path, shipment


def convert_amazon_hl_csv_to_source_workbook(
    csv_path: Path,
    output_dir: Path,
    *,
    resource_dir: Path | None = None,
) -> tuple[Path, AmazonHlShipment]:
    shipment = parse_amazon_hl_csv(csv_path)
    return convert_amazon_hl_shipment_to_source_workbook(shipment, output_dir, resource_dir=resource_dir)
