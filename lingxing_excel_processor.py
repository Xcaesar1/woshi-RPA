from __future__ import annotations

import argparse
import json
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, InlineFont, TextBlock
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


SOURCE_REQUIRED_HEADERS = [
    "序号",
    "MSKU",
    "FNSKU",
    "品名",
    "SKU",
    "发货量",
    "单箱数量",
    "箱数",
    "箱号",
]

MUL_SKU_REQUIRED_HEADERS = [
    "序号",
    "MSKU",
    "FNSKU",
    "品名",
    "SKU",
    "发货数量",
    "已装箱数",
]

MUL_SKU_BOX_HEADER_RE = re.compile(r"^第(\d+)箱$")
MUL_OUTPUT_HEADERS = ["NO.", "MSKU型号", "工厂型号", "箱号", "套", "箱数", "套/箱", "条码", "箱号", ""]
MUL_OUTPUT_COLUMN_WIDTHS = {
    1: 6.75,
    2: 18.5,
    3: 28.25,
    4: 24.9,
    5: 8.6,
    6: 13,
    7: 13,
    8: 15.75,
    9: 28.4,
    10: 25.4,
}

TEMPLATE_REQUIRED_HEADERS = [
    "NO.",
    "SKU",
    "工厂型号",
    "品名",
    "内盒标签",
    "箱号",
    "套",
    "箱数",
    "套/箱",
    "票数",
    "FBA号",
    "备注/品线",
    "说明书/包装上-品牌",
]

MSKU_MAP_REQUIRED_HEADERS = [
    "店铺",
    "MSKU",
]

STORE_DETAIL_REQUIRED_HEADERS = [
    "店铺+站点",
    "店铺",
    "店铺简称",
]

FIELD_MAPPING = {
    "NO.": "序号",
    "SKU": "MSKU",
    "工厂型号": "SKU",
    "品名": "品名",
    "内盒标签": "FNSKU",
    "箱号": "箱号清洗后加前缀“编号 ”",
    "套": "发货量",
    "箱数": "箱数",
    "套/箱": "单箱数量",
    "票数": "店铺简称+海运第x票",
    "FBA号": "货件单号",
    "备注/品线": "品名关键词归类(浴缸/厨房/淋浴/面盆)",
    "说明书/包装上-品牌": "MSKU -> MSKU对应品线表[店铺] -> 店铺明细表[店铺]",
}

PRODUCT_LINE_KEYWORDS = {
    "浴缸": "浴缸",
    "厨房": "厨房",
    "淋浴": "淋浴",
    "面盆": "面盆",
}

BOX_RANGE_RE = re.compile(r"(\d+)\s*[～~\-－—–至]+\s*(\d+)\s*$")
BOX_SINGLE_RE = re.compile(r"(\d+)\s*$")
DATE_IN_FILENAME_RE = re.compile(r"_(\d{8})(?:[-_]|$)")
CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
THIN_SIDE = Side(style="thin", color="000000")
FULL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")
HEADER_FBA_FONT = InlineFont(rFont="宋体", b=True, sz=14, color="FF0000")
DATA_FONT = Font(name="宋体", size=12)
TITLE_FONT = Font(name="宋体", size=20)
ASCII_FONT_NAME = "Arial"
CJK_FONT_NAME = "宋体"
HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
HIGHLIGHT_FONT_COLOR = "FF0000"
DATA_ROW_HEIGHT = 30
HEADER_ROW_HEIGHT = 20
OUTPUT_COLUMN_WIDTHS = {
    1: 8,
    2: 18,
    3: 18,
    4: 32,
    5: 20,
    6: 18,
    7: 10,
    8: 10,
    9: 12,
    10: 18,
    11: 20,
    12: 16,
    13: 24,
}


@dataclass
class WorkbookSelection:
    path: Path
    sheet_name: str
    header_row: int
    headers: dict[str, int]


@dataclass
class SourceWorkbookInfo:
    path: Path
    format_type: str
    selection: WorkbookSelection
    box_columns: list[tuple[int, int]]


@dataclass
class StoreLookupResult:
    row_number: int
    msku: str | None
    store_site: str | None
    store_name: str | None
    store_short: str | None
    mapped_line: str | None
    brand_name: str | None
    ticket_value: str | None
    lookup_ok: bool


def iter_xlsx_files(base_dir: Path) -> list[Path]:
    return sorted(path for path in base_dir.glob("*.xlsx") if not path.name.startswith("~$"))


def iter_source_files(base_dir: Path) -> list[Path]:
    candidates = [
        path
        for path in iter_xlsx_files(base_dir)
        if path.name.startswith("FBA")
    ]
    return sorted(candidates, key=lambda path: (path.stat().st_ctime_ns, path.name))


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def normalize_lookup_key(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).upper()


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def convert_numeric(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)

    text = str(value).strip()
    if not text:
        return None

    try:
        number = float(text)
    except ValueError:
        return text

    return int(number) if number.is_integer() else number


def dedupe_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            output.append(item)
    return output


def find_matching_sheet(workbook_path: Path, required_headers: list[str]) -> WorkbookSelection:
    workbook = load_workbook(workbook_path)
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row_idx in range(1, min(worksheet.max_row, 20) + 1):
                row_headers: dict[str, int] = {}
                for col_idx in range(1, worksheet.max_column + 1):
                    normalized = normalize_header(worksheet.cell(row=row_idx, column=col_idx).value)
                    if normalized:
                        row_headers[normalized] = col_idx
                if all(header in row_headers for header in required_headers):
                    return WorkbookSelection(
                        path=workbook_path,
                        sheet_name=sheet_name,
                        header_row=row_idx,
                        headers=row_headers,
                    )
    finally:
        workbook.close()

    raise ValueError(f"未在 {workbook_path.name} 中找到包含以下表头的工作表：{', '.join(required_headers)}")


def classify_source_workbook(workbook_path: Path) -> SourceWorkbookInfo | None:
    workbook = load_workbook(workbook_path, data_only=False, rich_text=True)
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row_idx in range(1, min(worksheet.max_row, 20) + 1):
                row_headers: dict[str, int] = {}
                box_columns: list[tuple[int, int]] = []
                for col_idx in range(1, worksheet.max_column + 1):
                    normalized = normalize_header(worksheet.cell(row=row_idx, column=col_idx).value)
                    if not normalized:
                        continue
                    row_headers[normalized] = col_idx
                    box_match = MUL_SKU_BOX_HEADER_RE.fullmatch(normalized)
                    if box_match:
                        box_columns.append((int(box_match.group(1)), col_idx))

                selection = WorkbookSelection(
                    path=workbook_path,
                    sheet_name=sheet_name,
                    header_row=row_idx,
                    headers=row_headers,
                )
                if all(header in row_headers for header in SOURCE_REQUIRED_HEADERS):
                    return SourceWorkbookInfo(workbook_path, "ONE_SKU", selection, [])
                if all(header in row_headers for header in MUL_SKU_REQUIRED_HEADERS) and box_columns:
                    return SourceWorkbookInfo(workbook_path, "MUL_SKU", selection, sorted(box_columns))
    finally:
        workbook.close()

    return None


def locate_template_file(base_dir: Path) -> Path:
    exact_template = base_dir / "模板.xlsx"
    if exact_template.exists():
        return exact_template

    candidates = [path for path in iter_xlsx_files(base_dir) if "模板" in path.stem]
    if candidates:
        return sorted(candidates, key=lambda path: (len(path.name), path.name))[0]

    raise FileNotFoundError("未找到模板 Excel 文件。")


def locate_source_file(base_dir: Path) -> Path:
    candidates = [path for path in iter_source_files(base_dir) if "_NO_PIC" in path.name]
    if candidates:
        return candidates[0]

    candidates = iter_source_files(base_dir)
    if candidates:
        return candidates[0]

    raise FileNotFoundError("未找到领星下载的源数据 Excel 文件。")


def locate_msku_mapping_file(base_dir: Path) -> Path:
    exact_path = base_dir / "MSKU对应品线表.xlsx"
    if exact_path.exists():
        return exact_path

    candidates = [path for path in iter_xlsx_files(base_dir) if "MSKU" in path.name]
    if candidates:
        return sorted(candidates, key=lambda path: (len(path.name), path.name))[0]

    raise FileNotFoundError("未找到 MSKU 对应品线表 Excel 文件。")


def locate_store_detail_file(base_dir: Path) -> Path:
    exact_path = base_dir / "店铺明细表.xlsx"
    if exact_path.exists():
        return exact_path

    candidates = [path for path in iter_xlsx_files(base_dir) if "店铺明细表" in path.name]
    if candidates:
        return sorted(candidates, key=lambda path: (len(path.name), path.name))[0]

    raise FileNotFoundError("未找到店铺明细表 Excel 文件。")


def extract_metadata(worksheet: Worksheet, header_row: int) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    for row_idx in range(1, header_row):
        for col_idx in range(1, worksheet.max_column):
            label = normalize_header(worksheet.cell(row=row_idx, column=col_idx).value)
            if not label:
                continue
            value = worksheet.cell(row=row_idx, column=col_idx + 1).value
            if value is not None and label not in metadata:
                metadata[label] = value
    return metadata


def extract_detail_rows(worksheet: Worksheet, selection: WorkbookSelection) -> list[dict[str, Any]]:
    detail_rows: list[dict[str, Any]] = []
    relevant_headers = [header for header in SOURCE_REQUIRED_HEADERS if header in selection.headers]

    for row_idx in range(selection.header_row + 1, worksheet.max_row + 1):
        row_data = {
            header: worksheet.cell(row=row_idx, column=selection.headers[header]).value
            for header in relevant_headers
        }
        if all(is_blank(row_data[header]) for header in relevant_headers):
            continue
        detail_rows.append(row_data)

    return detail_rows


def build_lookup_index(workbook_path: Path, selection: WorkbookSelection, key_header: str) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheet = workbook[selection.sheet_name]
        index: dict[str, list[dict[str, Any]]] = {}
        headers = list(selection.headers.keys())
        key_col = selection.headers[key_header]

        for row_idx in range(selection.header_row + 1, worksheet.max_row + 1):
            key_value = worksheet.cell(row=row_idx, column=key_col).value
            normalized_key = normalize_lookup_key(key_value)
            if not normalized_key:
                continue

            row_data = {
                header: worksheet.cell(row=row_idx, column=selection.headers[header]).value
                for header in headers
            }
            row_data["_row_number"] = row_idx
            index.setdefault(normalized_key, []).append(row_data)

        return index
    finally:
        workbook.close()


def clean_box_number(raw_value: Any) -> tuple[Any, str | None]:
    if is_blank(raw_value):
        return None, "箱号为空，已保留为空值"

    text = str(raw_value).strip().rstrip("；;，,")
    range_match = BOX_RANGE_RE.search(text)
    if range_match:
        start = str(int(range_match.group(1)))
        end = str(int(range_match.group(2)))
        return f"{start}-{end}", None

    single_match = BOX_SINGLE_RE.search(text)
    if single_match:
        return str(int(single_match.group(1))), None

    return text, f"箱号无法识别，保留原值：{text}"


def format_box_display(box_value: Any) -> Any:
    if is_blank(box_value):
        return None
    return f"编号 {box_value}"


def extract_fba_number(raw_value: Any) -> tuple[Any, str | None]:
    if is_blank(raw_value):
        return None, "未找到货件单号，FBA号留空"

    text = str(raw_value).strip()
    if re.fullmatch(r"[A-Za-z0-9-]+", text):
        return text, None

    tokens = re.findall(r"[A-Za-z0-9-]+", text)
    if len(tokens) == 1:
        return tokens[0], None

    return text, f"FBA号无法判断，已保留原始货件单号：{text}"


def classify_product_line(product_name: Any) -> tuple[Any, str | None]:
    if is_blank(product_name):
        return None, "品名为空，备注/品线留空"

    text = str(product_name).strip()
    for keyword, line_name in PRODUCT_LINE_KEYWORDS.items():
        if keyword in text:
            return line_name, None

    return None, f"备注/品线无法根据品名判断：{text}"


def format_store_brand(store_name: Any) -> str | None:
    if is_blank(store_name):
        return None

    text = str(store_name).strip()

    def repl(match: re.Match[str]) -> str:
        word = match.group(0)
        return word[:1].upper() + word[1:].lower()

    return re.sub(r"[A-Za-z]+", repl, text)


def select_source_files(base_dir: Path) -> list[Path]:
    preferred = [path for path in iter_source_files(base_dir) if "_NO_PIC" in path.name]
    return preferred or iter_source_files(base_dir)


def extract_ticket_date(metadata: dict[str, Any], source_path: Path) -> tuple[str, str | None]:
    cargo_name = "" if is_blank(metadata.get("货件名称")) else str(metadata.get("货件名称")).strip()
    cargo_match = re.search(r"(20\d{6})(?!\d)", cargo_name)
    if cargo_match:
        full_date = cargo_match.group(1)
        return full_date[2:], None

    filename_match = DATE_IN_FILENAME_RE.search(source_path.name)
    if filename_match:
        full_date = filename_match.group(1)
        return full_date[2:], "货件名称中未找到日期，已回退使用源文件名日期"

    return datetime.now().strftime("%y%m%d"), "货件名称和源文件名中都未找到日期，已回退使用系统日期"


def extract_station_code(store_site: Any) -> str | None:
    if is_blank(store_site):
        return None
    text = str(store_site).strip()
    parts = [part for part in re.split(r"[-_]", text) if part]
    return parts[-1] if parts else text


def build_title_store_label(store_short: Any, store_site: Any) -> str | None:
    if is_blank(store_short):
        return None
    station_code = extract_station_code(store_site)
    if is_blank(station_code):
        return None
    return f"{str(store_short).strip()}-{station_code}"


def ensure_merge_range(worksheet: Worksheet, cell_range: str) -> None:
    merge_ranges = {str(existing) for existing in worksheet.merged_cells.ranges}
    if cell_range not in merge_ranges:
        worksheet.merge_cells(cell_range)


def sync_merged_range_borders(worksheet: Worksheet, cell_range: str) -> None:
    for merged_range in worksheet.merged_cells.ranges:
        if str(merged_range) == cell_range:
            merged_range.format()
            break


def border_side_is_thin(side: Side) -> bool:
    return side.style == "thin"


def reinforce_header_block_borders(worksheet: Worksheet, header_row: int, header_blank_row: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        merged_range = (
            f"{get_column_letter(col_idx)}{header_row}:{get_column_letter(col_idx)}{header_blank_row}"
        )
        sync_merged_range_borders(worksheet, merged_range)


def header_block_borders_are_complete(
    worksheet: Worksheet,
    header_row: int,
    header_blank_row: int,
    max_col: int,
) -> bool:
    for col_idx in range(1, max_col + 1):
        top_cell = worksheet.cell(row=header_row, column=col_idx)
        bottom_cell = worksheet.cell(row=header_blank_row, column=col_idx)
        if not (
            border_side_is_thin(top_cell.border.left)
            and border_side_is_thin(top_cell.border.right)
            and border_side_is_thin(top_cell.border.top)
            and border_side_is_thin(top_cell.border.bottom)
        ):
            return False
        if not (
            border_side_is_thin(bottom_cell.border.left)
            and border_side_is_thin(bottom_cell.border.right)
            and border_side_is_thin(bottom_cell.border.bottom)
        ):
            return False
    return True


def prepare_output_sheet(workbook, keep_sheet_name: str, output_sheet_name: str) -> Worksheet:
    for sheet_name in list(workbook.sheetnames):
        if sheet_name != keep_sheet_name:
            workbook.remove(workbook[sheet_name])
    worksheet = workbook[keep_sheet_name]
    worksheet.title = output_sheet_name
    return worksheet


def create_output_workbook(output_sheet_name: str) -> tuple[Any, Worksheet, WorkbookSelection]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = output_sheet_name
    for col_idx, width in OUTPUT_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    header_map = {header: idx for idx, header in enumerate(TEMPLATE_REQUIRED_HEADERS, start=1)}
    selection = WorkbookSelection(
        path=Path("<generated>"),
        sheet_name=output_sheet_name,
        header_row=2,
        headers=header_map,
    )
    return workbook, worksheet, selection


def build_block_positions(start_row: int, detail_row_count: int) -> dict[str, int]:
    return {
        "title_row": start_row,
        "header_row": start_row + 1,
        "header_blank_row": start_row + 2,
        "detail_start_row": start_row + 3,
        "summary_row": start_row + 3 + detail_row_count,
        "next_start_row": start_row + 4 + detail_row_count,
    }


def apply_header_block(
    worksheet: Worksheet,
    block_start_row: int,
    header_values: dict[int, Any],
    fba_number: Any,
    max_col: int,
) -> None:
    header_row = block_start_row + 1
    header_blank_row = block_start_row + 2

    if block_start_row > 1:
        worksheet.row_dimensions[block_start_row].height = worksheet.row_dimensions[1].height
        clone_row_format(worksheet, 2, header_row, max_col)
        clone_row_format(worksheet, 3, header_blank_row, max_col)
    else:
        worksheet.row_dimensions[header_row].height = HEADER_ROW_HEIGHT
        worksheet.row_dimensions[header_blank_row].height = HEADER_ROW_HEIGHT

    ensure_merge_range(worksheet, build_title_merge_range(block_start_row))
    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        merged_range = f"{column_letter}{header_row}:{column_letter}{header_blank_row}"
        ensure_merge_range(worksheet, merged_range)
        header_cell = worksheet.cell(row=header_row, column=col_idx)
        if col_idx != 6:
            header_cell.value = header_values.get(col_idx)
        header_cell.alignment = copy(HEADER_ALIGNMENT)
        header_cell.border = FULL_BORDER
        sync_merged_range_borders(worksheet, merged_range)

    box_header_ref = f"{worksheet.cell(row=header_row, column=6).column_letter}{header_row}"
    apply_box_header_style(worksheet, box_header_ref, fba_number)
    reinforce_header_block_borders(worksheet, header_row, header_blank_row, max_col)


def detect_data_start_row(worksheet: Worksheet, header_row: int) -> int:
    max_header_row = header_row
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= header_row <= merged_range.max_row:
            max_header_row = max(max_header_row, merged_range.max_row)
    return max_header_row + 1


def clone_row_format(worksheet: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel

    for col_idx in range(1, max_col + 1):
        source_cell = worksheet.cell(row=source_row, column=col_idx)
        target_cell = worksheet.cell(row=target_row, column=col_idx)
        if isinstance(source_cell, MergedCell):
            continue
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format


def clear_target_range(worksheet: Worksheet, start_row: int, end_row: int, max_col: int) -> None:
    if end_row < start_row:
        return

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, max_col + 1):
            worksheet.cell(row=row_idx, column=col_idx).value = None


def ensure_header_room(worksheet: Worksheet, header_row: int) -> None:
    next_row = header_row + 1
    current_header_height = worksheet.row_dimensions[header_row].height or 15
    current_next_height = worksheet.row_dimensions[next_row].height or 15
    worksheet.row_dimensions[header_row].height = max(current_header_height, 20)
    worksheet.row_dimensions[next_row].height = max(current_next_height, 20)


def build_title_merge_range(title_row: int) -> str:
    return f"A{title_row}:M{title_row}"


def apply_title_style(worksheet: Worksheet, title_row: int, title_text: str) -> None:
    ensure_merge_range(worksheet, build_title_merge_range(title_row))
    title_cell = worksheet.cell(row=title_row, column=1)
    title_cell.value = title_text
    title_cell.font = copy(TITLE_FONT)
    title_cell.alignment = copy(TITLE_ALIGNMENT)
    worksheet.row_dimensions[title_row].height = 30


def apply_box_header_style(worksheet: Worksheet, header_cell_ref: str, fba_number: Any) -> None:
    header_cell = worksheet[header_cell_ref]
    fba_text = "" if is_blank(fba_number) else str(fba_number).strip()
    if fba_text:
        header_cell.value = CellRichText("箱号\n", TextBlock(HEADER_FBA_FONT, fba_text))
    else:
        header_cell.value = "箱号"
    header_cell.alignment = copy(HEADER_ALIGNMENT)
    header_cell.border = FULL_BORDER
    sync_merged_range_borders(
        worksheet,
        f"{header_cell.column_letter}{header_cell.row}:{header_cell.column_letter}{header_cell.row + 1}",
    )
    ensure_header_room(worksheet, header_cell.row)


def apply_data_cell_style(cell) -> None:
    wrap_text = bool(cell.alignment.wrapText) if cell.alignment else False
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)
    cell.border = FULL_BORDER
    cell.font = Font(name=font_name_for_value(cell.value), size=12)


def apply_data_row_style(worksheet: Worksheet, row_idx: int, max_col: int) -> None:
    worksheet.row_dimensions[row_idx].height = DATA_ROW_HEIGHT
    for col_idx in range(1, max_col + 1):
        apply_data_cell_style(worksheet.cell(row=row_idx, column=col_idx))


def build_output_name(store_short: str | None, mapped_line: str | None, ticket_date: str | None) -> str:
    short_part = store_short or "UNKNOWN"
    line_part = mapped_line or "未知品线"
    date_part = ticket_date or datetime.now().strftime("%y%m%d")
    return f"{short_part}-{line_part}-货代信息表（排序表）-{date_part}.xlsx"


def save_workbook_with_fallback(workbook, preferred_path: Path) -> Path:
    if not preferred_path.exists():
        try:
            workbook.save(preferred_path)
            return preferred_path
        except PermissionError:
            pass

    for index in range(1, 100):
        fallback_path = preferred_path.with_name(f"{preferred_path.stem}_样式调整{index}{preferred_path.suffix}")
        try:
            workbook.save(fallback_path)
            return fallback_path
        except PermissionError:
            continue

    raise PermissionError(f"无法保存输出文件：{preferred_path.name}")


def resolve_store_lookup(
    row_number: int,
    msku_value: Any,
    ticket_index: int,
    msku_index: dict[str, list[dict[str, Any]]],
    store_index: dict[str, list[dict[str, Any]]],
    anomalies: list[str],
) -> StoreLookupResult:
    msku_text = None if is_blank(msku_value) else str(msku_value).strip()
    if is_blank(msku_text):
        anomalies.append(f"第 {row_number} 行：MSKU 为空，无法查店铺")
        return StoreLookupResult(row_number, msku_text, None, None, None, None, None, None, False)

    msku_matches = msku_index.get(normalize_lookup_key(msku_text), [])
    if not msku_matches:
        anomalies.append(f"第 {row_number} 行：MSKU 在 MSKU对应品线表.xlsx 中找不到：{msku_text}")
        return StoreLookupResult(row_number, msku_text, None, None, None, None, None, None, False)

    selected_msku_match = msku_matches[0]
    selected_store_match: dict[str, Any] | None = None
    if len(msku_matches) > 1:
        equivalent_candidates: list[tuple[tuple[str | None, str | None, str | None], dict[str, Any], dict[str, Any]]] = []
        for msku_match in msku_matches:
            candidate_line_raw = msku_match.get("品线")
            candidate_line = None if is_blank(candidate_line_raw) else str(candidate_line_raw).strip()
            candidate_store_site_raw = msku_match.get("店铺")
            candidate_store_site = None if is_blank(candidate_store_site_raw) else str(candidate_store_site_raw).strip()
            if is_blank(candidate_store_site):
                continue

            candidate_store_matches = store_index.get(normalize_lookup_key(candidate_store_site), [])
            if len(candidate_store_matches) != 1:
                continue

            candidate_store = candidate_store_matches[0]
            candidate_store_name_raw = candidate_store.get("店铺")
            candidate_store_short_raw = candidate_store.get("店铺简称")
            candidate_store_name = None if is_blank(candidate_store_name_raw) else str(candidate_store_name_raw).strip()
            candidate_store_short = None if is_blank(candidate_store_short_raw) else str(candidate_store_short_raw).strip()
            signature = (candidate_store_short, candidate_store_name, candidate_line)
            equivalent_candidates.append((signature, msku_match, candidate_store))

        equivalent_signatures = dedupe_preserve_order([item[0] for item in equivalent_candidates])
        if len(equivalent_signatures) == 1:
            selected_msku_match = equivalent_candidates[0][1]
            selected_store_match = equivalent_candidates[0][2]
        else:
            anomalies.append(f"第 {row_number} 行：MSKU 在 MSKU对应品线表.xlsx 中匹配到多条记录且解析结果不一致：{msku_text}")
            return StoreLookupResult(row_number, msku_text, None, None, None, None, None, None, False)

    mapped_line_raw = selected_msku_match.get("品线")
    mapped_line = None if is_blank(mapped_line_raw) else str(mapped_line_raw).strip()
    store_site_value = selected_msku_match.get("店铺")
    store_site_text = None if is_blank(store_site_value) else str(store_site_value).strip()
    if is_blank(store_site_text):
        anomalies.append(f"第 {row_number} 行：MSKU 对应的店铺字段为空：{msku_text}")
        return StoreLookupResult(row_number, msku_text, None, None, None, mapped_line, None, None, False)

    if selected_store_match is None:
        store_matches = store_index.get(normalize_lookup_key(store_site_text), [])
        if not store_matches:
            anomalies.append(f"第 {row_number} 行：店铺明细表.xlsx 中找不到对应店铺+站点：{store_site_text}")
            return StoreLookupResult(row_number, msku_text, store_site_text, None, None, mapped_line, None, None, False)

        if len(store_matches) > 1:
            anomalies.append(f"第 {row_number} 行：店铺明细表.xlsx 中店铺+站点匹配到多条记录：{store_site_text}")
            return StoreLookupResult(row_number, msku_text, store_site_text, None, None, mapped_line, None, None, False)

        selected_store_match = store_matches[0]

    store_name_raw = selected_store_match.get("店铺")
    store_short_raw = selected_store_match.get("店铺简称")
    store_name = None if is_blank(store_name_raw) else str(store_name_raw).strip()
    store_short = None if is_blank(store_short_raw) else str(store_short_raw).strip()
    brand_name = format_store_brand(store_name)
    ticket_value = None if is_blank(store_short) else f"{store_short}海运第{ticket_index}票"

    if is_blank(store_short):
        anomalies.append(f"第 {row_number} 行：店铺简称为空，票数无法填写：{store_site_text}")
    if is_blank(store_name):
        anomalies.append(f"第 {row_number} 行：店铺名称为空，说明书/包装上-品牌无法填写：{store_site_text}")
    if is_blank(mapped_line):
        anomalies.append(f"第 {row_number} 行：MSKU 对应的品线为空：{msku_text}")

    lookup_ok = not is_blank(store_short) and not is_blank(store_name) and not is_blank(mapped_line)
    return StoreLookupResult(
        row_number=row_number,
        msku=msku_text,
        store_site=store_site_text,
        store_name=store_name,
        store_short=store_short,
        mapped_line=mapped_line,
        brand_name=brand_name,
        ticket_value=ticket_value,
        lookup_ok=lookup_ok,
    )


def as_plain_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_mul_shipment_date(cargo_name: Any, source_path: Path) -> tuple[str, str, str | None]:
    text = as_plain_text(cargo_name)
    full_match = re.search(r"(20\d{6})(?!\d)", text)
    if full_match:
        full_date = full_match.group(1)
        return f"{full_date[:4]}.{full_date[4:6]}.{full_date[6:8]}", full_date[2:], None

    short_matches = re.findall(r"(?<!\d)(\d{6})(?!\d)", text)
    if short_matches:
        short_date = short_matches[-1]
        full_date = f"20{short_date}"
        return f"{full_date[:4]}.{full_date[4:6]}.{full_date[6:8]}", short_date, None

    fallback_date, anomaly = extract_ticket_date({"货件名称": cargo_name}, source_path)
    full_date = f"20{fallback_date}"
    return f"{full_date[:4]}.{full_date[4:6]}.{full_date[6:8]}", fallback_date, anomaly


def format_country_name(country_code: str) -> str:
    return {
        "US": "United States",
        "USA": "United States",
        "CA": "Canada",
        "MX": "Mexico",
    }.get(country_code.upper(), country_code)


def format_mul_warehouse_address(fc_code: Any, raw_address: Any) -> str:
    fc = as_plain_text(fc_code) or "UNKNOWN"
    parts = [part.strip() for part in as_plain_text(raw_address).split(",") if part.strip()]
    if not parts:
        return f"仓库地址：{fc}"

    start_idx = 1 if parts and (parts[0].upper() == fc.upper() or "AMAZON" in parts[0].upper()) else 0
    street = parts[start_idx] if len(parts) > start_idx else ""
    city = parts[start_idx + 1] if len(parts) > start_idx + 1 else ""
    state = parts[start_idx + 2] if len(parts) > start_idx + 2 else ""
    postal = parts[start_idx + 3] if len(parts) > start_idx + 3 else ""
    country = format_country_name(parts[start_idx + 4]) if len(parts) > start_idx + 4 else ""

    address_parts = [fc]
    if street or postal:
        address_parts.append(" ".join(part for part in [street, postal] if part))
    if city or state:
        address_parts.append(", ".join(part for part in [city, state] if part))
    if country:
        address_parts.append(country)
    return f"仓库地址：{' - '.join(address_parts)}"


def sanitize_output_filename_part(value: Any) -> str:
    text = as_plain_text(value) or "UNKNOWN"
    return re.sub(r'[\\/:*?"<>|]+', "_", text).strip(" ._") or "UNKNOWN"


def build_mul_output_name(cargo_name: Any, fba_number: Any) -> str:
    return f"装箱信息-{sanitize_output_filename_part(cargo_name)}-{sanitize_output_filename_part(fba_number)}.xlsx"


def extract_mul_detail_rows(worksheet: Worksheet, selection: WorkbookSelection) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(selection.header_row + 1, worksheet.max_row + 1):
        sequence = worksheet.cell(row=row_idx, column=selection.headers["序号"]).value
        msku = worksheet.cell(row=row_idx, column=selection.headers["MSKU"]).value
        if is_blank(sequence) and is_blank(msku):
            continue
        if is_blank(msku):
            continue
        row_data = {
            header: worksheet.cell(row=row_idx, column=selection.headers[header]).value
            for header in MUL_SKU_REQUIRED_HEADERS
            if header in selection.headers
        }
        row_data["_worksheet_row"] = row_idx
        rows.append(row_data)
    return rows


def normalize_mul_quantity(value: Any) -> int | float | None:
    number = convert_numeric(value)
    if isinstance(number, (int, float)) and not isinstance(number, bool) and number > 0:
        return number
    return None


def format_box_sequence(start_box: int, end_box: int) -> str:
    if start_box == end_box:
        return f"编号{start_box:02d}"
    return f"编号{start_box:02d}-{end_box:02d}"


def format_carton_range(box_numbers: list[str]) -> str | None:
    clean_numbers = [number for number in box_numbers if not is_blank(number)]
    if not clean_numbers:
        return None
    if len(clean_numbers) == 1:
        return clean_numbers[0]
    return f"{clean_numbers[0]}-{clean_numbers[-1]}"


def same_mul_box_signature(left: list[dict[str, Any]], right: list[dict[str, Any]]) -> bool:
    return [
        (item["msku"], item["factory_sku"], item["fnsku"], item["quantity_per_box"])
        for item in left
    ] == [
        (item["msku"], item["factory_sku"], item["fnsku"], item["quantity_per_box"])
        for item in right
    ]


def build_mul_box_groups(
    worksheet: Worksheet,
    selection: WorkbookSelection,
    source_rows: list[dict[str, Any]],
    box_columns: list[tuple[int, int]],
) -> list[dict[str, Any]]:
    carton_number_row = selection.header_row + 9
    groups: list[dict[str, Any]] = []
    current_group: dict[str, Any] | None = None
    previous_primary_msku: Any = None

    for box_number, col_idx in box_columns:
        items: list[dict[str, Any]] = []
        for source_row in source_rows:
            quantity = normalize_mul_quantity(
                worksheet.cell(
                    row=source_row["_worksheet_row"],
                    column=col_idx,
                ).value
            )
            if quantity is None:
                continue
            items.append(
                {
                    "msku": source_row.get("MSKU"),
                    "factory_sku": source_row.get("SKU"),
                    "fnsku": source_row.get("FNSKU"),
                    "quantity_per_box": quantity,
                }
            )

        if not items:
            continue

        if previous_primary_msku is not None and any(item["msku"] == previous_primary_msku for item in items):
            items = sorted(items, key=lambda item: 0 if item["msku"] == previous_primary_msku else 1)

        carton_number = as_plain_text(worksheet.cell(row=carton_number_row, column=col_idx).value)
        if current_group and same_mul_box_signature(current_group["items"], items):
            current_group["end_box"] = box_number
            current_group["carton_numbers"].append(carton_number)
        else:
            current_group = {
                "start_box": box_number,
                "end_box": box_number,
                "items": items,
                "carton_numbers": [carton_number],
            }
            groups.append(current_group)
            previous_primary_msku = items[0]["msku"]

    return groups


def apply_mul_cell_style(cell, *, bold: bool = False, size: int = 12, wrap: bool = True) -> None:
    cell.font = Font(name=font_name_for_value(cell.value), size=size, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = FULL_BORDER


def apply_mul_output_layout(worksheet: Worksheet) -> None:
    for col_idx, width in MUL_OUTPUT_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width
    worksheet.row_dimensions[1].height = 77
    worksheet.row_dimensions[2].height = 23
    worksheet.row_dimensions[3].height = 23


def merge_and_style(worksheet: Worksheet, cell_range: str) -> None:
    ensure_merge_range(worksheet, cell_range)
    sync_merged_range_borders(worksheet, cell_range)


def contains_cjk(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, CellRichText):
        text = "".join(str(part) for part in value)
    else:
        text = str(value)
    return bool(CJK_RE.search(text))


def font_name_for_value(value: Any) -> str:
    return CJK_FONT_NAME if contains_cjk(value) else ASCII_FONT_NAME


def clone_font(
    source_font: Font,
    *,
    name: str | None = None,
    size: int | float | None = None,
    color: str | None = None,
    bold: bool | None = None,
) -> Font:
    return Font(
        name=name if name is not None else source_font.name,
        sz=size if size is not None else source_font.sz,
        b=bold if bold is not None else source_font.b,
        i=source_font.i,
        u=source_font.u,
        strike=source_font.strike,
        color=color if color is not None else copy(source_font.color),
        vertAlign=source_font.vertAlign,
        charset=source_font.charset,
        family=source_font.family,
        scheme=source_font.scheme,
        outline=source_font.outline,
        shadow=source_font.shadow,
        condense=source_font.condense,
        extend=source_font.extend,
    )


def apply_content_font_rule(cell, *, size: int | float | None = None, color: str | None = None) -> None:
    cell.font = clone_font(cell.font, name=font_name_for_value(cell.value), size=size, color=color)


def normalize_workbook_fonts(workbook) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                apply_content_font_rule(cell)


def read_source_metadata(source_info: SourceWorkbookInfo) -> dict[str, Any]:
    workbook = load_workbook(source_info.path, data_only=False, rich_text=True)
    try:
        worksheet = workbook[source_info.selection.sheet_name]
        return extract_metadata(worksheet, source_info.selection.header_row)
    finally:
        workbook.close()


def source_name_contains_ups(metadata: dict[str, Any]) -> bool:
    cargo_name = as_plain_text(metadata.get("货件名称"))
    return "UPS" in cargo_name.upper()


def build_one_sku_box_groups(
    worksheet: Worksheet,
    selection: WorkbookSelection,
    source_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    current_box = 1
    for source_row in source_rows:
        carton_count = convert_numeric(source_row.get("箱数")) or 1
        if not isinstance(carton_count, (int, float)) or carton_count <= 0:
            carton_count = 1
        carton_count = int(carton_count)
        quantity_per_box = convert_numeric(source_row.get("单箱数量")) or convert_numeric(source_row.get("发货量")) or 0
        raw_carton = as_plain_text(source_row.get("箱号")).rstrip("；;，,")
        groups.append(
            {
                "start_box": current_box,
                "end_box": current_box + carton_count - 1,
                "items": [
                    {
                        "msku": source_row.get("MSKU"),
                        "factory_sku": source_row.get("SKU"),
                        "fnsku": source_row.get("FNSKU"),
                        "quantity_per_box": quantity_per_box,
                    }
                ],
                "carton_numbers": [raw_carton] if raw_carton else [],
            }
        )
        current_box += carton_count
    return groups


def extract_ups_ticket_payload(
    source_info: SourceWorkbookInfo,
    ticket_index: int,
    msku_index: dict[str, list[dict[str, Any]]],
    store_index: dict[str, list[dict[str, Any]]],
    anomalies: list[str],
) -> dict[str, Any]:
    workbook = load_workbook(source_info.path, data_only=False, rich_text=True)
    try:
        worksheet = workbook[source_info.selection.sheet_name]
        metadata = extract_metadata(worksheet, source_info.selection.header_row)
        shipment_no = metadata.get("货件单号")
        fba_number, fba_anomaly = extract_fba_number(shipment_no)
        if fba_anomaly:
            anomalies.append(f"第{ticket_index}票：{fba_anomaly}")

        cargo_name = metadata.get("货件名称")
        fc_code = metadata.get("物流中心编码")
        date_display, short_date, date_anomaly = extract_mul_shipment_date(cargo_name, source_info.path)
        if date_anomaly:
            anomalies.append(f"第{ticket_index}票：{date_anomaly}")

        if source_info.format_type == "MUL_SKU":
            source_rows = extract_mul_detail_rows(worksheet, source_info.selection)
            groups = build_mul_box_groups(worksheet, source_info.selection, source_rows, source_info.box_columns)
            quantity_header = "发货数量"
        else:
            source_rows = extract_detail_rows(worksheet, source_info.selection)
            groups = build_one_sku_box_groups(worksheet, source_info.selection, source_rows)
            quantity_header = "发货量"

        if not source_rows:
            raise ValueError(f"{source_info.path.name} 未解析到 SKU 明细行")
        if not groups:
            raise ValueError(f"{source_info.path.name} 未解析到有效箱信息")

        lookup_results = [
            resolve_store_lookup(
                row_number=index,
                msku_value=row.get("MSKU"),
                ticket_index=ticket_index,
                msku_index=msku_index,
                store_index=store_index,
                anomalies=anomalies,
            )
            for index, row in enumerate(source_rows, start=1)
        ]
        store_shorts = dedupe_preserve_order(
            [result.store_short for result in lookup_results if not is_blank(result.store_short)]
        )
        store_short = store_shorts[0] if store_shorts else "UNKNOWN"
        if len(store_shorts) > 1:
            anomalies.append(f"第{ticket_index}票：UPS 文件存在多个店铺简称：{', '.join(store_shorts)}")

        total_units = sum(
            (convert_numeric(row.get(quantity_header)) or 0)
            for row in source_rows
            if isinstance(convert_numeric(row.get(quantity_header)) or 0, (int, float))
        )
        total_cartons = sum(group["end_box"] - group["start_box"] + 1 for group in groups)

        return {
            "source_info": source_info,
            "source_workbook": source_info.path.name,
            "source_sheet": source_info.selection.sheet_name,
            "source_structure": source_info.format_type,
            "metadata": metadata,
            "shipment_number": fba_number,
            "cargo_name": cargo_name,
            "warehouse_code": fc_code,
            "date_display": date_display,
            "ticket_date": short_date,
            "groups": groups,
            "store_short": store_short,
            "store_shorts": store_shorts,
            "lookup_results": lookup_results,
            "total_units": total_units,
            "total_cartons": total_cartons,
        }
    finally:
        workbook.close()


def write_ups_packing_header(
    worksheet: Worksheet,
    title_row: int,
    ticket_index: int,
    ticket: dict[str, Any],
) -> None:
    header_row = title_row + 1
    header_blank_row = title_row + 2
    worksheet.row_dimensions[title_row].height = 77
    worksheet.row_dimensions[header_row].height = 23
    worksheet.row_dimensions[header_blank_row].height = 23

    merge_and_style(worksheet, f"A{title_row}:G{title_row}")
    worksheet.cell(row=title_row, column=1).value = (
        f"装箱单（{ticket['store_short']}）第{ticket_index}票-{ticket['shipment_number'] or ''}"
    )
    worksheet.cell(row=title_row, column=1).font = Font(name="宋体", size=18, bold=True)
    worksheet.cell(row=title_row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.cell(row=title_row, column=1).border = FULL_BORDER

    worksheet.cell(row=title_row, column=8).value = f"{ticket['date_display']}\n-美西时间"
    worksheet.cell(row=title_row, column=9).value = format_mul_warehouse_address(
        ticket["warehouse_code"],
        ticket["metadata"].get("配送地址"),
    )
    worksheet.cell(row=title_row, column=10).value = "发往美国"
    for col_idx in range(8, 11):
        apply_mul_cell_style(worksheet.cell(row=title_row, column=col_idx), size=15)

    warehouse_cell = worksheet.cell(row=title_row, column=9)
    warehouse_cell.fill = copy(HIGHLIGHT_FILL)
    warehouse_cell.font = Font(
        name=font_name_for_value(warehouse_cell.value),
        size=12,
        color=HIGHLIGHT_FONT_COLOR,
    )

    for col_idx, header in enumerate(MUL_OUTPUT_HEADERS, start=1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_mul_cell_style(cell)
        apply_mul_cell_style(worksheet.cell(row=header_blank_row, column=col_idx))
        if col_idx <= 9:
            merge_and_style(
                worksheet,
                f"{get_column_letter(col_idx)}{header_row}:{get_column_letter(col_idx)}{header_blank_row}",
            )

    fba_number = ticket["shipment_number"] or ""
    worksheet.cell(row=header_row, column=4).value = CellRichText(
        "箱号  ",
        TextBlock(InlineFont(rFont="Arial", b=True, sz=12, color="FF0000"), "FBA"),
        TextBlock(InlineFont(rFont="宋体", b=True, sz=12, color="FF0000"), f"编号\n{fba_number}"),
    )
    apply_mul_cell_style(worksheet.cell(row=header_row, column=4), bold=True)


def write_ups_packing_detail_block(
    worksheet: Worksheet,
    start_row: int,
    groups: list[dict[str, Any]],
) -> tuple[int, int, int]:
    output_row = start_row
    output_no = 1
    for group in groups:
        group_start_row = output_row
        group_box_count = group["end_box"] - group["start_box"] + 1
        group_units_per_box = sum(convert_numeric(item["quantity_per_box"]) or 0 for item in group["items"])
        carton_range = format_carton_range(group["carton_numbers"])
        for item_index, item in enumerate(group["items"]):
            worksheet.row_dimensions[output_row].height = 23 if output_row != start_row else 33
            values = {
                1: output_no,
                2: item["msku"],
                3: item["factory_sku"],
                4: format_box_sequence(group["start_box"], group["end_box"]) if item_index == 0 else None,
                5: (convert_numeric(item["quantity_per_box"]) or 0) * group_box_count,
                6: group_box_count if item_index == 0 else None,
                7: group_units_per_box if item_index == 0 else None,
                8: item["fnsku"],
                9: carton_range if item_index == 0 else None,
            }
            for col_idx in range(1, 11):
                cell = worksheet.cell(row=output_row, column=col_idx)
                cell.value = values.get(col_idx)
                apply_mul_cell_style(cell)
                if col_idx == 8 and not is_blank(cell.value):
                    cell.fill = copy(HIGHLIGHT_FILL)
                    cell.font = Font(
                        name=font_name_for_value(cell.value),
                        size=12,
                        color=HIGHLIGHT_FONT_COLOR,
                    )
            output_no += 1
            output_row += 1

        if len(group["items"]) > 1:
            for col_idx in [4, 6, 7, 9]:
                merge_and_style(
                    worksheet,
                    f"{get_column_letter(col_idx)}{group_start_row}:{get_column_letter(col_idx)}{output_row - 1}",
                )

    summary_row = output_row
    worksheet.row_dimensions[summary_row].height = 23
    worksheet.cell(row=summary_row, column=1).value = "合计"
    worksheet.cell(row=summary_row, column=5).value = f"=SUM(E{start_row}:E{summary_row - 1})"
    worksheet.cell(row=summary_row, column=6).value = f"=SUM(F{start_row}:F{summary_row - 1})"
    for col_idx in range(1, 11):
        apply_mul_cell_style(worksheet.cell(row=summary_row, column=col_idx))
    return output_row - start_row, summary_row, summary_row + 2


def process_ups_packing_workbooks(
    resource_dir: Path,
    output_dir: Path,
    source_infos: list[SourceWorkbookInfo],
) -> dict[str, Any]:
    if not source_infos:
        raise ValueError("未提供 UPS 源文件")

    msku_map_path = locate_msku_mapping_file(resource_dir)
    store_detail_path = locate_store_detail_file(resource_dir)
    msku_selection = find_matching_sheet(msku_map_path, MSKU_MAP_REQUIRED_HEADERS)
    store_selection = find_matching_sheet(store_detail_path, STORE_DETAIL_REQUIRED_HEADERS)
    msku_index = build_lookup_index(msku_map_path, msku_selection, "MSKU")
    store_index = build_lookup_index(store_detail_path, store_selection, "店铺+站点")

    anomalies: list[str] = []
    tickets = [
        extract_ups_ticket_payload(source_info, ticket_index, msku_index, store_index, anomalies)
        for ticket_index, source_info in enumerate(source_infos, start=1)
    ]

    output_workbook = Workbook()
    try:
        worksheet = output_workbook.active
        apply_mul_output_layout(worksheet)

        total_units = sum(ticket["total_units"] for ticket in tickets)
        total_cartons = sum(ticket["total_cartons"] for ticket in tickets)
        total_units_display = int(total_units) if isinstance(total_units, (int, float)) and float(total_units).is_integer() else total_units
        worksheet.title = f"共{len(tickets)}票{total_units_display}件"

        block_reports: list[dict[str, Any]] = []
        current_title_row = 1
        for ticket_index, ticket in enumerate(tickets, start=1):
            write_ups_packing_header(worksheet, current_title_row, ticket_index, ticket)
            detail_count, summary_row, next_title_row = write_ups_packing_detail_block(
                worksheet,
                current_title_row + 3,
                ticket["groups"],
            )
            block_reports.append(
                {
                    "ticket_index": ticket_index,
                    "source_workbook": ticket["source_workbook"],
                    "source_structure": ticket["source_structure"],
                    "source_sheet": ticket["source_sheet"],
                    "shipment_number": ticket["shipment_number"],
                    "cargo_name": ticket["cargo_name"],
                    "warehouse_code": ticket["warehouse_code"],
                    "ticket_date": ticket["ticket_date"],
                    "title_row": current_title_row,
                    "detail_row_count": detail_count,
                    "summary_row": summary_row,
                    "total_units": ticket["total_units"],
                    "total_cartons": ticket["total_cartons"],
                    "resolved_store_shorts": ticket["store_shorts"],
                }
            )
            current_title_row = next_title_row

        normalize_workbook_fonts(output_workbook)
        first_ticket = tickets[0]
        output_path = save_workbook_with_fallback(
            output_workbook,
            output_dir / build_mul_output_name(first_ticket["cargo_name"], first_ticket["shipment_number"]),
        )

        report = {
            "format_type": "UPS_PACKING",
            "source_output_type": "UPS_PACKING",
            "source_workbooks": [ticket["source_workbook"] for ticket in tickets],
            "source_structures": [
                {"source_workbook": ticket["source_workbook"], "source_structure": ticket["source_structure"]}
                for ticket in tickets
            ],
            "ticket_count": len(tickets),
            "total_units": total_units,
            "total_cartons": total_cartons,
            "block_reports": block_reports,
            "styling_updates": {
                "warehouse_address": "黄色底红字12号",
                "barcode_column": "黄色底红字",
                "content_font_rule": "含中文宋体；不含中文 Arial",
            },
            "anomalies": dedupe_preserve_order(anomalies),
            "output_workbook": output_path.name,
        }
        report_path = output_path.with_name(f"{output_path.stem}_report.json")
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        report["report_file"] = report_path.name
        return report
    finally:
        output_workbook.close()


def font_matches_content_rule(cell, *, expected_size: int | None = None) -> bool:
    size_ok = True
    if expected_size is not None:
        size_ok = int(round((cell.font.sz or 0))) == expected_size
    return cell.font.name == font_name_for_value(cell.value) and size_ok


def process_mul_sku_workbook(
    resource_dir: Path,
    output_dir: Path,
    source_info: SourceWorkbookInfo,
    msku_index: dict[str, list[dict[str, Any]]],
    store_index: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    source_path = source_info.path
    anomalies: list[str] = []
    workbook = load_workbook(source_path, data_only=False, rich_text=True)
    output_workbook = Workbook()
    try:
        source_sheet = workbook[source_info.selection.sheet_name]
        metadata = extract_metadata(source_sheet, source_info.selection.header_row)
        shipment_no = metadata.get("货件单号")
        fba_number, fba_anomaly = extract_fba_number(shipment_no)
        if fba_anomaly:
            anomalies.append(fba_anomaly)

        cargo_name = metadata.get("货件名称")
        fc_code = metadata.get("物流中心编码")
        date_display, short_date, date_anomaly = extract_mul_shipment_date(cargo_name, source_path)
        if date_anomaly:
            anomalies.append(date_anomaly)

        source_rows = extract_mul_detail_rows(source_sheet, source_info.selection)
        if not source_rows:
            raise ValueError(f"{source_path.name} 未解析到混装 SKU 明细行")

        lookup_results = [
            resolve_store_lookup(
                row_number=index,
                msku_value=row.get("MSKU"),
                ticket_index=1,
                msku_index=msku_index,
                store_index=store_index,
                anomalies=anomalies,
            )
            for index, row in enumerate(source_rows, start=1)
        ]
        store_shorts = dedupe_preserve_order(
            [result.store_short for result in lookup_results if not is_blank(result.store_short)]
        )
        store_short = store_shorts[0] if store_shorts else "UNKNOWN"
        if len(store_shorts) > 1:
            anomalies.append(f"混装文件存在多个店铺简称：{', '.join(store_shorts)}")

        groups = build_mul_box_groups(source_sheet, source_info.selection, source_rows, source_info.box_columns)
        if not groups:
            raise ValueError(f"{source_path.name} 未解析到有效箱列")

        total_units = sum(
            (convert_numeric(row.get("发货数量")) or 0)
            for row in source_rows
            if isinstance(convert_numeric(row.get("发货数量")) or 0, (int, float))
        )
        total_cartons = sum(group["end_box"] - group["start_box"] + 1 for group in groups)

        worksheet = output_workbook.active
        worksheet.title = f"共1票{int(total_units) if float(total_units).is_integer() else total_units}件"
        apply_mul_output_layout(worksheet)

        merge_and_style(worksheet, "A1:G1")
        worksheet["A1"] = f"装箱单（{store_short}）第1票-{fba_number or ''}"
        worksheet["A1"].font = Font(name="宋体", size=18, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet["A1"].border = FULL_BORDER
        worksheet["H1"] = f"{date_display}\n-美西时间"
        worksheet["I1"] = format_mul_warehouse_address(fc_code, metadata.get("配送地址"))
        worksheet["J1"] = "发往美国"
        for col_idx in range(8, 11):
            apply_mul_cell_style(worksheet.cell(row=1, column=col_idx), size=15)
        warehouse_cell = worksheet["I1"]
        warehouse_cell.fill = copy(HIGHLIGHT_FILL)
        warehouse_cell.font = Font(
            name=font_name_for_value(warehouse_cell.value),
            size=12,
            color=HIGHLIGHT_FONT_COLOR,
        )

        for col_idx, header in enumerate(MUL_OUTPUT_HEADERS, start=1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.value = header
            apply_mul_cell_style(cell)
            apply_mul_cell_style(worksheet.cell(row=3, column=col_idx))
            if col_idx <= 9:
                merge_and_style(worksheet, f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}3")

        worksheet["D2"] = CellRichText(
            "箱号  ",
            TextBlock(InlineFont(rFont="Arial", b=True, sz=12, color="FF0000"), "FBA"),
            TextBlock(InlineFont(rFont="宋体", b=True, sz=12, color="FF0000"), f"编号\n{fba_number or ''}"),
        )
        apply_mul_cell_style(worksheet["D2"], bold=True)

        output_row = 4
        output_no = 1
        for group in groups:
            group_start_row = output_row
            group_box_count = group["end_box"] - group["start_box"] + 1
            group_units_per_box = sum(convert_numeric(item["quantity_per_box"]) or 0 for item in group["items"])
            carton_range = format_carton_range(group["carton_numbers"])
            for item_index, item in enumerate(group["items"]):
                worksheet.row_dimensions[output_row].height = 23 if output_row != 4 else 33
                values = {
                    1: output_no,
                    2: item["msku"],
                    3: item["factory_sku"],
                    4: format_box_sequence(group["start_box"], group["end_box"]) if item_index == 0 else None,
                    5: (convert_numeric(item["quantity_per_box"]) or 0) * group_box_count,
                    6: group_box_count if item_index == 0 else None,
                    7: group_units_per_box if item_index == 0 else None,
                    8: item["fnsku"],
                    9: carton_range if item_index == 0 else None,
                }
                for col_idx in range(1, 11):
                    cell = worksheet.cell(row=output_row, column=col_idx)
                    cell.value = values.get(col_idx)
                    apply_mul_cell_style(cell)
                    if col_idx == 8 and not is_blank(cell.value):
                        cell.fill = copy(HIGHLIGHT_FILL)
                        cell.font = Font(
                            name=font_name_for_value(cell.value),
                            size=12,
                            color=HIGHLIGHT_FONT_COLOR,
                        )
                output_no += 1
                output_row += 1

            if len(group["items"]) > 1:
                for col_idx in [4, 6, 7, 9]:
                    merge_and_style(
                        worksheet,
                        f"{get_column_letter(col_idx)}{group_start_row}:{get_column_letter(col_idx)}{output_row - 1}",
                    )

        summary_row = output_row
        worksheet.row_dimensions[summary_row].height = 23
        worksheet.cell(row=summary_row, column=1).value = "合计"
        worksheet.cell(row=summary_row, column=5).value = f"=SUM(E4:E{summary_row - 1})"
        worksheet.cell(row=summary_row, column=6).value = f"=SUM(F4:F{summary_row - 1})"
        for col_idx in range(1, 11):
            apply_mul_cell_style(worksheet.cell(row=summary_row, column=col_idx))

        normalize_workbook_fonts(output_workbook)
        output_path = save_workbook_with_fallback(output_workbook, output_dir / build_mul_output_name(cargo_name, fba_number))
        report = {
            "format_type": "MUL_SKU",
            "source_workbook": source_path.name,
            "source_sheet": source_info.selection.sheet_name,
            "shipment_number": fba_number,
            "cargo_name": cargo_name,
            "warehouse_code": fc_code,
            "ticket_date": short_date,
            "total_units": total_units,
            "total_cartons": total_cartons,
            "box_group_count": len(groups),
            "store_lookup_summary": {
                "resolved_store_shorts": store_shorts,
                "rows": [
                    {
                        "row_number": index,
                        "msku": result.msku,
                        "store_site": result.store_site,
                        "store_short": result.store_short,
                        "store_name": result.store_name,
                        "mapped_line": result.mapped_line,
                    }
                    for index, result in enumerate(lookup_results, start=1)
                ],
            },
            "anomalies": dedupe_preserve_order(anomalies),
            "output_workbook": output_path.name,
        }
        report_path = output_path.with_name(f"{output_path.stem}_report.json")
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        report["report_file"] = report_path.name
        return report
    finally:
        workbook.close()
        output_workbook.close()


def process_one_sku_workbooks(
    resource_dir: Path,
    source_dir: Path,
    output_dir: Path,
    source_files: list[Path] | None = None,
) -> dict[str, Any]:
    source_files = source_files or select_source_files(source_dir)
    if not source_files:
        raise FileNotFoundError("未找到可处理的领星源数据 Excel 文件。")

    msku_map_path = locate_msku_mapping_file(resource_dir)
    store_detail_path = locate_store_detail_file(resource_dir)

    msku_selection = find_matching_sheet(msku_map_path, MSKU_MAP_REQUIRED_HEADERS)
    store_selection = find_matching_sheet(store_detail_path, STORE_DETAIL_REQUIRED_HEADERS)

    msku_index = build_lookup_index(msku_map_path, msku_selection, "MSKU")
    store_index = build_lookup_index(store_detail_path, store_selection, "店铺+站点")

    output_sheet_name = f"{len(source_files)}票"
    template_workbook, template_sheet, template_selection = create_output_workbook(output_sheet_name)

    anomalies: list[str] = []
    try:
        target_max_col = len(TEMPLATE_REQUIRED_HEADERS)
        output_sheet_name = template_selection.sheet_name
        header_values = {
            col_idx: TEMPLATE_REQUIRED_HEADERS[col_idx - 1]
            for col_idx in range(1, target_max_col + 1)
        }

        block_reports: list[dict[str, Any]] = []
        lookup_rows_for_report: list[dict[str, Any]] = []
        all_store_sites: list[str] = []
        all_store_shorts: list[str] = []
        all_store_names: list[str] = []
        all_mapped_lines: list[str] = []
        ticket_date_for_filename: str | None = None
        filename_store_short: str | None = None
        filename_mapped_line: str | None = None
        current_start_row = 1

        for ticket_index, source_path in enumerate(source_files, start=1):
            source_selection = find_matching_sheet(source_path, SOURCE_REQUIRED_HEADERS)
            source_workbook = load_workbook(source_path, data_only=False)
            try:
                source_sheet = source_workbook[source_selection.sheet_name]
                metadata = extract_metadata(source_sheet, source_selection.header_row)
                shipment_no = metadata.get("货件单号")
                fba_number, fba_anomaly = extract_fba_number(shipment_no)
                if fba_anomaly:
                    anomalies.append(f"第{ticket_index}票：{fba_anomaly}")

                ticket_date, ticket_date_anomaly = extract_ticket_date(metadata, source_path)
                if ticket_index == 1:
                    ticket_date_for_filename = ticket_date
                if ticket_date_anomaly:
                    anomalies.append(f"第{ticket_index}票：{ticket_date_anomaly}")

                source_rows = extract_detail_rows(source_sheet, source_selection)
                positions = build_block_positions(current_start_row, len(source_rows))
                clear_target_range(
                    template_sheet,
                    positions["detail_start_row"],
                    max(template_sheet.max_row, positions["summary_row"]),
                    target_max_col,
                )
                apply_header_block(template_sheet, current_start_row, header_values, fba_number, target_max_col)

                lookup_results: list[StoreLookupResult] = []
                expected_set_total = sum(convert_numeric(row.get("发货量")) or 0 for row in source_rows)
                expected_carton_total = sum(convert_numeric(row.get("箱数")) or 0 for row in source_rows)

                for offset, source_row in enumerate(source_rows):
                    target_row = positions["detail_start_row"] + offset
                    if offset > 0:
                        clone_row_format(template_sheet, positions["detail_start_row"], target_row, target_max_col)

                    box_number, box_anomaly = clean_box_number(source_row.get("箱号"))
                    if box_anomaly:
                        anomalies.append(f"第{ticket_index}票第 {offset + 1} 行：{box_anomaly}")

                    product_line, product_anomaly = classify_product_line(source_row.get("品名"))
                    if product_anomaly:
                        anomalies.append(f"第{ticket_index}票第 {offset + 1} 行：{product_anomaly}")

                    lookup_result = resolve_store_lookup(
                        row_number=offset + 1,
                        msku_value=source_row.get("MSKU"),
                        ticket_index=ticket_index,
                        msku_index=msku_index,
                        store_index=store_index,
                        anomalies=anomalies,
                    )
                    lookup_results.append(lookup_result)

                    row_payload = {
                        "NO.": convert_numeric(source_row.get("序号")),
                        "SKU": source_row.get("MSKU"),
                        "工厂型号": source_row.get("SKU"),
                        "品名": source_row.get("品名"),
                        "内盒标签": source_row.get("FNSKU"),
                        "箱号": format_box_display(box_number),
                        "套": convert_numeric(source_row.get("发货量")),
                        "箱数": convert_numeric(source_row.get("箱数")),
                        "套/箱": convert_numeric(source_row.get("单箱数量")),
                        "票数": lookup_result.ticket_value,
                        "FBA号": fba_number,
                        "备注/品线": product_line,
                        "说明书/包装上-品牌": lookup_result.brand_name,
                    }

                    for header, value in row_payload.items():
                        col_idx = template_selection.headers[header]
                        template_sheet.cell(row=target_row, column=col_idx).value = value

                    apply_data_row_style(template_sheet, target_row, target_max_col)

                    lookup_rows_for_report.append(
                        {
                            "ticket_index": ticket_index,
                            "source_file": source_path.name,
                            "row_number": offset + 1,
                            "msku": lookup_result.msku,
                            "store_site": lookup_result.store_site,
                            "store_short": lookup_result.store_short,
                            "store_name": lookup_result.store_name,
                            "mapped_line": lookup_result.mapped_line,
                            "ticket_value": lookup_result.ticket_value,
                            "brand_value": lookup_result.brand_name,
                        }
                    )

                if source_rows:
                    clone_row_format(template_sheet, positions["detail_start_row"], positions["summary_row"], target_max_col)
                template_sheet.cell(row=positions["summary_row"], column=template_selection.headers["套"]).value = expected_set_total
                template_sheet.cell(row=positions["summary_row"], column=template_selection.headers["箱数"]).value = expected_carton_total
                apply_data_row_style(template_sheet, positions["summary_row"], target_max_col)

                resolved_store_sites = dedupe_preserve_order(
                    [result.store_site for result in lookup_results if not is_blank(result.store_site)]
                )
                resolved_store_shorts = dedupe_preserve_order(
                    [result.store_short for result in lookup_results if not is_blank(result.store_short)]
                )
                resolved_store_names = dedupe_preserve_order(
                    [result.store_name for result in lookup_results if not is_blank(result.store_name)]
                )
                resolved_mapped_lines = dedupe_preserve_order(
                    [result.mapped_line for result in lookup_results if not is_blank(result.mapped_line)]
                )

                all_store_sites.extend(resolved_store_sites)
                all_store_shorts.extend(resolved_store_shorts)
                all_store_names.extend(resolved_store_names)
                all_mapped_lines.extend(resolved_mapped_lines)

                if len(resolved_store_shorts) > 1:
                    anomalies.append(f"第{ticket_index}票：同一仓库文件存在多个店铺简称：{', '.join(resolved_store_shorts)}")
                if len(resolved_store_sites) > 1:
                    anomalies.append(f"第{ticket_index}票：同一仓库文件存在多个店铺+站点：{', '.join(resolved_store_sites)}")
                if len(resolved_mapped_lines) > 1:
                    anomalies.append(f"第{ticket_index}票：同一仓库文件存在多个品线：{', '.join(resolved_mapped_lines)}")

                title_label = None
                if lookup_results:
                    title_label = build_title_store_label(lookup_results[0].store_short, lookup_results[0].store_site)
                if is_blank(title_label) and resolved_store_shorts and resolved_store_sites:
                    title_label = build_title_store_label(resolved_store_shorts[0], resolved_store_sites[0])
                    anomalies.append(f"第{ticket_index}票：首行SKU未能成功解析标题店铺信息，标题改用首个可用店铺简称和站点")
                if is_blank(title_label):
                    title_label = "待确认"
                    anomalies.append(f"第{ticket_index}票：无法确定标题中的店铺简称和站点")

                title_text = f"装箱单（{title_label}）海运第{ticket_index}票"
                apply_title_style(template_sheet, positions["title_row"], title_text)

                if ticket_index == 1:
                    filename_store_short = resolved_store_shorts[0] if resolved_store_shorts else None
                    filename_mapped_line = resolved_mapped_lines[0] if resolved_mapped_lines else None

                block_reports.append(
                    {
                        "ticket_index": ticket_index,
                        "source_file": source_path.name,
                        "source_sheet": source_selection.sheet_name,
                        "source_header_row": source_selection.header_row,
                        "shipment_number": shipment_no,
                        "fba_number": fba_number,
                        "ticket_date": ticket_date,
                        "title_text": title_text,
                        "positions": positions,
                        "detail_row_count": len(source_rows),
                        "expected_set_total": expected_set_total,
                        "expected_carton_total": expected_carton_total,
                        "resolved_store_sites": resolved_store_sites,
                        "resolved_store_shorts": resolved_store_shorts,
                        "resolved_store_names": resolved_store_names,
                        "resolved_mapped_lines": resolved_mapped_lines,
                    }
                )
                current_start_row = positions["next_start_row"]
            finally:
                source_workbook.close()

        deduped_store_shorts = dedupe_preserve_order(all_store_shorts)
        deduped_mapped_lines = dedupe_preserve_order(all_mapped_lines)
        if len(deduped_store_shorts) > 1:
            anomalies.append(f"整个输出文件存在多个店铺简称：{', '.join(deduped_store_shorts)}")
        if len(deduped_mapped_lines) > 1:
            anomalies.append(f"整个输出文件存在多个品线：{', '.join(deduped_mapped_lines)}")

        preferred_output_path = output_dir / build_output_name(filename_store_short, filename_mapped_line, ticket_date_for_filename)
        normalize_workbook_fonts(template_workbook)
        output_path = save_workbook_with_fallback(template_workbook, preferred_output_path)

        written_workbook = load_workbook(output_path, data_only=False, rich_text=True)
        try:
            written_sheet = written_workbook[written_workbook.sheetnames[0]]
            actual_sheet_title = written_sheet.title
            actual_sheet_count = len(written_workbook.sheetnames)
            total_expected_rows = sum(block["detail_row_count"] for block in block_reports)
            total_actual_rows = 0
            total_expected_set = sum(block["expected_set_total"] for block in block_reports)
            total_actual_set = 0
            total_expected_carton = sum(block["expected_carton_total"] for block in block_reports)
            total_actual_carton = 0
            block_title_ok = True
            block_merge_ok = True
            block_title_style_ok = True
            block_box_header_ok = True
            block_summary_ok = True
            block_header_border_ok = True
            ticket_values_filled = True
            brand_values_filled = True
            box_numbers_prefixed = True
            added_cells_centered = True
            added_cells_bordered = True
            added_cells_font_ok = True
            added_rows_height_ok = True

            for block in block_reports:
                positions = block["positions"]
                detail_rows = list(range(positions["detail_start_row"], positions["detail_start_row"] + block["detail_row_count"]))
                styled_rows = detail_rows + [positions["summary_row"]]
                total_actual_rows += sum(
                    1
                    for row_idx in detail_rows
                    if any(not is_blank(written_sheet.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, target_max_col + 1))
                )
                block_set_total = sum(
                    convert_numeric(written_sheet.cell(row=row_idx, column=template_selection.headers["套"]).value) or 0
                    for row_idx in detail_rows
                )
                block_carton_total = sum(
                    convert_numeric(written_sheet.cell(row=row_idx, column=template_selection.headers["箱数"]).value) or 0
                    for row_idx in detail_rows
                )
                total_actual_set += block_set_total
                total_actual_carton += block_carton_total
                summary_set_total = convert_numeric(
                    written_sheet.cell(row=positions["summary_row"], column=template_selection.headers["套"]).value
                )
                summary_carton_total = convert_numeric(
                    written_sheet.cell(row=positions["summary_row"], column=template_selection.headers["箱数"]).value
                )
                block_summary_ok = (
                    block_summary_ok
                    and block["expected_set_total"] == summary_set_total
                    and block["expected_carton_total"] == summary_carton_total
                )

                actual_title = written_sheet.cell(row=positions["title_row"], column=1).value
                block_title_ok = block_title_ok and str(actual_title) == block["title_text"]
                block_merge_ok = block_merge_ok and build_title_merge_range(positions["title_row"]) in {
                    str(cell_range) for cell_range in written_sheet.merged_cells.ranges
                }
                block_title_style_ok = (
                    block_title_style_ok
                    and written_sheet.cell(row=positions["title_row"], column=1).font.name == "宋体"
                    and int(round(written_sheet.cell(row=positions["title_row"], column=1).font.sz or 0)) == 20
                    and written_sheet.cell(row=positions["title_row"], column=1).alignment.horizontal == "center"
                    and written_sheet.cell(row=positions["title_row"], column=1).alignment.vertical == "center"
                )

                box_header_value = written_sheet.cell(row=positions["header_row"], column=template_selection.headers["箱号"]).value
                block_box_header_ok = block_box_header_ok and (
                    is_blank(block["fba_number"]) or str(block["fba_number"]) in str(box_header_value)
                )
                block_header_border_ok = block_header_border_ok and header_block_borders_are_complete(
                    written_sheet,
                    positions["header_row"],
                    positions["header_blank_row"],
                    target_max_col,
                )

                ticket_values_filled = ticket_values_filled and all(
                    re.fullmatch(r".+海运第\d+票", str(written_sheet.cell(row=row_idx, column=template_selection.headers["票数"]).value or ""))
                    for row_idx in detail_rows
                )
                brand_values_filled = brand_values_filled and all(
                    not is_blank(written_sheet.cell(row=row_idx, column=template_selection.headers["说明书/包装上-品牌"]).value)
                    for row_idx in detail_rows
                )
                box_numbers_prefixed = box_numbers_prefixed and all(
                    re.fullmatch(r"编号 \d+(?:-\d+)?", str(written_sheet.cell(row=row_idx, column=template_selection.headers["箱号"]).value or ""))
                    for row_idx in detail_rows
                )
                added_cells_centered = added_cells_centered and all(
                    written_sheet.cell(row=row_idx, column=col_idx).alignment.horizontal == "center"
                    and written_sheet.cell(row=row_idx, column=col_idx).alignment.vertical == "center"
                    for row_idx in styled_rows
                    for col_idx in range(1, target_max_col + 1)
                )
                added_cells_bordered = added_cells_bordered and all(
                    written_sheet.cell(row=row_idx, column=col_idx).border.left.style == "thin"
                    and written_sheet.cell(row=row_idx, column=col_idx).border.right.style == "thin"
                    and written_sheet.cell(row=row_idx, column=col_idx).border.top.style == "thin"
                    and written_sheet.cell(row=row_idx, column=col_idx).border.bottom.style == "thin"
                    for row_idx in styled_rows
                    for col_idx in range(1, target_max_col + 1)
                )
                added_cells_font_ok = added_cells_font_ok and all(
                    font_matches_content_rule(written_sheet.cell(row=row_idx, column=col_idx), expected_size=12)
                    for row_idx in styled_rows
                    for col_idx in range(1, target_max_col + 1)
                )
                added_rows_height_ok = added_rows_height_ok and all(
                    int(round(written_sheet.row_dimensions[row_idx].height or 0)) == DATA_ROW_HEIGHT
                    for row_idx in styled_rows
                )
        finally:
            written_workbook.close()

        validations = {
            "sheet_name": {
                "expected": output_sheet_name,
                "actual": actual_sheet_title,
                "match": actual_sheet_title == output_sheet_name,
            },
            "sheet_count_is_one": actual_sheet_count == 1,
            "data_row_count": {
                "expected": total_expected_rows,
                "actual": total_actual_rows,
                "match": total_expected_rows == total_actual_rows,
            },
            "set_total": {
                "expected": total_expected_set,
                "actual": total_actual_set,
                "match": total_expected_set == total_actual_set,
            },
            "carton_total": {
                "expected": total_expected_carton,
                "actual": total_actual_carton,
                "match": total_expected_carton == total_actual_carton,
            },
            "summary_row_totals": block_summary_ok,
            "box_numbers_prefixed": box_numbers_prefixed,
            "ticket_values_filled": ticket_values_filled,
            "brand_values_filled": brand_values_filled,
            "all_titles_match": block_title_ok,
            "all_title_merges_ok": block_merge_ok,
            "all_title_styles_ok": block_title_style_ok,
            "all_box_headers_contain_fba": block_box_header_ok,
            "all_header_borders_complete": block_header_border_ok,
            "added_cells_centered": added_cells_centered,
            "added_cells_bordered": added_cells_bordered,
            "added_cells_content_font_rule_12": added_cells_font_ok,
            "added_rows_height_30": added_rows_height_ok,
        }

        if not validations["sheet_name"]["match"]:
            anomalies.append("工作表名称未按总票数命名")
        if not validations["sheet_count_is_one"]:
            anomalies.append("输出工作簿不是单工作表")
        if not validations["data_row_count"]["match"]:
            anomalies.append("数据行数校验不一致")
        if not validations["set_total"]["match"]:
            anomalies.append("套合计校验不一致")
        if not validations["carton_total"]["match"]:
            anomalies.append("箱数合计校验不一致")
        if not validations["summary_row_totals"]:
            anomalies.append("汇总行的套或箱数与明细合计不一致")
        if not validations["box_numbers_prefixed"]:
            anomalies.append("存在未按要求添加“编号”前缀的箱号")
        if not validations["ticket_values_filled"]:
            anomalies.append("票数未全部按“店铺简称+海运第x票”填写")
        if not validations["brand_values_filled"]:
            anomalies.append("说明书/包装上-品牌未全部根据SKU查表填写")
        if not validations["all_titles_match"]:
            anomalies.append("存在票块标题内容未按规则生成")
        if not validations["all_title_merges_ok"]:
            anomalies.append("存在票块标题未保持 A:M 合并")
        if not validations["all_title_styles_ok"]:
            anomalies.append("存在票块标题样式不是宋体20居中")
        if not validations["all_box_headers_contain_fba"]:
            anomalies.append("存在票块箱号表头未写入对应FBA号")
        if not validations["all_header_borders_complete"]:
            anomalies.append("存在票块表头边框不完整的情况")
        if not validations["added_cells_centered"]:
            anomalies.append("存在未居中的新增数据单元格")
        if not validations["added_cells_bordered"]:
            anomalies.append("存在未添加边框的新增数据单元格")
        if not validations["added_cells_content_font_rule_12"]:
            anomalies.append("存在未按“含中文宋体、不含中文Arial”规则设置 12 号字体的新增数据单元格")
        if not validations["added_rows_height_30"]:
            anomalies.append("存在行高不是 30 磅的新增数据行")

        anomalies = dedupe_preserve_order(anomalies)

        report = {
            "files_read": {
                "rpa_docx": "rpa.docx",
                "template_workbook": "generated_in_code",
                "template_sheet": template_selection.sheet_name,
                "resource_dir": str(resource_dir),
                "source_dir": str(source_dir),
                "output_dir": str(output_dir),
                "source_workbooks": [path.name for path in source_files],
                "msku_mapping_workbook": msku_map_path.name,
                "msku_mapping_sheet": msku_selection.sheet_name,
                "store_detail_workbook": store_detail_path.name,
                "store_detail_sheet": store_selection.sheet_name,
            },
            "header_rows": {
                "template_header_row": template_selection.header_row,
                "template_data_start_row": detect_data_start_row(template_sheet, template_selection.header_row),
                "msku_mapping_header_row": msku_selection.header_row,
                "store_detail_header_row": store_selection.header_row,
            },
            "source_file_order": {
                "ticket_count": len(source_files),
                "all_source_files": [path.name for path in source_files],
            },
            "field_mapping": FIELD_MAPPING,
            "store_lookup_summary": {
                "resolved_store_sites": dedupe_preserve_order(all_store_sites),
                "resolved_store_shorts": dedupe_preserve_order(all_store_shorts),
                "resolved_store_names": dedupe_preserve_order(all_store_names),
                "resolved_mapped_lines": dedupe_preserve_order(all_mapped_lines),
                "rows": lookup_rows_for_report,
            },
            "block_reports": block_reports,
            "styling_updates": {
                "sheet_name": output_sheet_name,
                "box_value_prefix": "编号",
                "title_format": "装箱单（店铺简称-站点）海运第x票",
                "box_header_extra_line": "每票块显示对应FBA号",
                "added_data_font": "含中文宋体 12；不含中文 Arial 12",
                "added_row_height": DATA_ROW_HEIGHT,
                "summary_row_added": True,
            },
            "validations": validations,
            "anomalies": anomalies,
            "output_workbook": output_path.name,
            "processing_output_files": [
                {
                    "format_type": "ONE_SKU",
                    "output_workbook": output_path.name,
                    "source_workbooks": [path.name for path in source_files],
                    "ticket_count": len(source_files),
                    "total_units": total_expected_set,
                    "total_cartons": total_expected_carton,
                }
            ],
        }

        report_path = output_path.with_name(f"{output_path.stem}_report.json")
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        report["report_file"] = report_path.name
        return report
    finally:
        template_workbook.close()


def process_workbooks(resource_dir: Path, source_dir: Path, output_dir: Path) -> dict[str, Any]:
    source_files = select_source_files(source_dir)
    if not source_files:
        raise FileNotFoundError("未找到可处理的领星源数据 Excel 文件。")

    output_dir.mkdir(parents=True, exist_ok=True)

    classified_files: list[SourceWorkbookInfo] = []
    unknown_files: list[Path] = []
    source_metadata: dict[Path, dict[str, Any]] = {}
    for source_path in source_files:
        source_info = classify_source_workbook(source_path)
        if source_info is None:
            unknown_files.append(source_path)
        else:
            classified_files.append(source_info)
            source_metadata[source_info.path] = read_source_metadata(source_info)

    ups_infos = [
        info
        for info in classified_files
        if source_name_contains_ups(source_metadata.get(info.path, {}))
    ]
    freight_one_sku_files = [
        info.path
        for info in classified_files
        if info not in ups_infos and info.format_type == "ONE_SKU"
    ]
    unsupported_non_ups_infos = [
        info
        for info in classified_files
        if info not in ups_infos and info.format_type != "ONE_SKU"
    ]

    if freight_one_sku_files and not ups_infos and not unknown_files and not unsupported_non_ups_infos:
        report = process_one_sku_workbooks(resource_dir, source_dir, output_dir, source_files=freight_one_sku_files)
        report["source_output_type"] = "FREIGHT_INFO"
        report["source_file_formats"] = [
            {
                "source_workbook": info.path.name,
                "source_structure": info.format_type,
                "source_output_type": "FREIGHT_INFO",
                "cargo_name": source_metadata.get(info.path, {}).get("货件名称"),
            }
            for info in classified_files
        ]
        for item in report.get("processing_output_files", []):
            item["source_output_type"] = "FREIGHT_INFO"
            item["source_structure"] = "ONE_SKU"
        report_file = report.get("report_file")
        if report_file:
            (output_dir / str(report_file)).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report

    output_files: list[dict[str, Any]] = []
    child_reports: list[dict[str, Any]] = []
    anomalies: list[str] = []

    if unknown_files:
        anomalies.extend([f"无法识别源文件格式，已跳过：{path.name}" for path in unknown_files])

    if unsupported_non_ups_infos:
        anomalies.extend(
            [
                f"非 UPS 源文件不是第一种单款箱结构，已跳过：{info.path.name}"
                for info in unsupported_non_ups_infos
            ]
        )

    if freight_one_sku_files:
        one_report = process_one_sku_workbooks(resource_dir, source_dir, output_dir, source_files=freight_one_sku_files)
        child_reports.append(one_report)
        output_files.append(
            {
                "format_type": "FREIGHT_INFO",
                "source_output_type": "FREIGHT_INFO",
                "output_workbook": one_report.get("output_workbook"),
                "report_file": one_report.get("report_file"),
                "source_workbooks": [path.name for path in freight_one_sku_files],
                "ticket_count": len(freight_one_sku_files),
            }
        )
        anomalies.extend(one_report.get("anomalies", []))

    if ups_infos:
        ups_report = process_ups_packing_workbooks(resource_dir, output_dir, ups_infos)
        child_reports.append(ups_report)
        output_files.append(
            {
                "format_type": "UPS_PACKING",
                "source_output_type": "UPS_PACKING",
                "output_workbook": ups_report.get("output_workbook"),
                "report_file": ups_report.get("report_file"),
                "source_workbooks": [info.path.name for info in ups_infos],
                "ticket_count": len(ups_infos),
                "total_units": ups_report.get("total_units"),
                "total_cartons": ups_report.get("total_cartons"),
            }
        )
        anomalies.extend(ups_report.get("anomalies", []))

    if not output_files:
        skipped_names = ", ".join(path.name for path in unknown_files) or "无"
        raise ValueError(f"未生成任何输出文件；无法识别或不可处理的源文件：{skipped_names}")

    primary_output = next((item.get("output_workbook") for item in output_files if item.get("output_workbook")), None)
    output_types = dedupe_preserve_order(
        [item.get("source_output_type") for item in output_files if item.get("source_output_type")]
    )
    report = {
        "files_read": {
            "resource_dir": str(resource_dir),
            "source_dir": str(source_dir),
            "output_dir": str(output_dir),
            "source_workbooks": [path.name for path in source_files],
        },
        "source_file_formats": [
            {
                "source_workbook": info.path.name,
                "source_structure": info.format_type,
                "source_output_type": "UPS_PACKING"
                if info in ups_infos
                else ("FREIGHT_INFO" if info.path in freight_one_sku_files else "UNSUPPORTED"),
                "cargo_name": source_metadata.get(info.path, {}).get("货件名称"),
            }
            for info in classified_files
        ]
        + [
            {"source_workbook": path.name, "source_structure": "UNKNOWN", "source_output_type": "UNKNOWN"}
            for path in unknown_files
        ],
        "child_reports": [
            {
                "format_type": child.get("format_type", "ONE_SKU"),
                "source_output_type": child.get("source_output_type"),
                "output_workbook": child.get("output_workbook"),
                "report_file": child.get("report_file"),
            }
            for child in child_reports
        ],
        "processing_output_files": output_files,
        "source_output_type": output_types[0] if len(output_types) == 1 else "MIXED",
        "anomalies": dedupe_preserve_order(anomalies),
        "output_workbook": primary_output,
    }
    report_path = output_dir / "processing_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    report["report_file"] = report_path.name
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="领星 Excel 批量整理工具")
    parser.add_argument("--resource-dir", default=str(Path(__file__).resolve().parent))
    parser.add_argument("--source-dir", default=str(Path(__file__).resolve().parent))
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    resource_dir = Path(args.resource_dir).resolve()
    source_dir = Path(args.source_dir).resolve()
    output_dir = Path(args.output_dir).resolve() if args.output_dir else source_dir
    report = process_workbooks(resource_dir, source_dir, output_dir)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
