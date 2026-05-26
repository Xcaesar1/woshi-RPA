from __future__ import annotations

import argparse
import concurrent.futures
import io
import json
import os
import re
import sys
import time
import traceback
import urllib.request
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable
from urllib.parse import unquote

from openpyxl import load_workbook

from lingxing_excel_processor import process_workbooks


LINGXING_HOME_URL = "https://erp.lingxing.com/erp/home"
LINGXING_FBA_CARGO_URL = "https://erp.lingxing.com/erp/msupply/fbaCargo"
LINGXING_ERP_BASE_URL = "https://erp.lingxing.com"
LINGXING_GW_BASE_URL = "https://gw.lingxingerp.com"
DEFAULT_CONFIG_FILE_NAME = "lingxing_rpa.local.json"
MANIFEST_HEADER_CANDIDATES = ["FBA号", "货件单号", "FBA"]
DEFAULT_PAGE_TIMEOUT = 30
DEFAULT_DOWNLOAD_TIMEOUT = 120
DEFAULT_MODAL_TIMEOUT = 20
DEFAULT_LOGIN_TIMEOUT = 20
DEFAULT_SEARCH_TIMEOUT = 30
DOWNLOAD_STABLE_CHECK_SECONDS = 1.0
VISIBLE_CLICKABLE_SELECTOR = "a,button,[role='button'],li,div,span,i,svg"
BEIJING_TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")
BLOCKED_RESOURCE_TYPES = {"image", "media", "font"}


def beijing_now() -> datetime:
    return datetime.now(BEIJING_TZ)


def beijing_now_text() -> str:
    return beijing_now().strftime("%Y-%m-%d %H:%M:%S")


class AutomationError(RuntimeError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


@dataclass
class LoginCredentials:
    username: str
    password: str


@dataclass
class RawExportRequest:
    url: str
    method: str
    headers: dict[str, str]
    data: bytes | None
    suggested_name: str | None


def dedupe_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            output.append(item)
    return output


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def normalize_fba(value: Any) -> str | None:
    text = normalize_text(value).upper()
    return text or None


def sanitize_filename_part(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", value).strip(" ._") or "UNKNOWN"


def extract_download_type_marker(filename: str) -> str | None:
    upper_name = filename.upper()
    if "MUL_SKU" in upper_name:
        return "MUL_SKU"
    if "ONE_SKU" in upper_name:
        return "ONE_SKU"
    return None


def build_download_filename(fba_code: str, index: int, warehouse_code: str, original_name: str, suffix: str) -> str:
    type_marker = extract_download_type_marker(original_name)
    parts = [sanitize_filename_part(fba_code), f"{index:02d}", warehouse_code]
    if type_marker:
        parts.append(type_marker)
    parts.append("NO_PIC")
    return "_".join(parts) + suffix


def filename_from_content_disposition(value: str | None) -> str | None:
    if not value:
        return None
    utf8_match = re.search(r"filename\*=UTF-8''([^;]+)", value, flags=re.IGNORECASE)
    if utf8_match:
        return sanitize_filename_part(unquote(utf8_match.group(1)).strip().strip('"'))
    plain_match = re.search(r'filename="?([^";]+)"?', value, flags=re.IGNORECASE)
    if plain_match:
        return sanitize_filename_part(unquote(plain_match.group(1)).strip().strip('"'))
    return None


def build_timestamp() -> str:
    return beijing_now().strftime("%Y%m%d-%H%M%S")


def elapsed_seconds(start: float) -> float:
    return round(time.perf_counter() - start, 3)


def is_valid_xlsx_payload(body: bytes) -> bool:
    if not body.startswith(b"PK"):
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(body)) as archive:
            return archive.testzip() is None
    except zipfile.BadZipFile:
        return False


def compact_response_preview(body: bytes, limit: int = 120) -> str:
    text = body[:1000].decode("utf-8", errors="ignore")
    text = normalize_text(text)
    if not text:
        return "空响应"
    if len(text) > limit:
        return f"{text[:limit]}..."
    return text


def env_int(name: str, default: int, *, minimum: int = 1, maximum: int | None = None) -> int:
    raw = os.environ.get(name)
    if raw is None:
        return default
    try:
        value = int(raw)
    except ValueError:
        return default
    value = max(minimum, value)
    if maximum is not None:
        value = min(value, maximum)
    return value


def json_default(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, datetime):
        return value.astimezone(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")
    raise TypeError(f"Object of type {type(value)!r} is not JSON serializable")


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")


def relative_to_base(path: Path | None, base_dir: Path) -> str | None:
    if path is None:
        return None
    try:
        return str(path.relative_to(base_dir))
    except ValueError:
        return str(path)


def load_login_credentials(config_path: Path | None) -> LoginCredentials | None:
    if config_path is None or not config_path.exists():
        return None

    payload = json.loads(config_path.read_text(encoding="utf-8"))
    username = normalize_text(payload.get("username"))
    password = normalize_text(payload.get("password"))
    if not username or not password:
        raise ValueError(f"登录配置文件缺少 username 或 password：{config_path}")
    return LoginCredentials(username=username, password=password)


def parse_manifest_txt(path: Path) -> list[str]:
    fba_codes: list[str] = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        fba_code = normalize_fba(line)
        if fba_code:
            fba_codes.append(fba_code)
    return dedupe_preserve_order(fba_codes)


def parse_manifest_xlsx(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_row_index: int | None = None
        value_column_index = 1

        for row_idx in range(1, min(worksheet.max_row, 5) + 1):
            row_values = [
                normalize_header(worksheet.cell(row=row_idx, column=col_idx).value)
                for col_idx in range(1, worksheet.max_column + 1)
            ]
            for candidate in MANIFEST_HEADER_CANDIDATES:
                if candidate in row_values:
                    header_row_index = row_idx
                    value_column_index = row_values.index(candidate) + 1
                    break
            if header_row_index is not None:
                break

        start_row = 2 if header_row_index == 1 else (header_row_index + 1 if header_row_index else 1)
        fba_codes: list[str] = []
        for row in worksheet.iter_rows(min_row=start_row, values_only=True):
            if not row:
                continue
            cell_value = row[value_column_index - 1] if len(row) >= value_column_index else row[0]
            fba_code = normalize_fba(cell_value)
            if fba_code:
                fba_codes.append(fba_code)

        return dedupe_preserve_order(fba_codes)
    finally:
        workbook.close()


def parse_manifest_file(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return parse_manifest_txt(path)
    if suffix == ".xlsx":
        return parse_manifest_xlsx(path)
    raise ValueError(f"不支持的 manifest 文件类型：{path.name}")


def locate_chrome_binary() -> Path | None:
    candidates = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def env_flag(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on"}


def import_selenium_bindings() -> dict[str, Any]:
    try:
        from selenium import webdriver
        from selenium.common.exceptions import (
            ElementClickInterceptedException,
            JavascriptException,
            MoveTargetOutOfBoundsException,
            NoSuchElementException,
            StaleElementReferenceException,
            TimeoutException,
            WebDriverException,
        )
        from selenium.webdriver import ActionChains
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait
    except ModuleNotFoundError as exc:
        raise RuntimeError("未安装 selenium，请先执行：python -m pip install -r requirements-rpa.txt") from exc

    return {
        "webdriver": webdriver,
        "ActionChains": ActionChains,
        "By": By,
        "Keys": Keys,
        "EC": EC,
        "WebDriverWait": WebDriverWait,
        "ElementClickInterceptedException": ElementClickInterceptedException,
        "JavascriptException": JavascriptException,
        "MoveTargetOutOfBoundsException": MoveTargetOutOfBoundsException,
        "NoSuchElementException": NoSuchElementException,
        "StaleElementReferenceException": StaleElementReferenceException,
        "TimeoutException": TimeoutException,
        "WebDriverException": WebDriverException,
    }


def import_playwright_bindings() -> dict[str, Any]:
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ModuleNotFoundError as exc:
        raise RuntimeError("未安装 playwright，请先执行：python -m pip install -r requirements-rpa.txt") from exc

    return {
        "sync_playwright": sync_playwright,
        "PlaywrightTimeoutError": PlaywrightTimeoutError,
    }


class LingxingBrowserAutomation:
    def __init__(self, profile_dir: Path, credentials: LoginCredentials | None):
        bindings = import_selenium_bindings()
        self.webdriver = bindings["webdriver"]
        self.ActionChains = bindings["ActionChains"]
        self.By = bindings["By"]
        self.Keys = bindings["Keys"]
        self.EC = bindings["EC"]
        self.WebDriverWait = bindings["WebDriverWait"]
        self.ElementClickInterceptedException = bindings["ElementClickInterceptedException"]
        self.JavascriptException = bindings["JavascriptException"]
        self.MoveTargetOutOfBoundsException = bindings["MoveTargetOutOfBoundsException"]
        self.NoSuchElementException = bindings["NoSuchElementException"]
        self.StaleElementReferenceException = bindings["StaleElementReferenceException"]
        self.TimeoutException = bindings["TimeoutException"]
        self.WebDriverException = bindings["WebDriverException"]
        self.profile_dir = profile_dir
        self.credentials = credentials
        self.chrome_binary = locate_chrome_binary()
        self.driver = None
        self.current_screenshot_dir: Path | None = None

    def start(self) -> None:
        self.profile_dir.mkdir(parents=True, exist_ok=True)
        options = self.webdriver.ChromeOptions()
        if self.chrome_binary:
            options.binary_location = str(self.chrome_binary)
        if hasattr(options, "ignore_local_proxy_environment_variables"):
            options.ignore_local_proxy_environment_variables()
        options.add_argument(f"--user-data-dir={self.profile_dir}")
        options.add_argument("--disable-popup-blocking")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        options.add_argument("--no-proxy-server")
        options.add_experimental_option(
            "prefs",
            {
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "profile.default_content_setting_values.automatic_downloads": 1,
                "safebrowsing.enabled": True,
            },
        )
        try:
            self.driver = self.webdriver.Chrome(options=options)
        except self.WebDriverException as exc:
            raise AutomationError("browser_start_failed", f"启动 Chrome 自动化失败：{exc}") from exc
        self.driver.set_window_size(1600, 1000)
        self.driver.implicitly_wait(1)

    def close(self) -> None:
        if self.driver is not None:
            self.driver.quit()
            self.driver = None

    def capture_screenshot(self, name: str) -> Path | None:
        if self.driver is None or self.current_screenshot_dir is None:
            return None
        self.current_screenshot_dir.mkdir(parents=True, exist_ok=True)
        screenshot_path = self.current_screenshot_dir / f"{sanitize_filename_part(name)}.png"
        try:
            self.driver.save_screenshot(str(screenshot_path))
            return screenshot_path
        except self.WebDriverException:
            return None

    def current_page_state(self) -> dict[str, Any]:
        if self.driver is None:
            return {"current_url": None, "page_title": None}
        try:
            return {
                "current_url": self.driver.current_url,
                "page_title": self.driver.title,
            }
        except self.WebDriverException:
            return {"current_url": None, "page_title": None}

    def _wait_until(self, predicate, timeout: int = DEFAULT_PAGE_TIMEOUT, message: str | None = None):
        return self.WebDriverWait(self.driver, timeout).until(lambda driver: predicate(), message=message)

    def _wait_for_document_ready(self, timeout: int = DEFAULT_PAGE_TIMEOUT) -> None:
        self._wait_until(
            lambda: self.driver.execute_script("return document.readyState") == "complete",
            timeout=timeout,
            message="页面加载超时",
        )

    def _page_text(self) -> str:
        if self.driver is None:
            return ""
        try:
            return normalize_text(self.driver.find_element(self.By.TAG_NAME, "body").text)
        except self.WebDriverException:
            return ""

    def _page_contains_any(self, texts: list[str]) -> bool:
        page_text = self._page_text()
        return any(text in page_text for text in texts)

    def _is_login_page(self) -> bool:
        if self.driver is None:
            return False
        try:
            current_url = (self.driver.current_url or "").lower()
        except self.WebDriverException:
            current_url = ""
        if "/login" in current_url:
            return True

        page_text = self._page_text()
        return "账号登录" in page_text and "登录" in page_text and bool(self._find_password_input(timeout=0))

    def _find_first_visible(self, locators: list[tuple[str, str]], root=None, timeout: float = 0):
        root = root or self.driver
        deadline = time.time() + timeout
        while True:
            for by, locator in locators:
                try:
                    elements = root.find_elements(by, locator)
                except self.WebDriverException:
                    elements = []
                for element in elements:
                    try:
                        if element.is_displayed():
                            return element
                    except (self.StaleElementReferenceException, self.WebDriverException):
                        continue
            if time.time() >= deadline:
                return None
            time.sleep(0.3)

    def _find_username_input(self, timeout: float = DEFAULT_LOGIN_TIMEOUT):
        return self._find_first_visible(
            [
                (self.By.CSS_SELECTOR, "input[name='account']"),
                (self.By.CSS_SELECTOR, "input[placeholder*='账号']"),
                (self.By.CSS_SELECTOR, "input[placeholder*='用户名']"),
                (self.By.CSS_SELECTOR, "input[type='text']"),
            ],
            timeout=timeout,
        )

    def _find_password_input(self, timeout: float = DEFAULT_LOGIN_TIMEOUT):
        return self._find_first_visible(
            [
                (self.By.CSS_SELECTOR, "input[name='pwd']"),
                (self.By.CSS_SELECTOR, "input[type='password']"),
            ],
            timeout=timeout,
        )

    def _find_login_button(self, timeout: float = DEFAULT_LOGIN_TIMEOUT):
        by_dom = self._find_dom_clickable(timeout=timeout, text="登录", exact_text=True, class_contains="loginBtn")
        if by_dom is not None:
            return by_dom
        return self._find_first_visible(
            [
                (self.By.CSS_SELECTOR, "button.loginBtn"),
                (self.By.CSS_SELECTOR, "button.el-button.loginBtn"),
                (self.By.XPATH, "//button[contains(@class, 'loginBtn') and .//span[normalize-space(.)='登录']]"),
                (self.By.XPATH, "//button[normalize-space(.)='登录']"),
            ],
            timeout=timeout,
        )

    def _fill_input(self, element, value: str) -> None:
        try:
            self._safe_click(element)
        except AutomationError:
            try:
                self.driver.execute_script("arguments[0].focus();", element)
            except self.JavascriptException:
                pass

        try:
            element.send_keys(self.Keys.CONTROL, "a")
            element.send_keys(self.Keys.DELETE)
            element.send_keys(value)
            return
        except self.WebDriverException:
            pass

        # send_keys 被拦截时，兜底使用 DOM 赋值并触发输入事件。
        try:
            self.driver.execute_script(
                """
const el = arguments[0];
const val = arguments[1];
el.focus();
el.value = '';
el.dispatchEvent(new Event('input', { bubbles: true }));
el.value = val;
el.dispatchEvent(new Event('input', { bubbles: true }));
el.dispatchEvent(new Event('change', { bubbles: true }));
""",
                element,
                value,
            )
        except self.JavascriptException as exc:
            raise AutomationError("input_fill_failed", f"无法写入输入框：{exc}") from exc

    def _scan_clickable_dom(self, container=None) -> list[dict[str, Any]]:
        container = container or self.driver.find_element(self.By.TAG_NAME, "body")
        script = """
const root = arguments[0];
const selector = arguments[1];
const isVisible = (el) => {
  if (!el) return false;
  const style = window.getComputedStyle(el);
  if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
  const rect = el.getBoundingClientRect();
  return rect.width > 0 && rect.height > 0;
};
const textFor = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
const classNameFor = (el) => {
  if (!el) return '';
  if (typeof el.className === 'string') return el.className;
  if (el.className && typeof el.className.baseVal === 'string') return el.className.baseVal;
  return '';
};
const isClickable = (el) => {
  if (!el) return false;
  if (['A', 'BUTTON', 'SUMMARY'].includes(el.tagName)) return true;
  if ((el.getAttribute('role') || '').toLowerCase() === 'button') return true;
  if (typeof el.onclick === 'function') return true;
  return window.getComputedStyle(el).cursor === 'pointer';
};
const output = [];
const seen = new Set();
for (const node of root.querySelectorAll(selector)) {
  if (!isVisible(node)) continue;
  let target = node;
  while (target && target !== root && !isClickable(target)) target = target.parentElement;
  if (!target || !isVisible(target)) target = node;
  const key = target.tagName + '|' + (target.id || '') + '|' + classNameFor(target) + '|' + textFor(target);
  if (seen.has(key)) continue;
  seen.add(key);
  output.push({
    element: target,
    tag: (target.tagName || '').toUpperCase(),
    text: textFor(target),
    class_name: classNameFor(target),
    id: target.id || '',
    name: target.getAttribute('name') || '',
    role: target.getAttribute('role') || '',
    title: target.getAttribute('title') || '',
    aria_label: target.getAttribute('aria-label') || '',
    data_testid: target.getAttribute('data-testid') || '',
  });
}
return output;
"""
        try:
            result = self.driver.execute_script(script, container, VISIBLE_CLICKABLE_SELECTOR)
            return result or []
        except self.JavascriptException:
            return []

    def _find_dom_clickable(
        self,
        *,
        timeout: float = DEFAULT_PAGE_TIMEOUT,
        container=None,
        text: str | None = None,
        exact_text: bool = True,
        class_contains: str | None = None,
        id_contains: str | None = None,
        name_equals: str | None = None,
        title_contains: str | None = None,
        aria_contains: str | None = None,
    ):
        deadline = time.time() + timeout
        expected_text = normalize_text(text) if text is not None else None
        expected_class = (class_contains or "").strip().lower()
        expected_id = (id_contains or "").strip().lower()
        expected_name = (name_equals or "").strip().lower()
        expected_title = (title_contains or "").strip().lower()
        expected_aria = (aria_contains or "").strip().lower()

        while True:
            for record in self._scan_clickable_dom(container=container):
                node_text = normalize_text(record.get("text"))
                node_class = str(record.get("class_name") or "").lower()
                node_id = str(record.get("id") or "").lower()
                node_name = str(record.get("name") or "").lower()
                node_title = str(record.get("title") or "").lower()
                node_aria = str(record.get("aria_label") or "").lower()

                if expected_text is not None:
                    matched = node_text == expected_text if exact_text else (expected_text in node_text)
                    if not matched:
                        continue
                if expected_class and expected_class not in node_class:
                    continue
                if expected_id and expected_id not in node_id:
                    continue
                if expected_name and expected_name != node_name:
                    continue
                if expected_title and expected_title not in node_title:
                    continue
                if expected_aria and expected_aria not in node_aria:
                    continue

                element = record.get("element")
                if element is not None:
                    return element

            if time.time() >= deadline:
                return None
            time.sleep(0.4)

    def _find_clickable_text_matches(self, text: str, exact: bool = True, container=None) -> list[Any]:
        container = container or self.driver.find_element(self.By.TAG_NAME, "body")
        script = """
const searchText = arguments[0].trim();
const exact = arguments[1];
const root = arguments[2];
const selector = arguments[3];
const isVisible = (el) => {
  if (!el) return false;
  const style = window.getComputedStyle(el);
  if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
  const rect = el.getBoundingClientRect();
  return rect.width > 0 && rect.height > 0;
};
const textFor = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
const isClickable = (el) => {
  if (!el) return false;
  if (['A', 'BUTTON', 'SUMMARY'].includes(el.tagName)) return true;
  if ((el.getAttribute('role') || '').toLowerCase() === 'button') return true;
  if (typeof el.onclick === 'function') return true;
  return window.getComputedStyle(el).cursor === 'pointer';
};
const seen = new Set();
const output = [];
for (const node of root.querySelectorAll(selector)) {
  if (!isVisible(node)) continue;
  const nodeText = textFor(node);
  if (!nodeText) continue;
  const matched = exact ? nodeText === searchText : nodeText.includes(searchText);
  if (!matched) continue;
  let target = node;
  while (target && target !== root && !isClickable(target)) target = target.parentElement;
  if (!target || !isVisible(target)) target = node;
  if (seen.has(target)) continue;
  seen.add(target);
  output.push(target);
}
return output;
"""
        try:
            return self.driver.execute_script(script, text, exact, container, VISIBLE_CLICKABLE_SELECTOR)
        except self.JavascriptException:
            return []

    def _safe_click(self, element) -> None:
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        except self.JavascriptException:
            pass

        try:
            element.click()
            return
        except (
            self.ElementClickInterceptedException,
            self.StaleElementReferenceException,
            self.WebDriverException,
        ):
            pass

        try:
            clickable = self.driver.execute_script(
                """
let el = arguments[0];
while (el) {
  if (['A', 'BUTTON', 'SUMMARY'].includes(el.tagName)) return el;
  if ((el.getAttribute('role') || '').toLowerCase() === 'button') return el;
  if (typeof el.onclick === 'function') return el;
  if (window.getComputedStyle(el).cursor === 'pointer') return el;
  el = el.parentElement;
}
return arguments[0];
""",
                element,
            )
            self.driver.execute_script("arguments[0].click();", clickable)
        except self.JavascriptException as exc:
            raise AutomationError("click_failed", f"无法点击页面元素：{exc}") from exc

    def _click_text(self, text: str, exact: bool = True, timeout: int = DEFAULT_PAGE_TIMEOUT, container=None) -> Any:
        dom_hit = self._find_dom_clickable(timeout=timeout, container=container, text=text, exact_text=exact)
        if dom_hit is not None:
            self._safe_click(dom_hit)
            return dom_hit

        deadline = time.time() + timeout
        while True:
            matches = self._find_clickable_text_matches(text, exact=exact, container=container)
            if matches:
                self._safe_click(matches[0])
                return matches[0]
            if time.time() >= deadline:
                break
            time.sleep(0.5)
        raise AutomationError("text_not_found", f"未找到可点击文本：{text}")

    def _find_search_input(self):
        body = self.driver.find_element(self.By.TAG_NAME, "body")

        # 先用 DOM 结构定位：输入框本身可编辑，且附近存在“搜索”图标/按钮。
        dom_candidate = self.driver.execute_script(
            """
const isVisible = (el) => {
  if (!el) return false;
  const style = window.getComputedStyle(el);
  if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
  const rect = el.getBoundingClientRect();
  return rect.width > 0 && rect.height > 0;
};
const hasSearchHint = (el) => {
  if (!el) return false;
  const q = '[class*=\"search\"],[class*=\"sousuo\"],[title*=\"搜索\"],[aria-label*=\"搜索\"],i[class*=\"icon-sousuo\"],i[class*=\"icon-search\"]';
  return !!el.querySelector(q);
};
const rowText = (el) => ((el && (el.innerText || el.textContent)) || '').replace(/\\s+/g, ' ');
let best = null;
let bestScore = -1;
for (const input of document.querySelectorAll('input.el-input__inner')) {
  if (!isVisible(input)) continue;
  if (input.readOnly || input.disabled) continue;
  const type = (input.getAttribute('type') || 'text').toLowerCase();
  if (type !== 'text' && type !== 'search') continue;
  const container = input.closest('.el-input-group') || input.closest('.el-input') || input.parentElement;
  const line = container ? (container.closest('.tool_row') || container.closest('.el-form-item') || container.parentElement) : null;
  const text = rowText(line || container || document.body);
  let score = 0;
  if ((input.getAttribute('inelement') || '') === '0') score += 3;
  if (hasSearchHint(container) || hasSearchHint(line)) score += 5;
  if (text.includes('货件单号')) score += 4;
  if (!input.getAttribute('placeholder')) score += 1;
  if (score > bestScore) {
    bestScore = score;
    best = input;
  }
}
return best;
""",
        )
        if dom_candidate is not None:
            return dom_candidate

        return self._find_first_visible(
            [
                (self.By.CSS_SELECTOR, "input[inelement='0']:not([readonly])"),
                (self.By.CSS_SELECTOR, "input.el-input__inner[autocomplete='off']:not([readonly])"),
                (
                    self.By.XPATH,
                    ".//*[contains(normalize-space(.), '货件单号') and .//input and not(.//*[contains(normalize-space(.), '货件单号') and .//input])]//input[not(@type='hidden') and not(@readonly)]",
                ),
                (self.By.CSS_SELECTOR, "input[placeholder*='货件单号']:not([readonly])"),
                (self.By.CSS_SELECTOR, "input[placeholder*='搜索']:not([readonly])"),
                (self.By.CSS_SELECTOR, "input[type='text']:not([readonly])"),
            ],
            root=body,
            timeout=DEFAULT_SEARCH_TIMEOUT,
        )

    def _click_shipment_search_icon(self, search_input) -> bool:
        # 优先在搜索输入框附近找“搜索”图标，避免误点页面其他按钮。
        script = """
const input = arguments[0];
if (!input) return null;
const isVisible = (el) => {
  if (!el) return false;
  const style = window.getComputedStyle(el);
  if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
  const rect = el.getBoundingClientRect();
  return rect.width > 0 && rect.height > 0;
};
const classNameFor = (el) => {
  if (!el) return '';
  if (typeof el.className === 'string') return el.className;
  if (el.className && typeof el.className.baseVal === 'string') return el.className.baseVal;
  return '';
};
const textFor = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
const isClickable = (el) => {
  if (!el) return false;
  if (['A', 'BUTTON', 'SUMMARY'].includes(el.tagName)) return true;
  if ((el.getAttribute('role') || '').toLowerCase() === 'button') return true;
  if (typeof el.onclick === 'function') return true;
  return window.getComputedStyle(el).cursor === 'pointer';
};
const climbClickable = (el, root) => {
  let cur = el;
  while (cur && cur !== root && !isClickable(cur)) cur = cur.parentElement;
  return cur || el;
};

const root = input.closest('.el-input-group') || input.closest('.el-form-item') || input.parentElement?.parentElement || document.body;
const inputRect = input.getBoundingClientRect();
let best = null;
let bestScore = Number.POSITIVE_INFINITY;
for (const node of root.querySelectorAll('button,[role=\"button\"],i,span,svg')) {
  if (!isVisible(node)) continue;
  const cls = classNameFor(node).toLowerCase();
  const title = (node.getAttribute('title') || '').toLowerCase();
  const aria = (node.getAttribute('aria-label') || '').toLowerCase();
  const text = textFor(node);
  const hints = cls.includes('search') || cls.includes('sousuo') || title.includes('搜索') || aria.includes('搜索') || text.includes('搜索');
  if (!hints) continue;

  const target = climbClickable(node, root);
  if (!isVisible(target)) continue;
  const rect = target.getBoundingClientRect();
  const dx = rect.left - inputRect.right;
  const dy = Math.abs((rect.top + rect.bottom) / 2 - (inputRect.top + inputRect.bottom) / 2);
  if (dx < -20) continue;
  const score = Math.max(0, dx) + dy * 0.6;
  if (score < bestScore) {
    best = target;
    bestScore = score;
  }
}
return best;
"""
        try:
            candidate = self.driver.execute_script(script, search_input)
        except self.JavascriptException:
            candidate = None

        if candidate is not None:
            self._safe_click(candidate)
            return True

        local_icon = self._find_first_visible(
            [
                (self.By.XPATH, ".//*[contains(@class, 'sousuo') or contains(@class, 'search') or contains(@title, '搜索') or contains(@aria-label, '搜索')]"),
            ],
            root=search_input.find_element(self.By.XPATH, "./ancestor::*[1]"),
            timeout=1,
        )
        if local_icon is not None:
            self._safe_click(local_icon)
            return True

        return False

    def _trigger_shipment_search(self, search_input) -> None:
        clicked = self._click_shipment_search_icon(search_input)
        if not clicked:
            search_input.send_keys(self.Keys.ENTER)

    def _ensure_search_dimension(self) -> None:
        if self._page_contains_any(["货件单号"]):
            return
        self._click_text("货件单号", exact=True, timeout=10)

    def _wait_for_login_resolution(self) -> None:
        deadline = time.time() + max(DEFAULT_LOGIN_TIMEOUT, 35)
        while time.time() < deadline:
            if not self._is_login_page():
                return
            time.sleep(0.5)

    def _pause_for_manual_step(self, prompt: str) -> None:
        print(prompt)
        if not sys.stdin or not sys.stdin.isatty():
            raise AutomationError("manual_step_unavailable", "当前运行环境不可交互，无法等待人工输入")
        input("完成后按 Enter 继续...")
        self._wait_for_document_ready(timeout=DEFAULT_PAGE_TIMEOUT)

    def _submit_login(self) -> None:
        login_button = self._find_login_button()
        if login_button is None:
            raise AutomationError("login_button_not_found", "未找到登录按钮（button.loginBtn）")
        self._safe_click(login_button)
        self._wait_for_login_resolution()

        # 某些页面状态下 click 可能未触发提交，补一个 Enter 提交兜底。
        if self._is_login_page():
            password_input = self._find_password_input(timeout=3)
            if password_input is not None:
                password_input.send_keys(self.Keys.ENTER)
                self._wait_for_login_resolution()

    def ensure_logged_in(self) -> None:
        self.driver.get(LINGXING_HOME_URL)
        self._wait_for_document_ready()
        if not self._is_login_page():
            return

        if self.credentials is None:
            raise AutomationError("credentials_missing", "仍在登录页，且未提供登录配置文件（username/password）")

        username_input = self._find_username_input()
        password_input = self._find_password_input()
        if username_input is None or password_input is None:
            self.capture_screenshot("login_fields_not_found")
            raise AutomationError("login_fields_not_found", "未找到登录输入框（name=account / name=pwd）")

        self._fill_input(username_input, self.credentials.username)
        self._fill_input(password_input, self.credentials.password)
        self._submit_login()

        if self._is_login_page():
            self.capture_screenshot("login_failed_still_on_login_page")
            try:
                iframe_count = len(self.driver.find_elements(self.By.CSS_SELECTOR, "iframe"))
            except self.WebDriverException:
                iframe_count = -1
            raise AutomationError(
                "login_failed",
                f"登录后仍停留在登录页（iframe_count={iframe_count}），请确认账号密码或站点风控状态",
            )

    def open_fba_shipments_page(self) -> None:
        self.ensure_logged_in()
        if self._page_contains_any(["FBA货件", "更多筛选", "旧版货件操作"]):
            return
        self._click_text("FBA", exact=True, timeout=20)
        time.sleep(1)
        self._click_text("FBA货件", exact=True, timeout=20)
        self._wait_until(
            lambda: self._page_contains_any(["货件单号", "更多筛选", "旧版货件操作"]),
            timeout=DEFAULT_PAGE_TIMEOUT,
            message="未进入 FBA货件 页面",
        )

    def _find_exact_shipment_matches(self, fba_code: str) -> list[Any]:
        # 主路径：按列表中的“货件号蓝色链接”结构定位，文本与目标 FBA 精确匹配。
        matches: list[Any] = []
        target_text = normalize_text(fba_code)
        try:
            dom_matches = self.driver.execute_script(
                """
const fba = arguments[0].trim();
const isVisible = (el) => {
  if (!el) return false;
  const style = window.getComputedStyle(el);
  if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
  const rect = el.getBoundingClientRect();
  return rect.width > 0 && rect.height > 0;
};
const out = [];
for (const el of document.querySelectorAll('div.oneLine.ak-blue-pointer')) {
  const text = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
  if (text === fba && isVisible(el)) out.push(el);
}
return out;
""",
                fba_code,
            )
        except self.JavascriptException:
            dom_matches = []
        if dom_matches:
            matches.extend(dom_matches)

        if not matches:
            try:
                matches.extend(self.driver.find_elements(
                    self.By.XPATH,
                    f"//div[contains(@class,'oneLine') and contains(@class,'ak-blue-pointer') and normalize-space(.)='{fba_code}']",
                ))
            except self.WebDriverException:
                pass

        if not matches:
            for record in self._scan_clickable_dom():
                if normalize_text(record.get("text")) == target_text:
                    element = record.get("element")
                    if element is not None:
                        matches.append(element)
        if not matches:
            matches = self._find_clickable_text_matches(fba_code, exact=True)

        filtered: list[Any] = []
        seen_ids: set[str] = set()
        for match in matches:
            try:
                element_id = match.id
                if element_id in seen_ids:
                    continue
                seen_ids.add(element_id)
                if match.is_displayed():
                    filtered.append(match)
            except (self.StaleElementReferenceException, self.WebDriverException):
                continue
        filtered.sort(key=lambda element: (element.location.get("y", 0), element.location.get("x", 0)))
        return filtered

    def search_shipment(self, fba_code: str) -> None:
        self.open_fba_shipments_page()
        self._ensure_search_dimension()
        search_input = self._find_search_input()
        if search_input is None:
            raise AutomationError("search_input_missing", "未找到货件搜索输入框")

        self._fill_input(search_input, fba_code)
        self._trigger_shipment_search(search_input)

        deadline = time.time() + DEFAULT_SEARCH_TIMEOUT
        while time.time() < deadline:
            matches = self._find_exact_shipment_matches(fba_code)
            if matches:
                return
            if self._page_contains_any(["暂无数据", "没有找到", "无数据"]):
                raise AutomationError("shipment_not_found", f"未搜索到 FBA {fba_code}")
            time.sleep(0.5)

        raise AutomationError("shipment_search_timeout", f"搜索 FBA {fba_code} 超时")

    def open_shipment_detail(self, fba_code: str) -> None:
        matches = self._find_exact_shipment_matches(fba_code)
        if not matches:
            raise AutomationError("shipment_not_found", f"未找到 FBA {fba_code} 的精确结果")
        self._safe_click(matches[0])
        self._wait_until(
            lambda: self._page_contains_any(["箱子标签", "装箱明细", "Reference ID", "FC Code"]),
            timeout=DEFAULT_PAGE_TIMEOUT,
            message="未进入货件详情/箱子标签页",
        )

    def _find_shipment_cards(self) -> list[Any]:
        candidate_xpaths = [
            "//*[contains(normalize-space(.), '货件单号') and (contains(normalize-space(.), '物流中心编码') or contains(normalize-space(.), 'FC Code')) and (contains(normalize-space(.), '配送地址') or contains(normalize-space(.), 'Reference ID')) and not(.//*[contains(normalize-space(.), '货件单号') and (contains(normalize-space(.), '物流中心编码') or contains(normalize-space(.), 'FC Code')) and (contains(normalize-space(.), '配送地址') or contains(normalize-space(.), 'Reference ID'))])]",
            "//*[contains(normalize-space(.), '装箱明细') and contains(normalize-space(.), 'SHIPPED') and (contains(normalize-space(.), '货件单号') or contains(normalize-space(.), 'Reference ID')) and not(.//*[contains(normalize-space(.), '装箱明细') and contains(normalize-space(.), 'SHIPPED') and (contains(normalize-space(.), '货件单号') or contains(normalize-space(.), 'Reference ID'))])]",
        ]
        cards: list[Any] = []
        seen_ids: set[str] = set()
        for xpath in candidate_xpaths:
            for element in self.driver.find_elements(self.By.XPATH, xpath):
                try:
                    if not element.is_displayed():
                        continue
                    if element.id in seen_ids:
                        continue
                    seen_ids.add(element.id)
                    cards.append(element)
                except (self.StaleElementReferenceException, self.WebDriverException):
                    continue
            if cards:
                break

        cards.sort(key=lambda element: (element.location.get("y", 0), element.location.get("x", 0)))
        return cards

    def _extract_warehouse_code(self, card_text: str, index: int) -> str:
        normalized = normalize_text(card_text)
        fc_code_match = re.search(r"FC\s*Code\s*[:：]?\s*([A-Z0-9-]+)", normalized, flags=re.IGNORECASE)
        if fc_code_match:
            return sanitize_filename_part(fc_code_match.group(1).upper())

        logistics_match = re.search(r"(?:物流中心编码|仓库编码)\s*[:：]?\s*([A-Z0-9-]+)", normalized)
        if logistics_match:
            return sanitize_filename_part(logistics_match.group(1).upper())

        generic_match = re.search(r"\b[A-Z]{3}\d\b", normalized)
        if generic_match:
            return sanitize_filename_part(generic_match.group(0).upper())

        return f"WAREHOUSE{index:02d}"

    def _wait_for_download_file(self, download_dir: Path, previous_names: set[str]) -> Path:
        deadline = time.time() + DEFAULT_DOWNLOAD_TIMEOUT
        while time.time() < deadline:
            new_files = [
                path
                for path in download_dir.iterdir()
                if path.is_file() and path.name not in previous_names and not path.name.endswith(".crdownload")
            ]
            if new_files:
                newest = max(new_files, key=lambda path: path.stat().st_mtime_ns)
                size_before = newest.stat().st_size
                time.sleep(DOWNLOAD_STABLE_CHECK_SECONDS)
                if newest.exists() and newest.stat().st_size == size_before:
                    return newest
            time.sleep(0.5)
        raise AutomationError("download_timeout", f"等待下载文件超时：{download_dir}")

    def _click_download_button_in_card(self, card) -> None:
        # DOM 主路径：优先匹配下载图标/下载按钮相关 class/title/aria。
        try:
            dom_hit = self.driver.execute_script(
                """
const card = arguments[0];
const isVisible = (el) => {
  if (!el) return false;
  const style = window.getComputedStyle(el);
  if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
  const rect = el.getBoundingClientRect();
  return rect.width > 0 && rect.height > 0;
};
const classNameFor = (el) => {
  if (!el) return '';
  if (typeof el.className === 'string') return el.className;
  if (el.className && typeof el.className.baseVal === 'string') return el.className.baseVal;
  return '';
};
const textFor = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
const isClickable = (el) => {
  if (!el) return false;
  if (['A', 'BUTTON', 'SUMMARY'].includes(el.tagName)) return true;
  if ((el.getAttribute('role') || '').toLowerCase() === 'button') return true;
  if (typeof el.onclick === 'function') return true;
  return window.getComputedStyle(el).cursor === 'pointer';
};
const climb = (el) => {
  let cur = el;
  while (cur && cur !== card && !isClickable(cur)) cur = cur.parentElement;
  return cur || el;
};
const hintMatched = (el) => {
  const cls = classNameFor(el).toLowerCase();
  const title = (el.getAttribute('title') || '').toLowerCase();
  const aria = (el.getAttribute('aria-label') || '').toLowerCase();
  const dataTitle = (el.getAttribute('data-title') || '').toLowerCase();
  const text = textFor(el);
  return (
    text.includes('下载装箱清单') ||
    title.includes('下载装箱清单') ||
    aria.includes('下载装箱清单') ||
    dataTitle.includes('下载装箱清单') ||
    cls.includes('download') ||
    cls.includes('xiazai') ||
    cls.includes('icon-download') ||
    cls.includes('icon-xiazai')
  );
};
let best = null;
let bestScore = Number.POSITIVE_INFINITY;
for (const node of card.querySelectorAll('button,a,i,span,svg,[role=\"button\"]')) {
  if (!isVisible(node)) continue;
  if (!hintMatched(node)) continue;
  const target = climb(node);
  if (!isVisible(target)) continue;
  const rect = target.getBoundingClientRect();
  // 卡片头部右上角通常是下载按钮，优先 y 小且 x 大。
  const score = rect.top * 10 - rect.left;
  if (score < bestScore) {
    best = target;
    bestScore = score;
  }
}
return best;
""",
                card,
            )
        except self.JavascriptException:
            dom_hit = None
        if dom_hit is not None:
            self._safe_click(dom_hit)
            return

        attribute_locators = [
            ".//*[contains(@title, '下载装箱清单') or contains(@aria-label, '下载装箱清单') or contains(@data-title, '下载装箱清单') or contains(normalize-space(.), '下载装箱清单')]",
            ".//*[contains(@title, '下载') or contains(@aria-label, '下载') or contains(@data-title, '下载') or contains(normalize-space(.), '下载')]",
        ]

        for locator in attribute_locators:
            matches = self._find_first_visible([(self.By.XPATH, locator)], root=card, timeout=1)
            if matches is not None:
                self._safe_click(matches)
                return

        # 详情页中的下载入口常常是纯图标，这里通过 hover 识别 tooltip 再点击。
        candidates = card.find_elements(self.By.CSS_SELECTOR, VISIBLE_CLICKABLE_SELECTOR)
        for candidate in candidates[:80]:
            try:
                if not candidate.is_displayed():
                    continue
                self.ActionChains(self.driver).move_to_element(candidate).pause(0.2).perform()
                time.sleep(0.2)
                if self._find_clickable_text_matches("下载装箱清单", exact=False):
                    self._safe_click(candidate)
                    return
            except (
                self.MoveTargetOutOfBoundsException,
                self.StaleElementReferenceException,
                self.WebDriverException,
            ):
                continue

        raise AutomationError("download_button_not_found", "未找到“下载装箱清单”按钮")

    def _choose_export_without_images(self) -> None:
        deadline = time.time() + DEFAULT_MODAL_TIMEOUT
        while time.time() < deadline:
            modal = self._find_first_visible(
                [
                    (self.By.XPATH, "//*[@role='dialog' and .//*[contains(normalize-space(.), '导出装箱清单')]]"),
                    (self.By.XPATH, "//*[contains(@class, 'dialog') and .//*[contains(normalize-space(.), '导出装箱清单')]]"),
                    (self.By.XPATH, "//*[contains(@class, 'modal') and .//*[contains(normalize-space(.), '导出装箱清单')]]"),
                ],
                timeout=0,
            )
            if modal is not None:
                radio = self._find_first_visible(
                    [
                        (self.By.XPATH, ".//*[contains(@class, 'el-radio') and contains(normalize-space(.), '导出不包含图片')]"),
                        (self.By.XPATH, ".//*[contains(normalize-space(.), '导出不包含图片') and (self::label or self::span or self::div)]"),
                    ],
                    root=modal,
                    timeout=3,
                )
                if radio is not None:
                    self._safe_click(radio)
                else:
                    self._click_text("导出不包含图片", exact=False, timeout=6, container=modal)

                confirm = self._find_dom_clickable(
                    timeout=6,
                    container=modal,
                    text="确定",
                    exact_text=True,
                    class_contains="el-button--primary",
                )
                if confirm is not None:
                    self._safe_click(confirm)
                else:
                    self._click_text("确定", exact=True, timeout=6, container=modal)
                return
            time.sleep(0.3)
        raise AutomationError("export_modal_not_found", "未弹出“导出装箱清单”窗口")

    def set_download_dir(self, download_dir: Path) -> None:
        download_dir.mkdir(parents=True, exist_ok=True)
        self.driver.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": str(download_dir),
            },
        )

    def download_for_fba(self, fba_code: str, download_dir: Path, screenshot_dir: Path) -> dict[str, Any]:
        self.current_screenshot_dir = screenshot_dir
        self.set_download_dir(download_dir)
        self.search_shipment(fba_code)
        self.open_shipment_detail(fba_code)

        cards = self._find_shipment_cards()
        if not cards:
            raise AutomationError("shipment_cards_not_found", f"FBA {fba_code} 页面中未识别到仓库卡片")

        downloaded_files: list[dict[str, Any]] = []
        for index, card in enumerate(cards, start=1):
            card_text = normalize_text(card.text)
            warehouse_code = self._extract_warehouse_code(card_text, index)
            previous_names = {path.name for path in download_dir.iterdir()}
            self._click_download_button_in_card(card)
            self._choose_export_without_images()
            downloaded_path = self._wait_for_download_file(download_dir, previous_names)
            target_path = download_dir / build_download_filename(
                fba_code,
                index,
                warehouse_code,
                downloaded_path.name,
                downloaded_path.suffix or ".xlsx",
            )
            if target_path.exists():
                target_path = download_dir / f"{target_path.stem}_{build_timestamp()}{target_path.suffix}"
            downloaded_path.rename(target_path)
            downloaded_files.append(
                {
                    "sequence": index,
                    "warehouse_code": warehouse_code,
                    "source_name": downloaded_path.name,
                    "saved_name": target_path.name,
                    "saved_path": str(target_path),
                }
            )

        return {
            "warehouse_count": len(cards),
            "downloaded_files": downloaded_files,
            **self.current_page_state(),
        }


class LingxingPlaywrightAutomation:
    def __init__(self, profile_dir: Path, credentials: LoginCredentials | None):
        bindings = import_playwright_bindings()
        self.sync_playwright = bindings["sync_playwright"]
        self.PlaywrightTimeoutError = bindings["PlaywrightTimeoutError"]
        self.profile_dir = profile_dir
        self.credentials = credentials
        self.headless = env_flag("PLAYWRIGHT_HEADLESS", False)
        self.block_static_resources = env_flag("PLAYWRIGHT_BLOCK_STATIC_RESOURCES", False)
        self.chrome_binary = locate_chrome_binary()
        self.playwright_manager = None
        self.context = None
        self.page = None
        self.current_screenshot_dir: Path | None = None
        self.api_headers_template: dict[str, str] | None = None

    def start(self) -> None:
        self.profile_dir.mkdir(parents=True, exist_ok=True)
        launch_args = ["--disable-popup-blocking", "--disable-notifications", "--disable-dev-shm-usage"]
        if self.headless:
            launch_args.append("--window-size=1600,900")
        else:
            launch_args.append("--start-maximized")
        if os.name != "nt":
            launch_args.append("--no-sandbox")
        launch_kwargs: dict[str, Any] = {
            "user_data_dir": str(self.profile_dir),
            "headless": self.headless,
            "args": launch_args,
            "viewport": {"width": 1600, "height": 900} if self.headless else None,
            "accept_downloads": True,
        }
        if self.chrome_binary:
            launch_kwargs["executable_path"] = str(self.chrome_binary)

        try:
            self.playwright_manager = self.sync_playwright().start()
            self.context = self.playwright_manager.chromium.launch_persistent_context(**launch_kwargs)
            if self.block_static_resources:
                self.context.route("**/*", self._route_static_resources)
            self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
        except Exception as exc:
            if self.context is not None:
                try:
                    self.context.close()
                except Exception:
                    pass
                self.context = None
            if self.playwright_manager is not None:
                try:
                    self.playwright_manager.stop()
                except Exception:
                    pass
                self.playwright_manager = None
            raise AutomationError("browser_start_failed", f"启动 Playwright Chrome 失败：{exc}") from exc

    def _route_static_resources(self, route: Any, request: Any) -> None:
        try:
            if request.resource_type in BLOCKED_RESOURCE_TYPES:
                route.abort()
                return
        except Exception:
            pass
        route.continue_()

    def close(self) -> None:
        if self.context is not None:
            try:
                self.context.close()
            finally:
                self.context = None
        if self.playwright_manager is not None:
            try:
                self.playwright_manager.stop()
            finally:
                self.playwright_manager = None
        self.page = None

    def capture_screenshot(self, name: str) -> Path | None:
        if self.page is None or self.current_screenshot_dir is None:
            return None
        self.current_screenshot_dir.mkdir(parents=True, exist_ok=True)
        screenshot_path = self.current_screenshot_dir / f"{sanitize_filename_part(name)}.png"
        try:
            self.page.screenshot(path=str(screenshot_path), full_page=True)
            return screenshot_path
        except Exception:
            return None

    def current_page_state(self) -> dict[str, Any]:
        if self.page is None:
            return {"current_url": None, "page_title": None}
        try:
            return {
                "current_url": self.page.url,
                "page_title": self.page.title(),
            }
        except Exception:
            return {"current_url": None, "page_title": None}

    def _page_text(self) -> str:
        if self.page is None:
            return ""
        try:
            return normalize_text(self.page.locator("body").inner_text(timeout=2000))
        except Exception:
            return ""

    def _page_contains_any(self, texts: list[str]) -> bool:
        page_text = self._page_text()
        return any(text in page_text for text in texts)

    def _wait_for_url_contains(self, text: str, timeout: int = DEFAULT_PAGE_TIMEOUT) -> bool:
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self.page is not None and text in self.page.url:
                return True
            if self.page is not None:
                self.page.wait_for_timeout(300)
            else:
                time.sleep(0.3)
        return False

    def _wait_for_url_contains_any(self, texts: list[str], timeout: int = DEFAULT_PAGE_TIMEOUT) -> bool:
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self.page is not None and any(text in self.page.url for text in texts):
                return True
            if self.page is not None:
                self.page.wait_for_timeout(300)
            else:
                time.sleep(0.3)
        return False

    def _has_login_fields(self) -> bool:
        if self.page is None:
            return False
        try:
            return self.page.locator("input[name='account']").count() > 0 and self.page.locator("input[name='pwd']").count() > 0
        except Exception:
            return False

    def _has_app_shell(self) -> bool:
        if self.page is None:
            return False
        try:
            if self.page.locator("li.el-menu-item:has(i.lx_nav_fba)").count() > 0:
                return True
            if self.page.locator("div.search-input input.el-input__inner:not([readonly])").count() > 0:
                return True
        except Exception:
            return False
        return False

    def _is_lingxing_app_url(self) -> bool:
        return bool(self.page is not None and "/erp/" in self.page.url and "/login" not in self.page.url)

    def _is_fba_cargo_ready(self) -> bool:
        if self.page is None:
            return False
        try:
            return (
                "/erp/msupply/fbaCargo" in self.page.url
                and self.page.locator("div.search-input input.el-input__inner:not([readonly])").count() > 0
            )
        except Exception:
            return False

    def _looks_like_blank_page(self) -> bool:
        if self.page is None:
            return False
        try:
            body_text = self.page.locator("body").inner_text(timeout=1500)
        except Exception:
            return False
        return normalize_text(body_text) == ""

    def _wait_for_login_or_app_shell(self, timeout: int = DEFAULT_PAGE_TIMEOUT) -> str:
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self._has_login_fields():
                return "login"
            if self._has_app_shell():
                return "app"
            if self.page is not None:
                self.page.wait_for_timeout(400)
            else:
                time.sleep(0.4)
        if self.page is not None and "/login" in self.page.url:
            return "login_pending"
        return "unknown"

    def ensure_logged_in(self) -> None:
        if self._is_lingxing_app_url():
            if self._has_app_shell():
                return
            state = self._wait_for_login_or_app_shell(timeout=3)
            if state == "app":
                return

        state = "unknown"
        for attempt in range(3):
            self.page.goto(LINGXING_HOME_URL, wait_until="domcontentloaded", timeout=120000)
            state = self._wait_for_login_or_app_shell(timeout=20 if attempt == 0 else 12)
            if state in {"app", "login", "login_pending"}:
                break
            if self._looks_like_blank_page():
                self.page.wait_for_timeout(1200)
                try:
                    self.page.reload(wait_until="domcontentloaded", timeout=120000)
                except Exception:
                    pass
                state = self._wait_for_login_or_app_shell(timeout=12)
                if state in {"app", "login", "login_pending"}:
                    break

        if state == "app":
            return

        if self.credentials is None:
            raise AutomationError("credentials_missing", "仍在登录页，且未提供登录配置文件(username/password)")

        if state == "login_pending":
            state = self._wait_for_login_or_app_shell(timeout=30)
            if state == "app":
                return

        if not self._has_login_fields():
            self.capture_screenshot("home_or_login_page_not_ready")
            raise AutomationError("home_or_login_page_not_ready", "领星页面没有加载完成，未识别到系统菜单或登录输入框")

        self.page.locator("input[name='account']").first.fill(self.credentials.username)
        self.page.locator("input[name='pwd']").first.fill(self.credentials.password)

        login_button = self.page.locator(
            "button.loginBtn, button:has-text('登录'), button:has-text('Login')"
        ).first
        if login_button.count() > 0:
            login_button.click(timeout=20000)
        else:
            self.page.locator("input[name='pwd']").first.press("Enter")

        deadline = time.time() + max(DEFAULT_LOGIN_TIMEOUT, 35)
        while time.time() < deadline:
            if self._has_app_shell():
                return
            if "/login" not in self.page.url and not self._has_login_fields():
                self.page.wait_for_timeout(1000)
                if self._has_app_shell():
                    return
            self.page.wait_for_timeout(500)

        self.capture_screenshot("login_failed_still_on_login_page")
        raise AutomationError("login_failed", "登录后仍停留在登录页，请确认账号密码或站点状态")

    def open_fba_shipments_page(self) -> None:
        self.ensure_logged_in()
        if self._is_fba_cargo_ready():
            return

        self.page.goto(LINGXING_FBA_CARGO_URL, wait_until="domcontentloaded", timeout=120000)
        deadline = time.time() + 18
        while time.time() < deadline:
            if self._is_fba_cargo_ready():
                return
            if self._has_login_fields() or "/login" in self.page.url:
                self.ensure_logged_in()
                self.page.goto(LINGXING_FBA_CARGO_URL, wait_until="domcontentloaded", timeout=120000)
                deadline = time.time() + 18
                continue
            self.page.wait_for_timeout(300)

        self.page.goto(LINGXING_HOME_URL, wait_until="domcontentloaded", timeout=120000)
        state = self._wait_for_login_or_app_shell(timeout=20)
        if state in {"login", "login_pending"}:
            self.ensure_logged_in()
        if not self._has_app_shell():
            self.capture_screenshot("home_app_shell_not_ready")
            raise AutomationError("home_app_shell_not_ready", "登录后未识别到领星系统菜单")
        self.page.locator("li.el-menu-item:has(i.lx_nav_fba)").first.click(timeout=20000)
        self.page.wait_for_timeout(600)

        submenu_anchor = self.page.locator("div.submenu-content a[href='/erp/msupply/fbaCargo']").first
        if submenu_anchor.count() > 0:
            submenu_anchor.click(timeout=10000, force=True)
        else:
            self.page.goto(LINGXING_FBA_CARGO_URL, wait_until="domcontentloaded", timeout=120000)

        deadline = time.time() + DEFAULT_PAGE_TIMEOUT
        while time.time() < deadline:
            if "/erp/msupply/fbaCargo" in self.page.url and self.page.locator("div.search-input input.el-input__inner:not([readonly])").count() > 0:
                return
            self.page.wait_for_timeout(400)

        self.page.goto(LINGXING_FBA_CARGO_URL, wait_until="domcontentloaded", timeout=120000)
        self.page.wait_for_timeout(2000)
        if self.page.locator("div.search-input input.el-input__inner:not([readonly])").count() == 0:
            raise AutomationError("fba_shipments_page_not_ready", "未能进入 FBA货件 页面")

    def _find_exact_shipment_matches(self, fba_code: str) -> list[Any]:
        matches: list[Any] = []
        normalized_fba = normalize_text(fba_code)
        selectors = [
            "div.oneLine.ak-blue-pointer",
            ".oneLine.ak-blue-pointer",
            ".ak-blue-pointer",
            "a",
            "span",
            "td",
        ]
        seen: set[str] = set()
        for selector in selectors:
            rows = self.page.locator(selector).filter(has_text=normalized_fba)
            try:
                row_count = min(rows.count(), 30)
            except Exception:
                continue
            for index in range(row_count):
                row = rows.nth(index)
                try:
                    if not row.is_visible(timeout=500):
                        continue
                    text = normalize_text(row.inner_text(timeout=1000))
                except Exception:
                    continue
                if text != normalized_fba:
                    continue
                identity = f"{selector}:{index}:{text}"
                if identity in seen:
                    continue
                seen.add(identity)
                matches.append(row)
        return matches

    def _has_exact_shipment_candidate(self, fba_code: str) -> bool:
        script = """
        (fbaCode) => {
          const expected = (fbaCode || '').replace(/\\s+/g, ' ').trim();
          const isVisible = (node) => {
            if (!node) return false;
            const style = window.getComputedStyle(node);
            if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
            const rect = node.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
          };
          const selectors = [
            'div.oneLine.ak-blue-pointer',
            '.oneLine.ak-blue-pointer',
            '.ak-blue-pointer',
            'a',
            'span',
            'td'
          ];
          for (const selector of selectors) {
            for (const node of document.querySelectorAll(selector)) {
              const text = (node.innerText || node.textContent || '').replace(/\\s+/g, ' ').trim();
              if (text === expected && isVisible(node)) return true;
            }
          }
          return false;
        }
        """
        try:
            return bool(self.page.evaluate(script, fba_code))
        except Exception:
            return False

    def _wait_for_shipment_detail_page(self, timeout: int = DEFAULT_PAGE_TIMEOUT) -> bool:
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self.page is not None and any(text in self.page.url for text in ["SendToAmazonDetail", "ShipmentDetail"]):
                detail_ready_deadline = time.time() + 2
                while time.time() < detail_ready_deadline:
                    if (
                        self.page.locator("div.delivery-ship").count() > 0
                        or self.page.get_by_text("Box Labels", exact=True).count() > 0
                        or self.page.get_by_text("箱子标签", exact=True).count() > 0
                    ):
                        return True
                    self.page.wait_for_timeout(150)
                return True
            if self.page is not None:
                body_text = self._page_text()
                if "箱子标签" in body_text or "Box Labels" in body_text:
                    return True
                self.page.wait_for_timeout(300)
            else:
                time.sleep(0.3)
        return False

    def _dispatch_shipment_dom_click(self, fba_code: str) -> bool:
        script = """
        (fbaCode) => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const expected = normalize(fbaCode);
          const selectors = [
            'div.oneLine.ak-blue-pointer',
            '.oneLine.ak-blue-pointer',
            '.ak-blue-pointer',
            'a',
            'span',
            'div'
          ];
          const seen = new Set();
          for (const selector of selectors) {
            for (const node of document.querySelectorAll(selector)) {
              if (seen.has(node)) continue;
              seen.add(node);
              if (normalize(node.innerText || node.textContent) !== expected) continue;
              const targets = [
                node,
                node.closest('a'),
                node.closest('.ak-blue-pointer'),
                node.closest('td'),
                node.parentElement
              ].filter(Boolean);
              for (const target of targets) {
                target.scrollIntoView({ block: 'center', inline: 'center' });
                for (const eventName of ['mouseover', 'mousedown', 'mouseup', 'click']) {
                  target.dispatchEvent(new MouseEvent(eventName, {
                    bubbles: true,
                    cancelable: true,
                    view: window
                  }));
                }
                if (typeof target.click === 'function') target.click();
              }
              return true;
            }
          }
          return false;
        }
        """
        try:
            return bool(self.page.evaluate(script, fba_code))
        except Exception:
            return False

    def _activate_shipment_match(self, match: Any, fba_code: str) -> bool:
        click_attempts = [
            lambda: match.click(timeout=5000),
            lambda: match.click(timeout=5000, force=True),
            lambda: match.dblclick(timeout=5000, force=True),
            lambda: match.press("Enter", timeout=3000),
            lambda: self._dispatch_shipment_dom_click(fba_code),
        ]
        for attempt in click_attempts:
            try:
                if self.context is not None:
                    try:
                        with self.context.expect_page(timeout=2500) as popup_info:
                            result = attempt()
                        new_page = popup_info.value
                        new_page.wait_for_load_state("domcontentloaded", timeout=30000)
                        self.page = new_page
                    except Exception:
                        result = attempt()
                else:
                    result = attempt()
                if result is False:
                    continue
            except Exception:
                continue
            if self._wait_for_shipment_detail_page(timeout=8):
                return True
        return False

    def search_shipment(self, fba_code: str) -> None:
        self.open_fba_shipments_page()
        query_input = self.page.locator(
            "div.search-input input.el-input__inner:not([readonly]):visible, "
            ".search-input input:not([readonly]):visible, "
            "input.el-input__inner:not([readonly]):visible"
        ).first
        try:
            query_input.scroll_into_view_if_needed(timeout=5000)
        except Exception:
            pass
        try:
            query_input.click(timeout=10000, force=True)
            query_input.fill("", timeout=10000)
            query_input.fill(fba_code, timeout=10000)
        except Exception:
            try:
                query_input.press("Control+A", timeout=5000)
                query_input.type(fba_code, timeout=20000)
            except Exception:
                query_input.evaluate(
                    "(node, value) => { node.focus(); node.value = value; node.dispatchEvent(new Event('input', { bubbles: true })); node.dispatchEvent(new Event('change', { bubbles: true })); }",
                    fba_code,
                )

        try:
            current_value = normalize_text(query_input.input_value(timeout=3000))
        except Exception:
            current_value = ""
        if current_value != normalize_text(fba_code):
            query_input.evaluate(
                "(node, value) => { node.focus(); node.value = value; node.dispatchEvent(new Event('input', { bubbles: true })); node.dispatchEvent(new Event('change', { bubbles: true })); }",
                fba_code,
            )

        search_icon = self.page.locator("i.lx_combo_search:visible").first
        if search_icon.count() > 0:
            search_icon.click(timeout=20000)
        else:
            try:
                query_input.press("Enter", timeout=5000)
            except Exception:
                pass

        deadline = time.time() + DEFAULT_SEARCH_TIMEOUT
        while time.time() < deadline:
            if self._has_exact_shipment_candidate(fba_code):
                return
            page_text = self._page_text()
            if any(text in page_text for text in ["暂无数据", "没有找到", "无数据", "No Data", "No data"]):
                raise AutomationError("shipment_not_found", f"未搜索到 FBA {fba_code}")
            self.page.wait_for_timeout(500)

        raise AutomationError("shipment_search_timeout", f"搜索 FBA {fba_code} 超时")

    def open_shipment_detail(self, fba_code: str) -> None:
        if self._dispatch_shipment_dom_click(fba_code) and self._wait_for_shipment_detail_page(timeout=10):
            return

        matches = self._find_exact_shipment_matches(fba_code)
        if not matches:
            raise AutomationError("shipment_not_found", f"未找到 FBA {fba_code} 的精确结果")

        for match in matches:
            if self._activate_shipment_match(match, fba_code):
                return
        raise AutomationError("shipment_detail_timeout", f"进入 FBA {fba_code} 详情页超时")

    def _detect_shipment_stage(self) -> str:
        page_text = self._page_text()
        stage_patterns = [
            ("Review/check Shipment Content", "Review/check Shipment Content"),
            ("选择发货商品", "选择发货商品"),
            ("Transportation Service", "Transportation Service"),
            ("配送服务", "配送服务"),
            ("Shipment Packaging", "Shipment Packaging"),
            ("商品装箱", "商品装箱"),
            ("Box Labels", "Box Labels"),
            ("箱子标签", "箱子标签"),
            ("Shipment Tracking", "Shipment Tracking"),
            ("货件追踪", "货件追踪"),
        ]
        for keyword, label in stage_patterns:
            if keyword in page_text:
                return label
        return "未知步骤"

    def _try_open_box_labels_step(self) -> None:
        candidates = [
            self.page.get_by_text("Box Labels", exact=True),
            self.page.get_by_text("箱子标签", exact=True),
            self.page.locator("text=Box Labels"),
            self.page.locator("text=箱子标签"),
        ]
        for locator in candidates:
            try:
                if locator.count() == 0:
                    continue
                target = locator.last
                target.scroll_into_view_if_needed(timeout=2000)
                target.click(timeout=3000, force=True)
                self.page.wait_for_timeout(700)
                return
            except Exception:
                continue

    def ensure_box_labels_ready(self, fba_code: str) -> None:
        if self._find_shipment_cards():
            return

        self._try_open_box_labels_step()
        deadline = time.time() + 10
        while time.time() < deadline:
            if self._find_shipment_cards():
                return
            self.page.wait_for_timeout(500)

        stage = self._detect_shipment_stage()
        if stage in {"Review/check Shipment Content", "选择发货商品", "Transportation Service", "配送服务", "Shipment Packaging", "商品装箱"}:
            raise AutomationError(
                "shipment_not_ready_for_box_labels",
                f"FBA {fba_code} 已进入详情页，但当前停在“{stage}”，尚未进入 Box Labels/箱子标签步骤，不能下载装箱清单。",
            )
        raise AutomationError(
            "shipment_cards_not_found",
            f"FBA {fba_code} 已进入详情页，但未识别到 Box Labels/箱子标签中的仓库卡片或下载按钮。",
        )

    def _find_shipment_cards(self) -> list[Any]:
        cards = self.page.locator("div.delivery-ship")
        matched_cards: list[Any] = []
        for index in range(cards.count()):
            card = cards.nth(index)
            try:
                text = normalize_text(card.inner_text(timeout=2000))
            except Exception:
                continue
            if "Reference ID" not in text:
                continue
            if "货件单号" not in text and "FBA" not in text:
                continue
            if card.locator("button:has(i.lx_img_download)").count() == 0:
                continue
            matched_cards.append(card)
        return matched_cards

    def _shipment_card_signature(self, card_text: str, index: int) -> str:
        normalized = normalize_text(card_text)
        reference_match = re.search(r"Reference\s*ID\s*[:：]?\s*([A-Z0-9-]+)", normalized, flags=re.IGNORECASE)
        fba_match = re.search(r"\bFBA[A-Z0-9-]+\b", normalized, flags=re.IGNORECASE)
        warehouse_code = self._extract_warehouse_code(normalized, index)
        parts = [
            warehouse_code,
            reference_match.group(1).upper() if reference_match else "",
            fba_match.group(0).upper() if fba_match else "",
        ]
        signature = "|".join(part for part in parts if part)
        return signature or normalized[:160]

    def _reset_shipment_card_scroll(self) -> None:
        script = """
        () => {
          const cards = Array.from(document.querySelectorAll('div.delivery-ship'));
          const containers = new Set();
          const addScrollableParents = (node) => {
            let current = node ? node.parentElement : null;
            while (current && current !== document.body) {
              if (current.scrollWidth > current.clientWidth + 20) {
                containers.add(current);
              }
              current = current.parentElement;
            }
          };
          cards.forEach(addScrollableParents);
          containers.forEach((node) => { node.scrollLeft = 0; });
          window.scrollTo(0, window.scrollY);
          return containers.size;
        }
        """
        try:
            self.page.evaluate(script)
            self.page.wait_for_timeout(250)
        except Exception:
            pass

    def _scroll_shipment_cards_right(self) -> bool:
        script = """
        () => {
          const cards = Array.from(document.querySelectorAll('div.delivery-ship'));
          const containers = new Set();
          const addScrollableParents = (node) => {
            let current = node ? node.parentElement : null;
            while (current && current !== document.body) {
              if (current.scrollWidth > current.clientWidth + 20) {
                containers.add(current);
              }
              current = current.parentElement;
            }
          };
          cards.forEach(addScrollableParents);
          const states = [];
          containers.forEach((node) => {
            const before = node.scrollLeft;
            const step = Math.max(360, Math.floor(node.clientWidth * 0.75));
            node.scrollLeft = Math.min(node.scrollLeft + step, node.scrollWidth);
            states.push(node.scrollLeft > before + 5);
          });
          const beforeWindow = window.scrollX;
          window.scrollBy({ left: Math.floor(window.innerWidth * 0.75), top: 0, behavior: 'auto' });
          states.push(window.scrollX > beforeWindow + 5);
          return states.some(Boolean);
        }
        """
        try:
            moved = bool(self.page.evaluate(script))
            self.page.wait_for_timeout(450)
            return moved
        except Exception:
            return False

    def _download_card_packing_list(
        self,
        *,
        card: Any,
        fba_code: str,
        index: int,
        warehouse_code: str,
        download_dir: Path,
        defer_raw_download: bool = False,
    ) -> dict[str, Any]:
        response = None
        download_source = "browser_download"
        try:
            with self.page.expect_response(
                lambda resp: "exportPackingListV2" in resp.url and resp.status == 200,
                timeout=DEFAULT_DOWNLOAD_TIMEOUT * 1000,
            ) as response_info:
                self._click_download_button_in_card(card)
                self._choose_export_without_images()
            response = response_info.value
        except Exception:
            response = None

        if response is None:
            with self.page.expect_download(timeout=DEFAULT_DOWNLOAD_TIMEOUT * 1000) as download_info:
                self._click_download_button_in_card(card)
                self._choose_export_without_images()
            download = download_info.value
            suggested_name = download.suggested_filename or f"{fba_code}_{index}.xlsx"
            suffix = Path(suggested_name).suffix or ".xlsx"
            target_path = download_dir / build_download_filename(fba_code, index, warehouse_code, suggested_name, suffix)
            if target_path.exists():
                target_path = download_dir / f"{target_path.stem}_{build_timestamp()}{target_path.suffix}"
            download.save_as(str(target_path))
        else:
            suggested_name = filename_from_content_disposition(response.headers.get("content-disposition")) or "packing_list.xlsx"
            download_source = "playwright_response"
            try:
                body = response.body()
            except Exception:
                body = b""
            if not is_valid_xlsx_payload(body):
                raw_request = self._build_export_raw_request(response)
                if defer_raw_download:
                    return {
                        "sequence": index,
                        "warehouse_code": warehouse_code,
                        "source_name": suggested_name,
                        "saved_name": None,
                        "saved_path": None,
                        "download_source": "raw_request_pending",
                        "_raw_request": raw_request,
                        "_fba_code": fba_code,
                        "_download_dir": str(download_dir),
                    }
                suggested_name, body = self._download_export_request_raw(raw_request)
                download_source = "raw_request"
            suffix = Path(suggested_name).suffix or ".xlsx"
            target_path = download_dir / build_download_filename(fba_code, index, warehouse_code, suggested_name, suffix)
            if target_path.exists():
                target_path = download_dir / f"{target_path.stem}_{build_timestamp()}{target_path.suffix}"
            if not is_valid_xlsx_payload(body):
                raise AutomationError("export_response_not_excel", "导出接口返回的不是 Excel 文件")
            target_path.write_bytes(body)

        return {
            "sequence": index,
            "warehouse_code": warehouse_code,
            "source_name": suggested_name,
            "saved_name": target_path.name,
            "saved_path": str(target_path),
            "download_source": download_source,
        }

    def _extract_warehouse_code(self, card_text: str, index: int) -> str:
        normalized = normalize_text(card_text)
        fc_code_match = re.search(r"FC\s*Code\s*[:：]?\s*([A-Z0-9-]+)", normalized, flags=re.IGNORECASE)
        if fc_code_match:
            return sanitize_filename_part(fc_code_match.group(1).upper())

        logistics_match = re.search(r"(?:物流中心编码|仓库编码)\s*[:：]?\s*([A-Z0-9-]+)", normalized)
        if logistics_match:
            return sanitize_filename_part(logistics_match.group(1).upper())

        header_match = re.search(r"AGL-([A-Z0-9]+)-", normalized)
        if header_match:
            return sanitize_filename_part(header_match.group(1).upper())

        generic_match = re.search(r"\b[A-Z]{3}\d\b", normalized)
        if generic_match:
            return sanitize_filename_part(generic_match.group(0).upper())

        return f"WAREHOUSE{index:02d}"

    def _click_download_button_in_card(self, card) -> None:
        button = card.locator(
            "button.el-button.btn-style.el-button--text.el-button--mini.is-round.is-icon:has(i.lx_img_download)"
        ).first
        if button.count() == 0:
            button = card.locator("button:has(i.lx_img_download)").first
        if button.count() == 0:
            raise AutomationError("download_button_not_found", "未找到下载装箱清单按钮")
        button.click(timeout=20000)

    def _choose_export_without_images(self) -> None:
        deadline = time.time() + DEFAULT_MODAL_TIMEOUT
        modal_title_pattern = re.compile(r"(导出装箱清单|Export Packing List|装箱清单)", re.IGNORECASE)
        without_images_pattern = re.compile(r"(导出不包含图片|Export Without Images)", re.IGNORECASE)
        confirm_pattern = re.compile(r"(确\s*定|Confirm)", re.IGNORECASE)
        while time.time() < deadline:
            modal_locator = self.page.locator(
                ".el-dialog:visible, .el-dialog__wrapper:visible, [role='dialog']:visible, .dialog:visible, .modal:visible"
            ).filter(has_text=modal_title_pattern)
            modal_count = modal_locator.count()
            if modal_count > 0:
                modal = modal_locator.nth(modal_count - 1)
                option = modal.locator("label.el-radio, .el-radio, span, div").filter(has_text=without_images_pattern)
                if option.count() > 0:
                    option.first.click(timeout=5000)
                else:
                    page_option = self.page.locator("label.el-radio, .el-radio, span, div").filter(has_text=without_images_pattern)
                    if page_option.count() > 0:
                        page_option.first.click(timeout=5000)

                confirm = modal.locator("button.el-button.el-button--primary.el-button--small.is-round").filter(
                    has_text=confirm_pattern
                )
                if confirm.count() == 0:
                    confirm = modal.locator("button.el-button--primary").filter(has_text=confirm_pattern)
                if confirm.count() == 0:
                    confirm = self.page.locator("button.el-button--primary").filter(has_text=confirm_pattern)
                if confirm.count() > 0:
                    confirm.first.click(timeout=10000)
                    return
            self.page.wait_for_timeout(300)
        raise AutomationError("export_modal_not_found", "未弹出导出装箱清单窗口")

    def _build_export_raw_request(self, response: Any) -> RawExportRequest:
        request = response.request
        headers = {
            key: value
            for key, value in request.headers.items()
            if key.lower() not in {"content-length", "cookie", "host", "accept-encoding"}
        }
        cookies = self.context.cookies() if self.context is not None else []
        cookie_header = "; ".join(
            f"{cookie['name']}={cookie['value']}"
            for cookie in cookies
            if "lingxing" in cookie.get("domain", "") or "lingxingerp" in cookie.get("domain", "")
        )
        if cookie_header:
            headers["Cookie"] = cookie_header

        post_data = request.post_data
        data = post_data.encode("utf-8") if isinstance(post_data, str) else post_data
        suggested_name = filename_from_content_disposition(response.headers.get("content-disposition"))
        return RawExportRequest(
            url=request.url,
            method=request.method,
            headers=headers,
            data=data,
            suggested_name=suggested_name,
        )

    def _download_export_request_raw(self, export_request: RawExportRequest) -> tuple[str, bytes]:
        last_error: Exception | None = None
        for attempt in range(3):
            raw_request = urllib.request.Request(
                export_request.url,
                data=export_request.data,
                headers=export_request.headers,
                method=export_request.method,
            )
            try:
                with urllib.request.urlopen(raw_request, timeout=DEFAULT_DOWNLOAD_TIMEOUT) as raw_response:
                    body = raw_response.read()
                    suggested_name = (
                        filename_from_content_disposition(raw_response.headers.get("content-disposition"))
                        or export_request.suggested_name
                    )
                if is_valid_xlsx_payload(body):
                    return suggested_name or "packing_list.xlsx", body
                last_error = AutomationError("export_response_not_excel", "导出接口返回的不是 Excel 文件")
            except Exception as exc:
                last_error = exc
            if attempt < 2:
                time.sleep(1.2 * (attempt + 1))

        if last_error is not None:
            raise last_error
        return export_request.suggested_name or "packing_list.xlsx", b""

    def _clone_api_headers(self, headers: dict[str, str]) -> dict[str, str]:
        output = {
            key: value
            for key, value in headers.items()
            if key.lower() not in {"content-length", "accept-encoding"}
        }
        output["content-type"] = "application/json;charset=UTF-8"
        output["accept"] = "application/json, text/plain, */*"
        output["x-ak-request-id"] = str(uuid.uuid4())
        return output

    def _api_post_json(
        self,
        headers: dict[str, str],
        url: str,
        payload: dict[str, Any],
    ) -> tuple[dict[str, Any], float]:
        if self.context is None:
            raise AutomationError("api_context_missing", "浏览器接口上下文未初始化")
        start = time.perf_counter()
        response = self.context.request.post(
            url,
            data=json.dumps(payload, ensure_ascii=False),
            headers=self._clone_api_headers(headers),
            timeout=env_int("LINGXING_API_JSON_TIMEOUT_MS", 20000, minimum=5000, maximum=120000),
        )
        body = response.body()
        if not response.ok:
            raise AutomationError("api_http_error", f"领星接口 HTTP {response.status}：{url}")
        try:
            data = json.loads(body.decode("utf-8"))
        except Exception as exc:
            raise AutomationError("api_invalid_json", f"领星接口返回不是 JSON：{url}") from exc
        if data.get("code") not in (1, 200, "1", "200", None) and data.get("success") is not True:
            raise AutomationError("api_business_error", str(data.get("msg") or data.get("message") or "领星接口返回业务错误"))
        return data, elapsed_seconds(start)

    def _api_download_excel(
        self,
        headers: dict[str, str],
        url: str,
        payload: dict[str, Any],
    ) -> tuple[str, bytes, float]:
        if self.context is None:
            raise AutomationError("api_context_missing", "浏览器接口上下文未初始化")
        start = time.perf_counter()
        attempts = env_int("LINGXING_API_EXPORT_ATTEMPTS", 2, minimum=1, maximum=3)
        timeout_ms = env_int("LINGXING_API_EXPORT_TIMEOUT_MS", 12000, minimum=3000, maximum=120000)
        last_error: Exception | None = None
        for attempt in range(1, attempts + 1):
            try:
                response = self.context.request.post(
                    url,
                    data=json.dumps(payload, ensure_ascii=False),
                    headers=self._clone_api_headers(headers),
                    timeout=timeout_ms,
                )
                body = response.body()
                if not response.ok:
                    raise AutomationError("api_export_http_error", f"领星导出接口 HTTP {response.status}")
                if not is_valid_xlsx_payload(body):
                    preview = compact_response_preview(body)
                    raise AutomationError("api_export_not_excel", f"领星导出接口返回的不是有效 Excel：{preview}")
                suggested_name = filename_from_content_disposition(response.headers.get("content-disposition")) or "packing_list.xlsx"
                return suggested_name, body, elapsed_seconds(start)
            except Exception as exc:
                last_error = exc
                if attempt < attempts:
                    time.sleep(0.8 * attempt)
        if last_error is not None:
            raise last_error
        raise AutomationError("api_export_not_excel", "领星导出接口返回的不是有效 Excel")

    def _capture_api_headers(self) -> tuple[dict[str, str], float]:
        if self.page is None:
            raise AutomationError("api_page_missing", "浏览器页面未初始化")
        captured: dict[str, str] = {}

        def on_request(request: Any) -> None:
            if captured:
                return
            if "/api/fba_shipment/showShipment_v2" not in request.url:
                return
            captured.update(request.headers)

        start = time.perf_counter()
        self.page.on("request", on_request)
        self.open_fba_shipments_page()
        deadline = time.time() + 12
        while time.time() < deadline and not captured:
            self.page.wait_for_timeout(200)
        if not captured:
            raise AutomationError("api_headers_not_captured", "未能捕获领星接口请求头")
        self.api_headers_template = captured
        return captured, elapsed_seconds(start)

    def _ensure_api_headers(self) -> tuple[dict[str, str], float]:
        if self.api_headers_template:
            return self.api_headers_template, 0.0
        return self._capture_api_headers()

    def _build_api_search_payload(self, fba_code: str) -> dict[str, Any]:
        end_date = beijing_now().date()
        start_date = end_date - timedelta(days=90)
        return {
            "search_field_time": "create_date",
            "is_sta": "",
            "is_awd": "",
            "ship_mode": "",
            "step": [],
            "is_closed": "",
            "application_diff": "",
            "received_diff": "",
            "application_received_diff": "",
            "is_relate_packing_task_sn": "",
            "is_add_tracking": "",
            "delivery_order_status": [],
            "box_type": "",
            "is_uploaded_box": "",
            "sta_transportation_mode": "",
            "delivery_mode": "",
            "carrier_type": "",
            "create_uids": [],
            "principal_uids": [],
            "is_store_diff": "",
            "search_field": "shipment_id",
            "search_value": fba_code,
            "shipment_status": [],
            "is_relate_shipment": "",
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "seniorSearchList": [],
            "shipment_type": [],
            "offset": 0,
            "length": 20,
            "req_time_sequence": "/api/fba_shipment/showShipment_v2$$1",
        }

    def _api_search_shipment(self, headers: dict[str, str], fba_code: str) -> tuple[dict[str, Any], float]:
        payload = self._build_api_search_payload(fba_code)
        data, seconds = self._api_post_json(
            headers,
            f"{LINGXING_ERP_BASE_URL}/api/fba_shipment/showShipment_v2",
            payload,
        )
        rows = ((data.get("data") or {}).get("list") or [])
        exact_rows = [row for row in rows if str(row.get("shipment_id", "")).upper() == fba_code.upper()]
        if len(exact_rows) != 1:
            raise AutomationError("api_shipment_not_unique", f"接口查询 FBA {fba_code} 返回 {len(exact_rows)} 条精确结果")
        return exact_rows[0], seconds

    def _api_get_plan_detail(
        self,
        headers: dict[str, str],
        local_task_id: str,
    ) -> tuple[dict[str, Any], float]:
        payload = {
            "localTaskId": local_task_id,
            "req_time_sequence": "/amz-sta-server/inbound-plan/detail$$1",
        }
        data, seconds = self._api_post_json(
            headers,
            f"{LINGXING_GW_BASE_URL}/amz-sta-server/inbound-plan/detail",
            payload,
        )
        detail = data.get("data") or {}
        if not detail.get("inboundPlanId"):
            raise AutomationError("api_inbound_plan_missing", "接口未返回 inboundPlanId")
        return detail, seconds

    def _api_get_label_page(
        self,
        headers: dict[str, str],
        *,
        sid: int,
        inbound_plan_id: str,
    ) -> tuple[list[dict[str, Any]], float]:
        payload = {
            "inboundPlanId": inbound_plan_id,
            "sid": sid,
            "req_time_sequence": "/amz-sta-server/inbound-shipment/shipmentLabelPage$$1",
        }
        data, seconds = self._api_post_json(
            headers,
            f"{LINGXING_GW_BASE_URL}/amz-sta-server/inbound-shipment/shipmentLabelPage",
            payload,
        )
        shipments = data.get("data") or []
        if not shipments:
            raise AutomationError("api_label_page_empty", "接口未返回箱子标签仓库数据")
        return shipments, seconds

    def _download_for_fba_via_api(self, fba_code: str, download_dir: Path, screenshot_dir: Path) -> dict[str, Any]:
        overall_start = time.perf_counter()
        timings: dict[str, float] = {}
        self.current_screenshot_dir = screenshot_dir
        download_dir.mkdir(parents=True, exist_ok=True)

        stage_start = time.perf_counter()
        self.ensure_logged_in()
        timings["login_ready_seconds"] = elapsed_seconds(stage_start)

        api_headers, timings["api_header_capture_seconds"] = self._ensure_api_headers()
        shipment, timings["api_search_seconds"] = self._api_search_shipment(api_headers, fba_code)
        local_task_id = str(shipment.get("local_sta_id") or "")
        if not local_task_id:
            raise AutomationError("api_local_task_id_missing", "接口结果没有 local_sta_id，无法走接口化下载")
        try:
            sid = int(shipment["sid"])
        except Exception as exc:
            raise AutomationError("api_sid_missing", "接口结果没有有效 sid") from exc

        plan_detail, timings["api_plan_detail_seconds"] = self._api_get_plan_detail(api_headers, local_task_id)
        label_shipments, timings["api_label_page_seconds"] = self._api_get_label_page(
            api_headers,
            sid=sid,
            inbound_plan_id=str(plan_detail["inboundPlanId"]),
        )

        export_results: list[dict[str, Any]] = []
        export_stage_start = time.perf_counter()
        for index, label_shipment in enumerate(label_shipments, start=1):
            shipment_id = label_shipment.get("shipmentId")
            if not shipment_id:
                raise AutomationError("api_label_shipment_id_missing", "接口箱子标签数据缺少 shipmentId")
            warehouse_code = sanitize_filename_part(str(label_shipment.get("warehouseId") or f"WAREHOUSE{index:02d}"))
            payload = {
                "isBatch": 0,
                "isPic": 0,
                "packingListBOList": [
                    {
                        "localTaskId": local_task_id,
                        "packingGroupId": None,
                        "shipmentId": shipment_id,
                    }
                ],
                "sid": sid,
            }
            suggested_name, body, seconds = self._api_download_excel(
                api_headers,
                f"{LINGXING_GW_BASE_URL}/amz-sta-server/inbound-packing/exportPackingListV2",
                payload,
            )
            export_results.append(
                {
                    "sequence": index,
                    "warehouse_code": warehouse_code,
                    "shipment_confirmation_id": label_shipment.get("shipmentConfirmationId"),
                    "shipment_id": shipment_id,
                    "source_name": suggested_name,
                    "body": body,
                    "api_export_seconds": seconds,
                }
            )
        timings["api_export_seconds"] = elapsed_seconds(export_stage_start)

        downloaded_files: list[dict[str, Any]] = []
        for item in export_results:
            suggested_name = str(item["source_name"])
            suffix = Path(suggested_name).suffix or ".xlsx"
            target_path = download_dir / build_download_filename(
                fba_code,
                int(item["sequence"]),
                str(item["warehouse_code"]),
                suggested_name,
                suffix,
            )
            if target_path.exists():
                target_path = download_dir / f"{target_path.stem}_{build_timestamp()}{target_path.suffix}"
            target_path.write_bytes(item["body"])
            downloaded_files.append(
                {
                    "sequence": item["sequence"],
                    "warehouse_code": item["warehouse_code"],
                    "shipment_confirmation_id": item.get("shipment_confirmation_id"),
                    "shipment_id": item.get("shipment_id"),
                    "source_name": suggested_name,
                    "saved_name": target_path.name,
                    "saved_path": str(target_path),
                    "download_source": "api_fast_path",
                    "duration_seconds": item["api_export_seconds"],
                }
            )

        return {
            "warehouse_count": len(downloaded_files),
            "downloaded_files": downloaded_files,
            "timings": timings,
            "download_strategy": "api_fast_path",
            "download_duration_seconds": elapsed_seconds(overall_start),
            **self.current_page_state(),
        }

    def _download_export_response_raw(self, response: Any) -> tuple[str, bytes]:
        export_request = self._build_export_raw_request(response)
        try:
            return self._download_export_request_raw(export_request)
        except Exception:
            body = response.body()
            return export_request.suggested_name or "packing_list.xlsx", body

    def _finalize_raw_download_metadata(
        self,
        *,
        pending_item: dict[str, Any],
        suggested_name: str,
        body: bytes,
        download_dir: Path,
        download_source: str,
    ) -> dict[str, Any]:
        if not is_valid_xlsx_payload(body):
            raise AutomationError("export_response_not_excel", "导出接口返回的不是 Excel 文件")
        suffix = Path(suggested_name).suffix or ".xlsx"
        target_path = download_dir / build_download_filename(
            pending_item["_fba_code"],
            int(pending_item["sequence"]),
            str(pending_item["warehouse_code"]),
            suggested_name,
            suffix,
        )
        if target_path.exists():
            target_path = download_dir / f"{target_path.stem}_{build_timestamp()}{target_path.suffix}"
        target_path.write_bytes(body)
        return {
            "sequence": pending_item["sequence"],
            "warehouse_code": pending_item["warehouse_code"],
            "source_name": suggested_name,
            "saved_name": target_path.name,
            "saved_path": str(target_path),
            "download_source": download_source,
        }

    def _resolve_pending_raw_downloads(
        self,
        pending_items: list[dict[str, Any]],
        download_dir: Path,
    ) -> list[dict[str, Any]]:
        if not pending_items:
            return []
        max_workers = min(
            len(pending_items),
            env_int("EXPORT_RAW_DOWNLOAD_CONCURRENCY", 2, minimum=1, maximum=4),
        )

        def download_one(item: dict[str, Any]) -> tuple[dict[str, Any], str, bytes, float]:
            start = time.perf_counter()
            suggested_name, body = self._download_export_request_raw(item["_raw_request"])
            return item, suggested_name, body, elapsed_seconds(start)

        resolved: list[dict[str, Any]] = []
        if max_workers <= 1:
            for item in pending_items:
                pending_item, suggested_name, body, raw_seconds = download_one(item)
                metadata = self._finalize_raw_download_metadata(
                    pending_item=pending_item,
                    suggested_name=suggested_name,
                    body=body,
                    download_dir=download_dir,
                    download_source="raw_request",
                )
                metadata["raw_download_seconds"] = raw_seconds
                metadata["capture_seconds"] = pending_item.get("capture_seconds")
                resolved.append(metadata)
            return resolved

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(download_one, item) for item in pending_items]
            for future in concurrent.futures.as_completed(futures):
                pending_item, suggested_name, body, raw_seconds = future.result()
                metadata = self._finalize_raw_download_metadata(
                    pending_item=pending_item,
                    suggested_name=suggested_name,
                    body=body,
                    download_dir=download_dir,
                    download_source="raw_request_parallel",
                )
                metadata["raw_download_seconds"] = raw_seconds
                metadata["capture_seconds"] = pending_item.get("capture_seconds")
                resolved.append(metadata)
        return sorted(resolved, key=lambda item: int(item["sequence"]))

    def download_for_fba(self, fba_code: str, download_dir: Path, screenshot_dir: Path) -> dict[str, Any]:
        api_fast_path_error: str | None = None
        api_fast_path_seconds: float | None = None
        if env_flag("LINGXING_API_FAST_PATH", True):
            api_fast_path_start = time.perf_counter()
            try:
                return self._download_for_fba_via_api(fba_code, download_dir, screenshot_dir)
            except Exception as exc:
                api_fast_path_seconds = elapsed_seconds(api_fast_path_start)
                api_fast_path_error = f"{type(exc).__name__}: {exc}"

        overall_start = time.perf_counter()
        timings: dict[str, float] = {}
        self.current_screenshot_dir = screenshot_dir
        download_dir.mkdir(parents=True, exist_ok=True)

        stage_start = time.perf_counter()
        self.search_shipment(fba_code)
        timings["search_seconds"] = elapsed_seconds(stage_start)

        stage_start = time.perf_counter()
        self.open_shipment_detail(fba_code)
        timings["open_detail_seconds"] = elapsed_seconds(stage_start)

        stage_start = time.perf_counter()
        self.ensure_box_labels_ready(fba_code)
        timings["box_labels_ready_seconds"] = elapsed_seconds(stage_start)

        stage_start = time.perf_counter()
        self._reset_shipment_card_scroll()
        if not self._find_shipment_cards():
            raise AutomationError("shipment_cards_not_found", f"FBA {fba_code} 页面中未识别到仓库卡片")
        timings["card_scan_seconds"] = elapsed_seconds(stage_start)

        downloaded_files: list[dict[str, Any]] = []
        pending_raw_downloads: list[dict[str, Any]] = []
        seen_card_signatures: set[str] = set()
        idle_scroll_count = 0
        stage_start = time.perf_counter()

        while idle_scroll_count < 3:
            cards = self._find_shipment_cards()
            downloaded_this_view = 0
            for card in cards:
                try:
                    card_text = normalize_text(card.inner_text(timeout=2000))
                except Exception:
                    continue
                signature = self._shipment_card_signature(card_text, len(seen_card_signatures) + 1)
                if signature in seen_card_signatures:
                    continue
                index = len(downloaded_files) + len(pending_raw_downloads) + 1
                warehouse_code = self._extract_warehouse_code(card_text, index)
                card_download_start = time.perf_counter()
                downloaded_file = self._download_card_packing_list(
                    card=card,
                    fba_code=fba_code,
                    index=index,
                    warehouse_code=warehouse_code,
                    download_dir=download_dir,
                    defer_raw_download=True,
                )
                downloaded_file["capture_seconds"] = elapsed_seconds(card_download_start)
                seen_card_signatures.add(signature)
                if downloaded_file.get("_raw_request") is not None:
                    pending_raw_downloads.append(downloaded_file)
                else:
                    downloaded_file["duration_seconds"] = downloaded_file["capture_seconds"]
                    downloaded_files.append(downloaded_file)
                downloaded_this_view += 1
                self.page.wait_for_timeout(100)

            moved = self._scroll_shipment_cards_right()
            if not moved:
                break
            if downloaded_this_view == 0:
                idle_scroll_count += 1
            else:
                idle_scroll_count = 0

        raw_resolve_start = time.perf_counter()
        resolved_raw_downloads = self._resolve_pending_raw_downloads(pending_raw_downloads, download_dir)
        for item in resolved_raw_downloads:
            capture_seconds = float(item.get("capture_seconds") or 0)
            raw_seconds = float(item.get("raw_download_seconds") or 0)
            item["duration_seconds"] = round(capture_seconds + raw_seconds, 3)
        downloaded_files.extend(resolved_raw_downloads)
        downloaded_files.sort(key=lambda item: int(item["sequence"]))
        timings["download_cards_seconds"] = elapsed_seconds(stage_start)
        timings["raw_download_resolve_seconds"] = elapsed_seconds(raw_resolve_start)
        timings["raw_download_concurrency"] = min(
            len(pending_raw_downloads),
            env_int("EXPORT_RAW_DOWNLOAD_CONCURRENCY", 2, minimum=1, maximum=4),
        ) if pending_raw_downloads else 0
        return {
            "warehouse_count": len(downloaded_files),
            "downloaded_files": downloaded_files,
            "timings": timings,
            "download_strategy": "page_fallback" if api_fast_path_error else "page",
            "api_fast_path_error": api_fast_path_error,
            "api_fast_path_seconds": api_fast_path_seconds,
            "download_duration_seconds": elapsed_seconds(overall_start),
            **self.current_page_state(),
        }


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="领星浏览器自动化下载 + Excel 一体化处理工具")
    parser.add_argument("--manifest", required=True, help="TXT 或 XLSX 的 FBA 清单文件")
    parser.add_argument("--resource-dir", default=str(Path(__file__).resolve().parent))
    parser.add_argument("--work-dir", default=str(Path(__file__).resolve().parent))
    parser.add_argument("--config-file", default=None, help="登录配置 JSON 文件，可选")
    return parser


def build_batch_report(
    batch_dir: Path,
    manifest_path: Path,
    resource_dir: Path,
    work_dir: Path,
    config_path: Path | None,
    fba_codes: list[str],
    results: list[dict[str, Any]],
    fatal_error: dict[str, Any] | None = None,
    report_path: Path | None = None,
) -> dict[str, Any]:
    success_count = sum(1 for item in results if item.get("status") == "success")
    failed_count = sum(1 for item in results if item.get("status") != "success")
    if failed_count == 0 and results:
        batch_status = "success"
    elif success_count == 0:
        batch_status = "failed"
    else:
        batch_status = "partial_success"
    payload = {
        "batch_dir": str(batch_dir),
        "manifest_path": str(manifest_path),
        "resource_dir": str(resource_dir),
        "work_dir": str(work_dir),
        "config_file": str(config_path) if config_path and config_path.exists() else None,
        "started_at": results[0]["started_at"] if results else beijing_now_text(),
        "finished_at": beijing_now_text(),
        "fba_codes": fba_codes,
        "success_count": success_count,
        "failed_count": failed_count,
        "status": batch_status,
        "results": results,
        "fatal_error": fatal_error,
    }
    write_json(report_path or (batch_dir / "batch_report.json"), payload)
    return payload


def emit_log(log_callback: Callable[[str], None] | None, message: str) -> None:
    if log_callback is None:
        return
    log_callback(message)


def run_single_fba(
    automation: Any,
    fba_code: str,
    resource_dir: Path,
    batch_dir: Path,
    download_dir: Path | None = None,
    output_dir: Path | None = None,
    screenshot_dir: Path | None = None,
    report_path: Path | None = None,
    log_callback: Callable[[str], None] | None = None,
) -> dict[str, Any]:
    fba_start = time.perf_counter()
    fba_root = batch_dir / sanitize_filename_part(fba_code)
    download_dir = download_dir or (fba_root / "downloads")
    output_dir = output_dir or (fba_root / "output")
    screenshot_dir = screenshot_dir or (fba_root / "screenshots")
    for path in [download_dir, output_dir, screenshot_dir]:
        path.mkdir(parents=True, exist_ok=True)

    report: dict[str, Any] = {
        "fba_code": fba_code,
        "status": "pending",
        "started_at": beijing_now_text(),
        "fba_root": str(fba_root),
        "downloads_dir": str(download_dir),
        "output_dir": str(output_dir),
        "screenshots_dir": str(screenshot_dir),
        "downloaded_files": [],
        "processing_output_workbook": None,
        "processing_output_files": [],
        "processing_report_file": None,
        "error_code": None,
        "error": None,
        "traceback": None,
        "failure_screenshot": None,
    }

    try:
        emit_log(log_callback, f"[{fba_code}] 开始浏览器自动化下载")
        download_stage_start = time.perf_counter()
        download_info = automation.download_for_fba(fba_code, download_dir, screenshot_dir)
        report.update(download_info)
        report["downloaded_files"] = download_info.get("downloaded_files", [])
        report["browser_download_seconds"] = elapsed_seconds(download_stage_start)
        emit_log(log_callback, f"[{fba_code}] 下载完成，耗时 {report['browser_download_seconds']} 秒，开始整理 Excel")

        process_stage_start = time.perf_counter()
        process_report = process_workbooks(resource_dir, download_dir, output_dir)
        report["excel_process_seconds"] = elapsed_seconds(process_stage_start)
        report["processing_output_workbook"] = process_report.get("output_workbook")
        report["processing_output_files"] = process_report.get("processing_output_files", [])
        report["processing_report_file"] = process_report.get("report_file")
        report["processing_anomalies"] = process_report.get("anomalies", [])
        report["status"] = "success"
        emit_log(log_callback, f"[{fba_code}] 整理完成，耗时 {report['excel_process_seconds']} 秒")
    except Exception as exc:
        report["status"] = "failed"
        report["error_code"] = exc.code if isinstance(exc, AutomationError) else "unexpected_error"
        report["error"] = str(exc)
        report["traceback"] = traceback.format_exc()
        screenshot_path = automation.capture_screenshot(f"{fba_code}_{report['error_code']}")
        report["failure_screenshot"] = str(screenshot_path) if screenshot_path else None
        report.update(automation.current_page_state())
        emit_log(log_callback, f"[{fba_code}] 执行失败：{report['error']}")
    finally:
        report["duration_seconds"] = elapsed_seconds(fba_start)
        report["finished_at"] = beijing_now_text()
        final_report_path = report_path or (fba_root / "automation_report.json")
        final_report_path.parent.mkdir(parents=True, exist_ok=True)
        write_json(final_report_path, report)

    return report


def run_manifest_job(
    manifest_path: Path,
    resource_dir: Path,
    job_dir: Path,
    profile_dir: Path,
    config_path: Path | None = None,
    log_callback: Callable[[str], None] | None = None,
) -> dict[str, Any]:
    manifest_path = manifest_path.resolve()
    resource_dir = resource_dir.resolve()
    job_dir = job_dir.resolve()
    profile_dir = profile_dir.resolve()
    config_path = config_path.resolve() if config_path else (resource_dir / DEFAULT_CONFIG_FILE_NAME)

    if not manifest_path.exists():
        raise FileNotFoundError(f"manifest 文件不存在：{manifest_path}")
    if not resource_dir.exists():
        raise FileNotFoundError(f"resource-dir 不存在：{resource_dir}")

    job_dir.mkdir(parents=True, exist_ok=True)
    downloads_root = job_dir / "downloads"
    output_root = job_dir / "output"
    screenshots_root = job_dir / "screenshots"
    reports_root = job_dir / "reports"
    for path in [downloads_root, output_root, screenshots_root, reports_root]:
        path.mkdir(parents=True, exist_ok=True)

    fba_codes = parse_manifest_file(manifest_path)
    if not fba_codes:
        raise ValueError(f"manifest 中未解析到任何 FBA 号：{manifest_path.name}")

    emit_log(log_callback, f"解析到 {len(fba_codes)} 个 FBA，开始初始化浏览器")
    credentials = load_login_credentials(config_path)
    results: list[dict[str, Any]] = []
    automation = None
    fatal_error: dict[str, Any] | None = None
    fatal_exception: Exception | None = None
    try:
        automation = LingxingPlaywrightAutomation(
            profile_dir=profile_dir,
            credentials=credentials,
        )
        automation.start()
        emit_log(log_callback, "浏览器初始化完成")
        for fba_code in fba_codes:
            results.append(
                run_single_fba(
                    automation=automation,
                    fba_code=fba_code,
                    resource_dir=resource_dir,
                    batch_dir=job_dir,
                    download_dir=downloads_root / sanitize_filename_part(fba_code),
                    output_dir=output_root / sanitize_filename_part(fba_code),
                    screenshot_dir=screenshots_root / sanitize_filename_part(fba_code),
                    report_path=reports_root / f"{sanitize_filename_part(fba_code)}_automation_report.json",
                    log_callback=log_callback,
                )
            )
    except Exception as exc:
        fatal_error = {
            "error": str(exc),
            "traceback": traceback.format_exc(),
        }
        fatal_exception = exc
        emit_log(log_callback, f"批次执行中断：{exc}")
    finally:
        if automation is not None:
            automation.close()
            emit_log(log_callback, "浏览器已关闭")
        batch_report = build_batch_report(
            batch_dir=job_dir,
            manifest_path=manifest_path,
            resource_dir=resource_dir,
            work_dir=job_dir.parent,
            config_path=config_path if config_path.exists() else None,
            fba_codes=fba_codes,
            results=results,
            fatal_error=fatal_error,
            report_path=reports_root / "batch_report.json",
        )
        emit_log(log_callback, f"批次完成，状态：{batch_report['status']}")

    if fatal_exception is not None:
        batch_report["fatal_exception_message"] = str(fatal_exception)
    return batch_report


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()

    manifest_path = Path(args.manifest).resolve()
    resource_dir = Path(args.resource_dir).resolve()
    work_dir = Path(args.work_dir).resolve()
    config_path = Path(args.config_file).resolve() if args.config_file else None
    work_dir.mkdir(parents=True, exist_ok=True)
    batch_dir = work_dir / "runs" / build_timestamp()
    batch_dir.mkdir(parents=True, exist_ok=True)
    batch_report = run_manifest_job(
        manifest_path=manifest_path,
        resource_dir=resource_dir,
        job_dir=batch_dir,
        profile_dir=work_dir / "chrome_profile_playwright",
        config_path=config_path,
    )
    print(json.dumps(batch_report, ensure_ascii=False, indent=2, default=json_default))
    if batch_report.get("fatal_error"):
        raise RuntimeError(batch_report["fatal_error"]["error"])


if __name__ == "__main__":
    main()
